from dotenv import load_dotenv
load_dotenv()
from flask import Flask, render_template, request, redirect, url_for, jsonify, session, Response
import hmac
import hashlib
import json
import re
import base64
import requests
from urllib.parse import urlparse
from datetime import datetime, timezone, timedelta
from psycopg import connect
from psycopg.rows import dict_row
import os
import mimetypes
import secrets
import string
import tempfile
import csv
import logging
import time
from io import BytesIO, TextIOWrapper, StringIO
from functools import wraps
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook


def get_positive_int_env(name: str, default: int) -> int:
    raw_value = os.environ.get(name, "").strip()
    if not raw_value:
        return default
    try:
        value = int(raw_value)
    except ValueError:
        return default
    return value if value > 0 else default


APP_ENV = os.environ.get("APP_ENV", "development").strip().lower()
IS_PROD = APP_ENV == "production"
BOOTSTRAP_WARNINGS = []


def emit_bootstrap_warning(message: str) -> None:
    BOOTSTRAP_WARNINGS.append(message)
    print(f"[config warning] {message}")


def require_env(name: str, default_dev: str | None = None, allow_empty_in_dev: bool = False) -> str:
    value = os.environ.get(name, "").strip()
    if value:
        return value
    if not IS_PROD:
        if default_dev is not None:
            emit_bootstrap_warning(
                f"Environment variable {name} belum diisi. Memakai default development sementara."
            )
            return default_dev
        if allow_empty_in_dev:
            emit_bootstrap_warning(
                f"Environment variable {name} belum diisi. Fitur yang butuh konfigurasi ini akan nonaktif."
            )
            return ""
    raise RuntimeError(f"Environment variable {name} wajib diisi.")

app = Flask(__name__)
app.secret_key = require_env("FLASK_SECRET_KEY", default_dev="dev-secret-key-change-me")

logger = logging.getLogger("kc_submit_app")
if not logger.handlers:
    logger.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
    file_handler = logging.FileHandler("submit.log", encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)
for bootstrap_warning in BOOTSTRAP_WARNINGS:
    logger.warning(bootstrap_warning)


APP_HMAC_SECRET = require_env("APP_HMAC_SECRET", default_dev="dev-app-hmac-secret-change-me")
MASTERDATA_HMAC_SECRET = os.environ.get("MASTERDATA_HMAC_SECRET", APP_HMAC_SECRET).strip()
DEFAULT_BASE_URL = os.environ.get(
    "DEFAULT_BASE_URL",
    "https://ca-msfsax05-22-be-letscml-prd.mangosmoke-bb4ae1b7.southeastasia.azurecontainerapps.io",
).strip()
DEFAULT_ENDPOINT = os.environ.get(
    "DEFAULT_ENDPOINT",
    "/api/survey-questionnaire-cmkt-v2s/submit",
).strip()
DEFAULT_BUMO_ENDPOINT = os.environ.get(
    "DEFAULT_BUMO_ENDPOINT",
    "/api/bumos",
).strip()
DEFAULT_KC_AREA_ENDPOINT = os.environ.get(
    "DEFAULT_KC_AREA_ENDPOINT",
    "/api/kc-areas",
).strip()

PRODUCT_PACK_OPTIONS = ["0 pack", "1 pack", "2 pack"]
DEFAULT_SP12_PACK = "1 pack"
NON_PURCHASE_REASON_OPTIONS = [
    "Harga terlalu mahal",
    "Tidak tertarik dengan persepsi rasa produknya",
    "Tidak tertarik dengan kesan mereknya",
    "Pernah mencoba, tapi tidak suka rasa produknya",
    "Pernah mencoba, tapi terlalu dingin/kurang memuaskan",
]
VALID_NON_PURCHASE_REASONS = set(NON_PURCHASE_REASON_OPTIONS)

AGE_RANGE_OPTIONS = [
    ("age-21-25", "21 - 25"),
    ("age-26-30", "26 - 30"),
    ("age-31-35", "31 - 35"),
    ("age-36-40", "36 - 40"),
    ("age-40+", "40+"),
]
VALID_AGE_RANGES = {value for value, _label in AGE_RANGE_OPTIONS}


DATABASE_URL = require_env("DATABASE_URL", allow_empty_in_dev=True)
DEFAULT_DAILY_LIMIT = 40
RESERVED_PHONE_TIMEOUT_MINUTES = get_positive_int_env("RESERVED_PHONE_TIMEOUT_MINUTES", 120)
DUPLICATE_SUBMISSION_WINDOW_MINUTES = get_positive_int_env("DUPLICATE_SUBMISSION_WINDOW_MINUTES", 10)
PENDING_DUPLICATE_BLOCK_SECONDS = min(get_positive_int_env("PENDING_DUPLICATE_BLOCK_SECONDS", 90), 90)
MASTERDATA_API_TIMEOUT_SECONDS = get_positive_int_env("MASTERDATA_API_TIMEOUT_SECONDS", 300)
SUBMISSION_LOG_LIMIT_OPTIONS = ["25", "50", "100", "200", "500", "1000", "10000", "all"]
MAX_SUBMISSION_LOG_LIMIT = 10000

ADMIN_PAGE_USERNAME = require_env("ADMIN_PAGE_USERNAME", default_dev="admin")
ADMIN_PAGE_PASSWORD = require_env("ADMIN_PAGE_PASSWORD", default_dev="admin123")


def get_db_connection():
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL belum di-set. Tambahkan ke environment atau file .env untuk mengaktifkan fitur database.")
    conn = connect(DATABASE_URL, row_factory=dict_row, autocommit=False)
    return conn


def normalize_submission_log_limit(value, default=100):
    if value is None:
        return None
    current = str(value).strip().lower()
    if current in {"all", "semua"}:
        return None
    try:
        limit = int(current or default)
    except (TypeError, ValueError):
        limit = default
    return max(1, min(limit, MAX_SUBMISSION_LOG_LIMIT))


def normalize_submission_date_filter(value):
    current = str(value or "").strip()
    if not current:
        return ""
    try:
        return datetime.strptime(current, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        return ""


def normalize_has_purchased_value(value):
    current = str(value or "").strip().lower()
    if current not in {"true", "false"}:
        raise ValueError("Jawaban pembelian tidak valid.")
    return current


def init_db():
    if not DATABASE_URL:
        logger.warning("init_db dilewati karena DATABASE_URL belum tersedia.")
        return
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS kc_token_usage (
            kc_token TEXT NOT NULL,
            usage_date TEXT NOT NULL,
            total_submit INTEGER NOT NULL DEFAULT 0,
            PRIMARY KEY (kc_token, usage_date)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS valid_kc_tokens (
            kc_token TEXT PRIMARY KEY,
            kc_name TEXT NOT NULL,
            team TEXT NOT NULL DEFAULT '',
            token_area TEXT NOT NULL DEFAULT '',
            kc_username TEXT NOT NULL DEFAULT '',
            kc_password TEXT NOT NULL DEFAULT '',
            bearer_token TEXT NOT NULL,
            daily_limit INTEGER NOT NULL DEFAULT 40,
            is_active INTEGER NOT NULL DEFAULT 1,
            last_bearer_updated_at TEXT NOT NULL DEFAULT ''
        )
    """)

    cur.execute("""
        ALTER TABLE valid_kc_tokens
        ADD COLUMN IF NOT EXISTS token_area TEXT NOT NULL DEFAULT ''
    """)

    cur.execute("""
        ALTER TABLE valid_kc_tokens
        ADD COLUMN IF NOT EXISTS team TEXT NOT NULL DEFAULT ''
    """)

    cur.execute("""
        ALTER TABLE valid_kc_tokens
        ADD COLUMN IF NOT EXISTS kc_username TEXT NOT NULL DEFAULT ''
    """)

    cur.execute("""
        ALTER TABLE valid_kc_tokens
        ADD COLUMN IF NOT EXISTS kc_password TEXT NOT NULL DEFAULT ''
    """)

    cur.execute("""
        ALTER TABLE valid_kc_tokens
        ADD COLUMN IF NOT EXISTS created_date TEXT NOT NULL DEFAULT ''
    """)

    cur.execute("""
        ALTER TABLE valid_kc_tokens
        ADD COLUMN IF NOT EXISTS last_bearer_updated_at TEXT NOT NULL DEFAULT ''
    """)

    cur.execute("""
        SELECT LOWER(BTRIM(kc_username)) AS normalized_username, COUNT(*) AS total
        FROM valid_kc_tokens
        WHERE BTRIM(COALESCE(kc_username, '')) != ''
        GROUP BY LOWER(BTRIM(kc_username))
        HAVING COUNT(*) > 1
        LIMIT 1
    """)
    duplicate_username = cur.fetchone()
    if duplicate_username:
        emit_bootstrap_warning(
            "Ada kc_username duplikat di valid_kc_tokens. Unique index username dilewati sampai data duplikat dibersihkan."
        )
    else:
        cur.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS idx_valid_kc_tokens_unique_username
            ON valid_kc_tokens (LOWER(BTRIM(kc_username)))
            WHERE BTRIM(COALESCE(kc_username, '')) != ''
        """)

    cur.execute("""
        SELECT 1
        FROM information_schema.columns
        WHERE table_name = 'valid_kc_tokens'
          AND column_name = 'kc_area'
        LIMIT 1
    """)
    if cur.fetchone():
        cur.execute("""
            UPDATE valid_kc_tokens
            SET token_area = kc_area
            WHERE COALESCE(token_area, '') = ''
              AND COALESCE(kc_area, '') != ''
        """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS customer_directory (
            phone_number TEXT PRIMARY KEY,
            is_active INTEGER NOT NULL DEFAULT 1,
            is_used INTEGER NOT NULL DEFAULT 0,
            reserved_by_token TEXT,
            reserved_at TEXT,
            shuffle_order BIGINT,
            created_at TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL DEFAULT ''
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS submission_attempts (
            submission_id TEXT PRIMARY KEY,
            phone_number TEXT NOT NULL,
            kc_token TEXT NOT NULL,
            status_local TEXT NOT NULL DEFAULT 'PENDING',
            final_status_code INTEGER,
            final_response_text TEXT,
            attempts_json TEXT NOT NULL DEFAULT '[]',
            request_summary_json TEXT NOT NULL DEFAULT '{}',
            customer_name TEXT NOT NULL DEFAULT '',
            has_purchased TEXT NOT NULL DEFAULT '',
            lighter TEXT NOT NULL DEFAULT '',
            lighter_purchased INTEGER NOT NULL DEFAULT 0,
            kc_area TEXT NOT NULL DEFAULT '',
            kc_area_label TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
    """)

    for column_name, column_type in {
        "customer_name": "TEXT NOT NULL DEFAULT ''",
        "has_purchased": "TEXT NOT NULL DEFAULT ''",
        "lighter": "TEXT NOT NULL DEFAULT ''",
        "lighter_purchased": "INTEGER NOT NULL DEFAULT 0",
        "kc_area": "TEXT NOT NULL DEFAULT ''",
        "kc_area_label": "TEXT NOT NULL DEFAULT ''",
    }.items():
        cur.execute(f"""
            ALTER TABLE submission_attempts
            ADD COLUMN IF NOT EXISTS {column_name} {column_type}
        """)

    cur.execute("""
        UPDATE submission_attempts
        SET customer_name = COALESCE(NULLIF(customer_name, ''), COALESCE(request_summary_json::jsonb ->> 'customer_name', '')),
            has_purchased = COALESCE(NULLIF(has_purchased, ''), COALESCE(request_summary_json::jsonb ->> 'has_purchased', '')),
            lighter = COALESCE(NULLIF(lighter, ''), COALESCE(request_summary_json::jsonb ->> 'lighter', '')),
            kc_area = COALESCE(NULLIF(kc_area, ''), COALESCE(request_summary_json::jsonb ->> 'kc_area', '')),
            kc_area_label = COALESCE(NULLIF(kc_area_label, ''), COALESCE(request_summary_json::jsonb ->> 'kc_area_label', '')),
            lighter_purchased = CASE
                WHEN lighter_purchased = 1 THEN 1
                WHEN COALESCE(request_summary_json::jsonb ->> 'has_purchased', '') = 'true'
                 AND COALESCE(request_summary_json::jsonb ->> 'lighter', '') = 'Ya'
                THEN 1 ELSE 0
            END
        WHERE request_summary_json IS NOT NULL
          AND request_summary_json != ''
          AND request_summary_json != '{}'
          AND (
            customer_name = ''
            OR has_purchased = ''
            OR kc_area = ''
            OR kc_area_label = ''
          )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS bumo_master (
            name TEXT PRIMARY KEY,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL DEFAULT ''
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS kc_area_master (
            area_id TEXT PRIMARY KEY,
            area_name TEXT NOT NULL,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL DEFAULT ''
        )
    """)

    cur.execute("""
        UPDATE customer_directory
        SET created_at = COALESCE(NULLIF(created_at, ''), NOW()::text),
            updated_at = COALESCE(NULLIF(updated_at, ''), NOW()::text)
    """)

    cur.execute("SELECT phone_number FROM customer_directory WHERE shuffle_order IS NULL")
    rows_to_backfill = cur.fetchall()
    for row in rows_to_backfill:
        cur.execute(
            "UPDATE customer_directory SET shuffle_order = %s WHERE phone_number = %s",
            (secrets.randbits(63), row["phone_number"]),
        )

    cur.execute("CREATE INDEX IF NOT EXISTS idx_customer_directory_pick ON customer_directory (is_active, is_used, reserved_by_token, shuffle_order)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_submission_attempts_phone ON submission_attempts (phone_number, created_at DESC)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_submission_attempts_status ON submission_attempts (status_local, created_at DESC)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_submission_attempts_kc ON submission_attempts (kc_token, created_at DESC)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_submission_attempts_created_at ON submission_attempts (created_at DESC)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_submission_attempts_purchase_counts ON submission_attempts (status_local, created_at DESC, kc_token)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_bumo_master_active_name ON bumo_master (is_active, name)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_kc_area_master_active_name ON kc_area_master (is_active, area_name)")

    cur.execute("""
        CREATE TABLE IF NOT EXISTS team_leaders (
            username TEXT PRIMARY KEY,
            password_hash TEXT NOT NULL,
            leader_name TEXT NOT NULL DEFAULT '',
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL DEFAULT ''
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS team_leader_kc_access (
            id BIGSERIAL PRIMARY KEY,
            leader_username TEXT NOT NULL,
            kc_token TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT '',
            UNIQUE (leader_username, kc_token)
        )
    """)

    cur.execute("CREATE INDEX IF NOT EXISTS idx_team_leader_access_username ON team_leader_kc_access (leader_username)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_team_leader_access_kc_token ON team_leader_kc_access (kc_token)")

    conn.commit()
    conn.close()


def seed_kc_tokens():
    sample_tokens = [
        ("KC-JKT-001", "KC Jakarta", "Jakarta", "isi_bearer_jakarta", 40, 1),
        ("KC-BDG-001", "KC Bandung", "Bandung", "isi_bearer_bandung", 40, 1),
        ("KC-SBY-001", "KC Surabaya", "Surabaya", "isi_bearer_surabaya", 40, 1),
    ]

    conn = get_db_connection()
    cur = conn.cursor()

    for kc_token, kc_name, token_area, bearer_token, daily_limit, is_active in sample_tokens:
        cur.execute("""
            INSERT INTO valid_kc_tokens (kc_token, kc_name, token_area, bearer_token, daily_limit, is_active)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (kc_token) DO NOTHING
        """, (kc_token, kc_name, token_area, bearer_token, daily_limit, is_active))

    conn.commit()
    conn.close()


def get_now_wib():
    return datetime.now(timezone.utc) + timedelta(hours=7)


def get_today_wib():
    return get_now_wib().date().isoformat()


def get_now_wib_string():
    return get_now_wib().strftime("%Y-%m-%d %H:%M:%S")


def is_token_expired():
    token_login_date = session.get("token_login_date")
    return token_login_date != get_today_wib()


def clear_user_session(release_phone=True):
    if release_phone:
        release_current_reserved_phone()
    session.pop("bearer_token", None)
    session.pop("kc_token", None)
    session.pop("token_login_date", None)
    session.pop("kc_name", None)
    session.pop("daily_limit", None)
    session.pop("assigned_phone_number", None)


def clear_expired_user_session():
    if is_token_expired():
        clear_user_session()
        return True
    return False


def is_admin_logged_in():
    return session.get("is_admin_logged_in") is True


def is_team_leader_logged_in():
    return session.get("is_team_leader_logged_in") is True and bool((session.get("team_leader_username") or "").strip())


def admin_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not is_admin_logged_in():
            return redirect(url_for("admin_login"))
        return view_func(*args, **kwargs)
    return wrapper


def leader_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if is_admin_logged_in() or is_team_leader_logged_in():
            return view_func(*args, **kwargs)
        return redirect(url_for("team_leader_login"))
    return wrapper


def wants_json_response():
    accept = (request.headers.get("Accept") or "").lower()
    requested_with = (request.headers.get("X-Requested-With") or "").lower()
    return "application/json" in accept or requested_with == "xmlhttprequest"


def get_now_db_string():
    return get_now_wib().strftime("%Y-%m-%d %H:%M:%S")


def response_body_to_text(response_body):
    if isinstance(response_body, (dict, list)):
        return safe_json_dumps(response_body, ensure_ascii=False)
    return str(response_body or "")


def is_duplicate_response(response_body):
    return "sudah melakukan pengisian form" in response_body_to_text(response_body).lower()


def normalize_final_submit_state(result):
    attempts = result.get("attempts") or []

    if attempts:
        final_attempt = attempts[-1]
        final_status = final_attempt.get("status_code")

        if final_status and 200 <= final_status < 300:
            return "SUCCESS"
        if final_status == 400:
            return "LIKELY_SUCCESS"
        if should_mark_phone_invalid(result):
            return "INVALID"
        return "FAILED"

    status_code = result.get("status_code")
    if status_code and 200 <= status_code < 300:
        return "SUCCESS"
    if status_code == 400:
        return "LIKELY_SUCCESS"
    return "FAILED"


def should_mark_phone_invalid(result):
    return False


def is_all_attempts_unauthorized(result):
    attempts = result.get("attempts") or []
    if attempts:
        return all(attempt.get("status_code") == 401 for attempt in attempts)
    return result.get("status_code") == 401


def reserve_next_phone_for_session(kc_token, previous_phone_number=None):
    session.pop("assigned_phone_number", None)
    next_phone_number = reserve_phone_for_kc(kc_token)
    if previous_phone_number and next_phone_number == previous_phone_number:
        logger.warning(
            "reserve next phone returned previous phone kc_token=%s previous_phone=%s",
            kc_token,
            previous_phone_number,
        )
        mark_phone_as_used(previous_phone_number, kc_token)
        session.pop("assigned_phone_number", None)
        next_phone_number = reserve_phone_for_kc(kc_token)
        if next_phone_number == previous_phone_number:
            session.pop("assigned_phone_number", None)
            next_phone_number = ""
    if next_phone_number:
        session["assigned_phone_number"] = next_phone_number
    return next_phone_number or ""


def summarize_submit_result(result):
    attempts = result.get("attempts") or []
    if not attempts:
        status = result.get("status_code")
        return f"1x:{status}"
    return " | ".join(
        f"{attempt.get('attempt_no', idx + 1)}x:{attempt.get('status_code')}"
        for idx, attempt in enumerate(attempts)
    )


def request_summary_has_lighter_purchase(request_summary):
    summary = request_summary or {}
    if str(summary.get("has_purchased") or "").strip().lower() != "true":
        return 0
    if str(summary.get("lighter") or "").strip() == "Ya":
        return 1

    product_transactions = summary.get("product_transactions")
    if isinstance(product_transactions, str):
        try:
            product_transactions = json.loads(product_transactions or "[]")
        except Exception:
            product_transactions = []
    if not isinstance(product_transactions, list):
        return 0

    for item in product_transactions:
        if not isinstance(item, dict):
            continue
        if str(item.get("product_name") or "").strip() != "Lighter":
            continue
        try:
            quantity = int(item.get("quantity") or 0)
        except (TypeError, ValueError):
            quantity = 0
        if quantity > 0:
            return 1
    return 0


def extract_submission_summary_columns(request_summary):
    summary = request_summary or {}
    return {
        "customer_name": str(summary.get("customer_name") or "").strip(),
        "has_purchased": str(summary.get("has_purchased") or "").strip().lower(),
        "lighter": str(summary.get("lighter") or "").strip(),
        "lighter_purchased": request_summary_has_lighter_purchase(summary),
        "kc_area": str(summary.get("kc_area") or "").strip(),
        "kc_area_label": str(summary.get("kc_area_label") or "").strip(),
    }


def create_submission_attempt(submission_id, phone_number, kc_token, request_summary):
    now_str = get_now_db_string()
    summary_columns = extract_submission_summary_columns(request_summary)
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO submission_attempts (
            submission_id, phone_number, kc_token, status_local, final_status_code,
            final_response_text, attempts_json, request_summary_json, customer_name,
            has_purchased, lighter, lighter_purchased, kc_area, kc_area_label,
            created_at, updated_at
        ) VALUES (%s, %s, %s, 'PENDING', NULL, NULL, '[]', %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """,
        (
            submission_id,
            phone_number,
            kc_token,
            safe_json_dumps(request_summary, ensure_ascii=False),
            summary_columns["customer_name"],
            summary_columns["has_purchased"],
            summary_columns["lighter"],
            summary_columns["lighter_purchased"],
            summary_columns["kc_area"],
            summary_columns["kc_area_label"],
            now_str,
            now_str,
        ),
    )
    conn.commit()
    conn.close()


def update_submission_request_summary(submission_id, phone_number, kc_token, request_summary):
    now_str = get_now_db_string()
    summary_columns = extract_submission_summary_columns(request_summary)
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE submission_attempts
        SET phone_number = %s,
            kc_token = %s,
            request_summary_json = %s,
            customer_name = %s,
            has_purchased = %s,
            lighter = %s,
            lighter_purchased = %s,
            kc_area = %s,
            kc_area_label = %s,
            updated_at = %s
        WHERE submission_id = %s
        """,
        (
            phone_number,
            kc_token,
            safe_json_dumps(request_summary, ensure_ascii=False),
            summary_columns["customer_name"],
            summary_columns["has_purchased"],
            summary_columns["lighter"],
            summary_columns["lighter_purchased"],
            summary_columns["kc_area"],
            summary_columns["kc_area_label"],
            now_str,
            submission_id,
        ),
    )
    conn.commit()
    conn.close()


def update_submission_attempt(submission_id, status_local, final_status_code, final_response_text, attempts):
    now_str = get_now_db_string()
    response_text = safe_json_dumps(final_response_text, ensure_ascii=False) if isinstance(final_response_text, (dict, list)) else str(final_response_text or "")
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE submission_attempts
        SET status_local = %s,
            final_status_code = %s,
            final_response_text = %s,
            attempts_json = %s,
            updated_at = %s
        WHERE submission_id = %s
        """,
        (status_local, final_status_code, response_text, safe_json_dumps(attempts, ensure_ascii=False), now_str, submission_id),
    )
    conn.commit()
    conn.close()


def summarize_response_text(response_text, max_length=220):
    text = str(response_text or "").strip()
    if not text:
        return "-"
    if len(text) <= max_length:
        return text
    return text[:max_length].rstrip() + "..."


def get_recent_submission_attempts(limit=100, status_filter="", kc_token_filter="", phone_filter="", date_from="", date_to="", include_response_text=True):
    limit = normalize_submission_log_limit(limit, default=100)
    date_from = normalize_submission_date_filter(date_from)
    date_to = normalize_submission_date_filter(date_to)
    response_select = (
        "s.final_response_text"
        if include_response_text
        else "LEFT(COALESCE(s.final_response_text, ''), 240) AS final_response_text, LENGTH(COALESCE(s.final_response_text, '')) AS final_response_length"
    )
    query = [
        f"SELECT s.submission_id, s.phone_number, s.kc_token, COALESCE(v.kc_name, '') AS kc_name, s.status_local, s.final_status_code, {response_select}, s.attempts_json, s.request_summary_json, s.created_at, s.updated_at",
        "FROM submission_attempts s",
        "LEFT JOIN valid_kc_tokens v ON s.kc_token = v.kc_token",
        "WHERE 1=1",
    ]
    params = []

    if status_filter:
        query.append("AND s.status_local = %s")
        params.append(status_filter)
    if kc_token_filter:
        query.append("AND s.kc_token LIKE %s")
        params.append(f"%{kc_token_filter}%")
    if phone_filter:
        query.append("AND s.phone_number LIKE %s")
        params.append(f"%{phone_filter}%")
    if date_from:
        query.append("AND s.created_at >= %s")
        params.append(f"{date_from} 00:00:00")
    if date_to:
        query.append("AND s.created_at <= %s")
        params.append(f"{date_to} 23:59:59")

    query.append("ORDER BY s.created_at DESC")
    if limit is not None:
        query.append("LIMIT %s")
        params.append(limit)

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("\n".join(query), params)
    rows = cur.fetchall()
    conn.close()

    results = []
    for row in rows:
        try:
            attempts = json.loads(row["attempts_json"] or "[]")
        except Exception:
            attempts = []
        try:
            request_summary = json.loads(row["request_summary_json"] or "{}")
        except Exception:
            request_summary = {}
        response_text = row["final_response_text"] or ""
        response_length = int(row.get("final_response_length") or len(response_text) or 0)
        results.append({
            "submission_id": row["submission_id"],
            "phone_number": row["phone_number"],
            "kc_token": row["kc_token"],
            "kc_name": row["kc_name"] or request_summary.get("kc_name") or "-",
            "status_local": row["status_local"],
            "final_status_code": row["final_status_code"],
            "final_response_text": response_text if include_response_text else "",
            "final_response_preview": summarize_response_text(response_text),
            "final_response_length": response_length,
            "has_full_response": response_length > 240 if not include_response_text else bool(response_text),
            "attempts": attempts,
            "attempt_count": len(attempts),
            "attempt_summary": " | ".join([f"{a.get('attempt_no', i+1)}x:{a.get('status_code')}" for i, a in enumerate(attempts)]) if attempts else "-",
            "request_summary": request_summary,
            "created_at": row["created_at"],
            "updated_at": row["updated_at"],
        })
    return results


def get_submission_status_counts(kc_token_filter="", phone_filter="", date_from="", date_to=""):
    kc_token_filter = str(kc_token_filter or "").strip()
    phone_filter = str(phone_filter or "").strip()
    date_from = normalize_submission_date_filter(date_from)
    date_to = normalize_submission_date_filter(date_to)
    query = [
        "SELECT status_local, COUNT(*) AS total",
        "FROM submission_attempts",
        "WHERE 1=1",
    ]
    params = []

    if kc_token_filter:
        query.append("AND kc_token LIKE %s")
        params.append(f"%{kc_token_filter}%")
    if phone_filter:
        query.append("AND phone_number LIKE %s")
        params.append(f"%{phone_filter}%")
    if date_from:
        query.append("AND created_at >= %s")
        params.append(f"{date_from} 00:00:00")
    if date_to:
        query.append("AND created_at <= %s")
        params.append(f"{date_to} 23:59:59")
    query.append("GROUP BY status_local")

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("\n".join(query), params)
    rows = cur.fetchall()
    conn.close()
    counts = {"SUCCESS": 0, "LIKELY_SUCCESS": 0, "INVALID": 0, "FAILED": 0, "PENDING": 0}
    for row in rows:
        counts[row["status_local"]] = row["total"]
    return counts


def get_submission_response_detail(submission_id):
    current_id = str(submission_id or "").strip()
    if not current_id:
        return None
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT submission_id, final_response_text
        FROM submission_attempts
        WHERE submission_id = %s
        LIMIT 1
        """,
        (current_id,),
    )
    row = cur.fetchone()
    conn.close()
    return row


def delete_submission_attempts_by_filter(status_filter="", kc_token_filter="", phone_filter="", date_from="", date_to=""):
    status_filter = str(status_filter or "").strip()
    kc_token_filter = str(kc_token_filter or "").strip()
    phone_filter = str(phone_filter or "").strip()
    date_from = normalize_submission_date_filter(date_from)
    date_to = normalize_submission_date_filter(date_to)

    query = ["DELETE FROM submission_attempts WHERE 1=1"]
    params = []
    if status_filter:
        query.append("AND status_local = %s")
        params.append(status_filter)
    if kc_token_filter:
        query.append("AND kc_token LIKE %s")
        params.append(f"%{kc_token_filter}%")
    if phone_filter:
        query.append("AND phone_number LIKE %s")
        params.append(f"%{phone_filter}%")
    if date_from:
        query.append("AND created_at >= %s")
        params.append(f"{date_from} 00:00:00")
    if date_to:
        query.append("AND created_at <= %s")
        params.append(f"{date_to} 23:59:59")

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("\n".join(query), params)
    deleted_count = cur.rowcount
    conn.commit()
    conn.close()
    return deleted_count


def get_kc_purchase_counts(date_from="", date_to=""):
    date_from = normalize_submission_date_filter(date_from) or get_today_wib()
    date_to = normalize_submission_date_filter(date_to) or date_from

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT
            s.kc_token,
            SUM(CASE WHEN s.has_purchased = 'true' THEN 1 ELSE 0 END) AS purchase_yes,
            SUM(CASE WHEN s.has_purchased = 'false' THEN 1 ELSE 0 END) AS purchase_no,
            SUM(CASE WHEN s.has_purchased = 'true' AND s.lighter_purchased = 1 THEN 1 ELSE 0 END) AS lighter_yes,
            SUM(CASE WHEN s.has_purchased = 'true' AND s.lighter_purchased != 1 THEN 1 ELSE 0 END) AS lighter_no
        FROM submission_attempts s
        WHERE s.created_at >= %s
          AND s.created_at <= %s
          AND s.status_local IN ('SUCCESS', 'LIKELY_SUCCESS')
        GROUP BY s.kc_token
        """,
        (f"{date_from} 00:00:00", f"{date_to} 23:59:59"),
    )
    rows = cur.fetchall()
    conn.close()
    return {
        row["kc_token"]: {
            "purchase_yes": int(row["purchase_yes"] or 0),
            "purchase_no": int(row["purchase_no"] or 0),
            "lighter_yes": int(row["lighter_yes"] or 0),
            "lighter_no": int(row["lighter_no"] or 0),
        }
        for row in rows
    }


def get_kc_token_card_stats(kc_token, daily_limit):
    used_today, remaining_today, quota_date = get_remaining_quota(kc_token, daily_limit)
    purchase_counts = get_kc_purchase_counts(quota_date, quota_date).get(kc_token, {})
    return {
        "used_today": int(used_today or 0),
        "remaining_today": int(remaining_today or 0),
        "quota_date": quota_date,
        "purchase_yes": int(purchase_counts.get("purchase_yes", 0) or 0),
        "purchase_no": int(purchase_counts.get("purchase_no", 0) or 0),
        "lighter_yes": int(purchase_counts.get("lighter_yes", 0) or 0),
    }


def build_submission_attempts_export_csv(limit=10000, status_filter="", kc_token_filter="", phone_filter="", date_from="", date_to=""):
    rows = get_recent_submission_attempts(
        limit=limit,
        status_filter=status_filter,
        kc_token_filter=kc_token_filter,
        phone_filter=phone_filter,
        date_from=date_from,
        date_to=date_to,
    )
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "created_at",
        "updated_at",
        "status",
        "http_status",
        "attempt_count",
        "attempt_summary",
        "phone_number",
        "kc_token",
        "kc_name",
        "stage",
        "customer_name",
        "age_range",
        "current_bumo",
        "kc_area",
        "kc_area_name",
        "has_purchased",
        "lighter",
        "non_purchase_reasons",
        "has_transaction_photo",
        "has_chat_photo",
        "product_transactions",
        "username",
        "response",
        "submission_id",
        "attempts_json",
        "request_summary_json",
    ])
    for row in rows:
        request_summary = row.get("request_summary") or {}
        response_username = ""
        final_response_text = row.get("final_response_text") or ""
        if final_response_text:
            try:
                response_payload = json.loads(final_response_text)
                if isinstance(response_payload, dict):
                    response_data = response_payload.get("data") or {}
                    if isinstance(response_data, dict):
                        response_username = response_data.get("username") or ""
            except Exception:
                response_username = ""
        writer.writerow([
            row.get("created_at") or "",
            row.get("updated_at") or "",
            row.get("status_local") or "",
            row.get("final_status_code") or "",
            row.get("attempt_count") or 0,
            row.get("attempt_summary") or "",
            row.get("phone_number") or "",
            row.get("kc_token") or "",
            row.get("kc_name") or "",
            request_summary.get("stage") or "",
            request_summary.get("customer_name") or "",
            request_summary.get("age_range") or "",
            request_summary.get("current_bumo") or "",
            request_summary.get("kc_area") or "",
            request_summary.get("kc_area_label") or request_summary.get("kc_area") or "",
            request_summary.get("has_purchased") or "",
            request_summary.get("lighter") or "",
            request_summary.get("non_purchase_reasons") or "",
            request_summary.get("has_transaction_photo") or "",
            request_summary.get("has_chat_photo") or "",
            request_summary.get("product_transactions") or "",
            response_username,
            row.get("final_response_text") or "",
            row.get("submission_id") or "",
            safe_json_dumps(row.get("attempts") or [], ensure_ascii=False),
            safe_json_dumps(request_summary, ensure_ascii=False),
        ])
    return "\ufeff" + output.getvalue()


def normalize_phone_number(raw_value):
    digits = "".join(ch for ch in str(raw_value or "") if ch.isdigit())
    if not digits:
        return ""
    if digits.startswith("62"):
        digits = "0" + digits[2:]
    elif digits.startswith("8"):
        digits = "0" + digits
    if not digits.startswith("08"):
        return ""
    if len(digits) < 10 or len(digits) > 14:
        return ""
    return digits


def get_reserved_phone_for_kc(kc_token):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT phone_number
        FROM customer_directory
        WHERE reserved_by_token = %s AND is_active = 1 AND is_used = 0
        ORDER BY reserved_at ASC, phone_number ASC
        LIMIT 1
        """,
        (kc_token,),
    )
    row = cur.fetchone()
    conn.close()
    return row["phone_number"] if row else None


def get_reserved_phone_timeout_cutoff():
    return (get_now_wib() - timedelta(minutes=RESERVED_PHONE_TIMEOUT_MINUTES)).strftime("%Y-%m-%d %H:%M:%S")


def release_stale_reserved_phones(kc_token=None):
    cutoff_str = get_reserved_phone_timeout_cutoff()
    conn = get_db_connection()
    cur = conn.cursor()
    query = [
        "UPDATE customer_directory",
        "SET reserved_by_token = NULL, reserved_at = NULL, updated_at = %s",
        "WHERE is_used = 0",
        "AND reserved_by_token IS NOT NULL",
        "AND reserved_by_token != ''",
        "AND reserved_at IS NOT NULL",
        "AND reserved_at != ''",
        "AND reserved_at < %s",
    ]
    params = [get_now_db_string(), cutoff_str]
    if kc_token:
        query.append("AND reserved_by_token = %s")
        params.append(kc_token)
    cur.execute(" ".join(query), tuple(params))
    released_count = cur.rowcount
    conn.commit()
    conn.close()
    if released_count:
        logger.info(
            "released stale reserved phones count=%s timeout_minutes=%s kc_token=%s",
            released_count,
            RESERVED_PHONE_TIMEOUT_MINUTES,
            kc_token or "-",
        )
    return released_count


def refresh_reserved_phone(phone_number, kc_token):
    if not phone_number or not kc_token:
        return False

    now_str = get_now_db_string()
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE customer_directory
        SET reserved_at = %s, updated_at = %s
        WHERE phone_number = %s
          AND reserved_by_token = %s
          AND is_active = 1
          AND is_used = 0
        """,
        (now_str, now_str, phone_number, kc_token),
    )
    refreshed = cur.rowcount == 1
    conn.commit()
    conn.close()
    return refreshed


def reserve_phone_for_kc(kc_token):
    release_stale_reserved_phones()
    existing = get_reserved_phone_for_kc(kc_token)
    if existing:
        refresh_reserved_phone(existing, kc_token)
        return existing

    for _ in range(5):
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                """
                SELECT phone_number
                FROM customer_directory
                WHERE is_active = 1
                  AND is_used = 0
                  AND (reserved_by_token IS NULL OR reserved_by_token = '')
                ORDER BY shuffle_order ASC, created_at ASC, phone_number ASC
                FOR UPDATE SKIP LOCKED
                LIMIT 1
                """
            )
            row = cur.fetchone()
            if not row:
                conn.commit()
                return None

            phone_number = row["phone_number"]
            now_str = get_now_db_string()
            cur.execute(
                """
                UPDATE customer_directory
                SET reserved_by_token = %s, reserved_at = %s, updated_at = %s
                WHERE phone_number = %s
                  AND is_active = 1
                  AND is_used = 0
                  AND (reserved_by_token IS NULL OR reserved_by_token = '')
                """,
                (kc_token, now_str, now_str, phone_number),
            )
            if cur.rowcount == 1:
                conn.commit()
                return phone_number
            conn.rollback()
        except Exception:
            conn.rollback()
        finally:
            conn.close()
    return None


def release_reserved_phone(phone_number=None, kc_token=None):
    if not phone_number and not kc_token:
        return

    conn = get_db_connection()
    cur = conn.cursor()
    query = [
        "UPDATE customer_directory",
        "SET reserved_by_token = NULL, reserved_at = NULL, updated_at = %s",
        "WHERE is_used = 0",
    ]
    params = [get_now_db_string()]

    if phone_number:
        query.append("AND phone_number = %s")
        params.append(phone_number)
    if kc_token:
        query.append("AND reserved_by_token = %s")
        params.append(kc_token)

    cur.execute(" ".join(query), tuple(params))
    conn.commit()
    conn.close()


def release_current_reserved_phone():
    assigned_phone = session.get("assigned_phone_number")
    kc_token = session.get("kc_token")
    if assigned_phone and kc_token:
        release_reserved_phone(phone_number=assigned_phone, kc_token=kc_token)


def mark_phone_as_used(phone_number, kc_token):
    if not phone_number:
        return
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE customer_directory
        SET is_used = 1,
            reserved_by_token = NULL,
            reserved_at = NULL,
            updated_at = %s
        WHERE phone_number = %s
          AND (reserved_by_token = %s OR reserved_by_token IS NULL OR reserved_by_token = '')
        """,
        (get_now_db_string(), phone_number, kc_token),
    )
    conn.commit()
    conn.close()


def upsert_customer_number(phone_number, is_active=1, old_phone_number=None):
    normalized = normalize_phone_number(phone_number)
    if not normalized:
        raise ValueError("Nomor HP harus diawali 08 dan panjang 10-14 digit.")

    now_str = get_now_db_string()
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        if old_phone_number:
            old_phone_number = normalize_phone_number(old_phone_number)
            if not old_phone_number:
                raise ValueError("Nomor lama tidak valid.")
            if old_phone_number != normalized:
                cur.execute("SELECT 1 FROM customer_directory WHERE phone_number = %s", (normalized,))
                if cur.fetchone():
                    raise ValueError("Nomor HP sudah ada di database.")
            cur.execute(
                """
                UPDATE customer_directory
                SET phone_number = %s, is_active = %s, updated_at = %s
                WHERE phone_number = %s
                """,
                (normalized, 1 if is_active else 0, now_str, old_phone_number),
            )
        else:
            cur.execute("SELECT 1 FROM customer_directory WHERE phone_number = %s", (normalized,))
            if cur.fetchone():
                raise ValueError("Nomor HP sudah ada di database.")
            cur.execute(
                """
                INSERT INTO customer_directory (
                    phone_number, is_active, is_used, reserved_by_token, reserved_at, shuffle_order, created_at, updated_at
                ) VALUES (%s, %s, 0, NULL, NULL, %s, %s, %s)
                """,
                (normalized, 1 if is_active else 0, secrets.randbits(63), now_str, now_str),
            )
        conn.commit()
        return normalized
    finally:
        conn.close()


def delete_customer_number(phone_number):
    normalized = normalize_phone_number(phone_number)
    if not normalized:
        return
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM customer_directory WHERE phone_number = %s", (normalized,))
    conn.commit()
    conn.close()


def reset_customer_status(phone_number):
    normalized = normalize_phone_number(phone_number)
    if not normalized:
        return
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE customer_directory
        SET is_used = 0,
            reserved_by_token = NULL,
            reserved_at = NULL,
            updated_at = %s
        WHERE phone_number = %s
        """,
        (get_now_db_string(), normalized),
    )
    conn.commit()
    conn.close()


def get_customer_number(phone_number):
    normalized = normalize_phone_number(phone_number)
    if not normalized:
        return None

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT phone_number, is_active, is_used, reserved_by_token, reserved_at, shuffle_order, created_at, updated_at
        FROM customer_directory
        WHERE phone_number = %s
        LIMIT 1
        """,
        (normalized,),
    )
    row = cur.fetchone()
    conn.close()
    return row


def mark_phone_as_invalid(phone_number, kc_token=None):
    normalized = normalize_phone_number(phone_number)
    if not normalized:
        return

    conn = get_db_connection()
    cur = conn.cursor()
    if kc_token:
        cur.execute(
            """
            UPDATE customer_directory
            SET is_active = 0,
                is_used = 0,
                reserved_by_token = NULL,
                reserved_at = NULL,
                updated_at = %s
            WHERE phone_number = %s
              AND (reserved_by_token = %s OR reserved_by_token IS NULL OR reserved_by_token = '')
            """,
            (get_now_db_string(), normalized, kc_token),
        )
    else:
        cur.execute(
            """
            UPDATE customer_directory
            SET is_active = 0,
                is_used = 0,
                reserved_by_token = NULL,
                reserved_at = NULL,
                updated_at = %s
            WHERE phone_number = %s
            """,
            (get_now_db_string(), normalized),
        )
    conn.commit()
    conn.close()


def normalize_customer_rows_param(raw_value):
    allowed_values = {"3", "10", "100", "1000", "10000", "all"}
    current = str(raw_value or "10").strip().lower()
    if current not in allowed_values:
        current = "10"
    return current


def customer_rows_value_to_limit(rows_value):
    return None if rows_value == "all" else int(rows_value)


def normalize_customer_sort(sort_by, sort_dir):
    allowed_sort_by = {
        "phone_number",
        "is_active",
        "status_pemakaian",
        "reserved_by_token",
        "reserved_at",
        "updated_at",
    }
    current_sort_by = str(sort_by or "reserved_at").strip().lower()
    if current_sort_by not in allowed_sort_by:
        current_sort_by = "reserved_at"

    current_sort_dir = "asc" if str(sort_dir or "desc").strip().lower() == "asc" else "desc"
    return current_sort_by, current_sort_dir


def build_customer_order_clause(sort_by, sort_dir):
    sort_by, sort_dir = normalize_customer_sort(sort_by, sort_dir)
    direction = sort_dir.upper()

    if sort_by == "phone_number":
        order_parts = [f"phone_number {direction}"]
    elif sort_by == "is_active":
        order_parts = [f"is_active {direction}"]
    elif sort_by == "status_pemakaian":
        expr = "CASE WHEN is_used = 1 THEN 2 WHEN reserved_by_token IS NOT NULL AND reserved_by_token != '' THEN 1 ELSE 0 END"
        order_parts = [f"{expr} {direction}"]
    elif sort_by == "reserved_by_token":
        order_parts = [
            "CASE WHEN reserved_by_token IS NULL OR reserved_by_token = '' THEN 1 ELSE 0 END ASC",
            f"reserved_by_token {direction}",
        ]
    elif sort_by == "updated_at":
        order_parts = [
            "CASE WHEN updated_at IS NULL OR updated_at = '' THEN 1 ELSE 0 END ASC",
            f"updated_at {direction}",
        ]
    else:
        order_parts = [
            "CASE WHEN reserved_at IS NULL OR reserved_at = '' THEN 1 ELSE 0 END ASC",
            f"reserved_at {direction}",
        ]

    if sort_by != "updated_at":
        order_parts.append("updated_at DESC")
    if sort_by != "phone_number":
        order_parts.append("phone_number ASC")

    return ", ".join(order_parts)


def get_all_customer_numbers(limit=None, sort_by="reserved_at", sort_dir="desc"):
    conn = get_db_connection()
    cur = conn.cursor()

    order_clause = build_customer_order_clause(sort_by, sort_dir)

    query = f"""
        SELECT phone_number, is_active, is_used, reserved_by_token, reserved_at, shuffle_order, created_at, updated_at
        FROM customer_directory
        ORDER BY {order_clause}
    """

    params = []
    if isinstance(limit, int) and limit > 0:
        query += "\n        LIMIT %s"
        params.append(limit)

    cur.execute(query, params)
    rows = cur.fetchall()
    conn.close()
    return rows


def get_customer_stats():
    rows = get_all_customer_numbers()
    total = len(rows)
    active = len([r for r in rows if r["is_active"] == 1])
    ready = len([r for r in rows if r["is_active"] == 1 and r["is_used"] == 0 and not r["reserved_by_token"]])
    used = len([r for r in rows if r["is_used"] == 1])
    return {
        "total_numbers": total,
        "active_numbers": active,
        "ready_numbers": ready,
        "used_numbers": used,
    }


def serialize_customer_row(row):
    return {
        "phone_number": row["phone_number"],
        "is_active": row["is_active"],
        "is_used": row["is_used"],
        "reserved_by_token": row["reserved_by_token"],
        "reserved_at": row["reserved_at"],
        "shuffle_order": row["shuffle_order"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def get_customer_active_status_label(row):
    return "Aktif" if row["is_active"] == 1 else "Nonaktif"


def get_customer_usage_status_label(row):
    if row["is_used"] == 1:
        return "Terpakai"
    if row["reserved_by_token"]:
        return "Sedang Dipakai"
    return "Siap"


def build_customer_numbers_export_excel(sort_by="reserved_at", sort_dir="desc", limit=None):
    rows = get_all_customer_numbers(limit=limit, sort_by=sort_by, sort_dir=sort_dir)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Database Customer"
    worksheet.freeze_panes = "A2"

    headers = [
        "Nomor HP",
        "Status Aktif",
        "Status Pemakaian",
        "Reserved KC",
        "Reserved At",
        "Created At",
        "Updated At",
        "Shuffle Order",
        "Is Active",
        "Is Used",
    ]
    worksheet.append(headers)

    for row in rows:
        worksheet.append([
            row["phone_number"] or "",
            get_customer_active_status_label(row),
            get_customer_usage_status_label(row),
            row["reserved_by_token"] or "",
            row["reserved_at"] or "",
            row["created_at"] or "",
            row["updated_at"] or "",
            row["shuffle_order"] if row["shuffle_order"] is not None else "",
            row["is_active"],
            row["is_used"],
        ])

    for column, width in {
        "A": 18,
        "B": 14,
        "C": 18,
        "D": 24,
        "E": 22,
        "F": 22,
        "G": 22,
        "H": 20,
        "I": 10,
        "J": 10,
    }.items():
        worksheet.column_dimensions[column].width = width

    worksheet.auto_filter.ref = worksheet.dimensions

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output.getvalue()


def import_customer_numbers(uploaded_file, is_active=1):
    filename = secure_filename(uploaded_file.filename or "")
    if not filename:
        raise ValueError("File import wajib dipilih.")

    ext = os.path.splitext(filename)[1].lower()
    if ext not in {".xlsx", ".csv"}:
        raise ValueError("Format file harus .xlsx atau .csv.")

    if ext == ".csv":
        uploaded_file.stream.seek(0)
        wrapper = TextIOWrapper(uploaded_file.stream, encoding="utf-8-sig", newline="")
        reader = csv.reader(wrapper)
        rows = list(reader)
        wrapper.detach()
    else:
        uploaded_file.stream.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()

    if not rows:
        raise ValueError("File import kosong.")

    first_row = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    header_candidates = {"nomor", "nomor hp", "phone", "msisdn", "telp", "telepon", "phone_number"}
    target_index = 0
    has_header = False
    for idx, cell in enumerate(first_row):
        if cell.lower() in header_candidates:
            target_index = idx
            has_header = True
            break

    start_index = 1 if has_header else 0
    inserted = 0
    duplicates = 0
    invalid = 0

    conn = get_db_connection()
    cur = conn.cursor()
    now_str = get_now_db_string()
    try:
        for row in rows[start_index:]:
            if row is None:
                continue
            if target_index >= len(row):
                invalid += 1
                continue
            phone = normalize_phone_number(row[target_index])
            if not phone:
                invalid += 1
                continue

            cur.execute("SELECT 1 FROM customer_directory WHERE phone_number = %s", (phone,))
            if cur.fetchone():
                duplicates += 1
                continue

            cur.execute(
                """
                INSERT INTO customer_directory (
                    phone_number, is_active, is_used, reserved_by_token, reserved_at, shuffle_order, created_at, updated_at
                ) VALUES (%s, %s, 0, NULL, NULL, %s, %s, %s)
                """,
                (phone, 1 if is_active else 0, secrets.randbits(63), now_str, now_str),
            )
            inserted += 1
        conn.commit()
    finally:
        conn.close()

    return {
        "inserted": inserted,
        "duplicates": duplicates,
        "invalid": invalid,
    }


def normalize_import_header(value):
    return " ".join(str(value or "").strip().lower().replace("_", " ").replace("-", " ").split())


def generate_kc_token(length=16):
    chars = string.ascii_uppercase + string.digits
    return "KC-" + "".join(secrets.choice(chars) for _ in range(length))


def generate_unique_kc_token(cur):
    for _ in range(10):
        kc_token = generate_kc_token()
        cur.execute("SELECT 1 FROM valid_kc_tokens WHERE kc_token = %s", (kc_token,))
        if not cur.fetchone():
            return kc_token
    raise RuntimeError("Gagal generate KC token unik. Silakan coba import ulang.")


def get_import_cell(row, header_index, field_name):
    idx = header_index.get(field_name)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def parse_positive_int(value, field_label):
    if value is None or str(value).strip() == "":
        raise ValueError(f"{field_label} wajib diisi.")

    if isinstance(value, float):
        if not value.is_integer():
            raise ValueError(f"{field_label} harus angka bulat.")
        parsed = int(value)
    else:
        current = str(value).strip()
        if current.endswith(".0"):
            current = current[:-2]
        if not current.isdigit():
            raise ValueError(f"{field_label} harus angka bulat.")
        parsed = int(current)

    if parsed < 1:
        raise ValueError(f"{field_label} minimal 1.")
    return parsed


def parse_nonnegative_int(value, field_label):
    if value is None or str(value).strip() == "":
        raise ValueError(f"{field_label} wajib diisi.")

    if isinstance(value, float):
        if not value.is_integer():
            raise ValueError(f"{field_label} harus angka bulat.")
        parsed = int(value)
    else:
        current = str(value).strip()
        if current.endswith(".0"):
            current = current[:-2]
        if not current.isdigit():
            raise ValueError(f"{field_label} harus angka bulat.")
        parsed = int(current)

    if parsed < 0:
        raise ValueError(f"{field_label} minimal 0.")
    return parsed


def parse_optional_positive_int(value, default_value):
    if value is None or str(value).strip() == "":
        return default_value

    return parse_positive_int(value, "Daily limit")


def parse_optional_active_value(value, default_value):
    if value is None or str(value).strip() == "":
        return int(default_value)

    return parse_active_value(value)


def parse_active_value(value):
    if value is None or str(value).strip() == "":
        raise ValueError("Status aktif wajib diisi.")

    current = str(value).strip().lower()
    if current in {"1", "true", "yes", "ya", "y", "aktif", "active"}:
        return 1
    if current in {"0", "false", "no", "tidak", "n", "nonaktif", "non aktif", "inactive"}:
        return 0
    raise ValueError("Status aktif harus berisi 1/0, ya/tidak, atau aktif/nonaktif.")


def import_cell_to_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def get_import_rows(uploaded_file):
    filename = secure_filename(uploaded_file.filename or "")
    if not filename:
        raise ValueError("File import wajib dipilih.")

    ext = os.path.splitext(filename)[1].lower()
    if ext not in {".xlsx", ".csv"}:
        raise ValueError("Format file harus .xlsx atau .csv.")

    if ext == ".csv":
        uploaded_file.stream.seek(0)
        wrapper = TextIOWrapper(uploaded_file.stream, encoding="utf-8-sig", newline="")
        reader = csv.reader(wrapper)
        rows = list(reader)
        wrapper.detach()
    else:
        uploaded_file.stream.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()

    if not rows:
        raise ValueError("File import kosong.")
    return rows


def extract_master_json_items(payload, master_type):
    current_type = str(master_type or "").strip().lower()
    data = payload.get("data") if isinstance(payload, dict) else payload

    if current_type == "bumo":
        if isinstance(data, dict):
            items = data.get("data")
        else:
            items = data
    elif current_type == "kc_area":
        if isinstance(data, dict):
            items = data.get("areas") or data.get("data")
        else:
            items = data
    else:
        raise ValueError("Jenis master data tidak valid.")

    if not isinstance(items, list):
        raise ValueError("Format JSON master data tidak sesuai.")
    return items


def get_master_import_rows(uploaded_file, master_type):
    filename = secure_filename(uploaded_file.filename or "")
    if not filename:
        raise ValueError("File import wajib dipilih.")

    ext = os.path.splitext(filename)[1].lower()
    if ext != ".json":
        return get_import_rows(uploaded_file)

    uploaded_file.stream.seek(0)
    payload = json.load(TextIOWrapper(uploaded_file.stream, encoding="utf-8-sig"))
    items = extract_master_json_items(payload, master_type)
    current_type = str(master_type or "").strip().lower()

    if current_type == "bumo":
        rows = [["id", "name", "category", "campaign_type", "is_active"]]
        for item in items:
            if not isinstance(item, dict):
                continue
            rows.append([
                item.get("id", ""),
                item.get("name", ""),
                item.get("category", ""),
                item.get("campaign_type", ""),
                item.get("is_active", 1),
            ])
    elif current_type == "kc_area":
        rows = [["area_id", "area_name", "is_active"]]
        for item in items:
            if not isinstance(item, dict):
                continue
            rows.append([
                item.get("id") or item.get("area_id") or item.get("value") or "",
                item.get("name") or item.get("area_name") or item.get("label") or "",
                item.get("is_active", 1),
            ])
    else:
        raise ValueError("Jenis master data tidak valid.")

    if len(rows) == 1:
        raise ValueError("File JSON tidak berisi data master yang bisa diimport.")
    return rows


def import_kc_tokens(uploaded_file):
    rows = get_import_rows(uploaded_file)
    first_row = [normalize_import_header(cell) for cell in rows[0]]

    header_aliases = {
        "kc_token": {"kc token", "token kc", "token"},
        "kc_name": {"kc name", "nama kc", "nama"},
        "team": {"team", "tim", "group", "grup"},
        "token_area": {"token area", "area token", "area", "kc_area", "kc area", "area kc"},
        "kc_username": {"kc username", "username kc", "username", "user", "login username"},
        "kc_password": {"kc password", "password kc", "password", "pass", "login password"},
        "bearer_token": {"bearer token", "bearer", "token bearer"},
        "daily_limit": {"daily limit", "limit harian", "limit", "kuota", "kuota harian"},
        "is_active": {"is active", "aktif", "active", "status"},
        "used_today": {"used today", "sudah terpakai", "terpakai", "total submit", "total submit hari ini", "submit hari ini", "pemakaian", "usage", "used"},
    }

    header_index = {}
    for field_name, aliases in header_aliases.items():
        for idx, header in enumerate(first_row):
            if header in aliases:
                header_index[field_name] = idx
                break

    required_fields = ["kc_name", "bearer_token", "daily_limit", "is_active"]
    missing_fields = [field for field in required_fields if field not in header_index]
    if missing_fields:
        raise ValueError(
            "Header wajib: kc_name, bearer_token, daily_limit, is_active. "
            f"Kolom belum ditemukan: {', '.join(missing_fields)}."
        )

    inserted = 0
    updated = 0
    invalid = 0
    skipped = 0
    sample_errors = []

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        for row_number, row in enumerate(rows[1:], start=2):
            if row is None or all(str(cell or "").strip() == "" for cell in row):
                skipped += 1
                continue

            try:
                kc_token = str(get_import_cell(row, header_index, "kc_token") or "").strip()
                kc_name = str(get_import_cell(row, header_index, "kc_name") or "").strip()
                team = str(get_import_cell(row, header_index, "team") or "").strip()
                token_area = str(get_import_cell(row, header_index, "token_area") or "").strip()
                kc_username = str(get_import_cell(row, header_index, "kc_username") or "").strip()
                kc_password = str(get_import_cell(row, header_index, "kc_password") or "").strip()
                bearer_token = str(get_import_cell(row, header_index, "bearer_token") or "").strip()

                if not kc_name:
                    raise ValueError("Nama KC kosong.")
                if not bearer_token:
                    raise ValueError("Bearer token kosong.")

                if not kc_token:
                    kc_token = generate_unique_kc_token(cur)

                cur.execute(
                    """
                    SELECT bearer_token, daily_limit, is_active, team, token_area, kc_username, kc_password, last_bearer_updated_at
                    FROM valid_kc_tokens
                    WHERE kc_token = %s
                    """,
                    (kc_token,),
                )
                existing = cur.fetchone()
                if existing and "team" not in header_index:
                    team = existing["team"] or ""
                if existing and "token_area" not in header_index:
                    token_area = existing["token_area"] or ""
                if existing and "kc_username" not in header_index:
                    kc_username = existing["kc_username"] or ""
                if existing and "kc_password" not in header_index:
                    kc_password = existing["kc_password"] or ""

                daily_limit = parse_positive_int(get_import_cell(row, header_index, "daily_limit"), "Daily limit")
                is_active = parse_active_value(get_import_cell(row, header_index, "is_active"))
                ensure_kc_username_unique(kc_username, exclude_kc_token=kc_token, cur=cur)

                cur.execute(
                    """
                    INSERT INTO valid_kc_tokens (
                        kc_token, kc_name, team, token_area, kc_username, kc_password,
                        bearer_token, daily_limit, is_active, last_bearer_updated_at
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (kc_token)
                    DO UPDATE SET
                        kc_name = EXCLUDED.kc_name,
                        team = EXCLUDED.team,
                        token_area = EXCLUDED.token_area,
                        kc_username = EXCLUDED.kc_username,
                        kc_password = EXCLUDED.kc_password,
                        bearer_token = EXCLUDED.bearer_token,
                        last_bearer_updated_at = CASE
                            WHEN valid_kc_tokens.bearer_token IS DISTINCT FROM EXCLUDED.bearer_token
                            THEN EXCLUDED.last_bearer_updated_at
                            ELSE valid_kc_tokens.last_bearer_updated_at
                        END,
                        daily_limit = EXCLUDED.daily_limit,
                        is_active = EXCLUDED.is_active
                    """,
                    (
                        kc_token,
                        kc_name,
                        team,
                        token_area,
                        kc_username,
                        kc_password,
                        bearer_token,
                        daily_limit,
                        int(is_active),
                        get_now_wib_string(),
                    ),
                )
                if "used_today" in header_index:
                    used_today = parse_nonnegative_int(get_import_cell(row, header_index, "used_today"), "Sudah terpakai")
                    cur.execute(
                        """
                        INSERT INTO kc_token_usage (kc_token, usage_date, total_submit)
                        VALUES (%s, %s, %s)
                        ON CONFLICT (kc_token, usage_date)
                        DO UPDATE SET total_submit = EXCLUDED.total_submit
                        """,
                        (kc_token, get_today_wib(), used_today),
                    )
                if existing:
                    updated += 1
                else:
                    inserted += 1
            except Exception as exc:
                invalid += 1
                if len(sample_errors) < 3:
                    sample_errors.append(f"Baris {row_number}: {exc}")

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    if inserted + updated == 0:
        detail = f" Contoh error: {' | '.join(sample_errors)}" if sample_errors else ""
        raise ValueError(f"Tidak ada KC token valid yang berhasil diimport.{detail}")

    return {
        "inserted": inserted,
        "updated": updated,
        "invalid": invalid,
        "skipped": skipped,
        "sample_errors": sample_errors,
    }


def find_kc_token_by_username(username, cur):
    normalized = str(username or "").strip()
    if not normalized:
        return None
    cur.execute(
        """
        SELECT kc_token, kc_name, team, token_area, daily_limit, is_active
        FROM valid_kc_tokens
        WHERE LOWER(BTRIM(kc_username)) = LOWER(BTRIM(%s))
        LIMIT 1
        """,
        (normalized,),
    )
    return cur.fetchone()


def login_kc_credentials_for_bearer(username, password, captcha_key, secret):
    recaptcha_token = _sl_solve_recaptcha(captcha_key)
    ts = _sl_utc_timestamp_ms()
    body_obj = {
        "identifier": username,
        "password": base64.b64encode(password.encode()).decode(),
        "type_of_web": _SL_DEFAULT_TYPE_OF_WEB,
        "recaptcha_token": recaptcha_token,
    }
    hash_val = _sl_build_hash(secret, ts, _SL_AUTH_ENDPOINT, body_obj)
    headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
        "Content-Type": "application/json",
        "Dnt": "1",
        "Hash": hash_val,
        "Origin": "https://letscml.id",
        "Referer": "https://letscml.id/",
        "Timestamp": ts,
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36",
    }
    resp = requests.post(_SL_BASE_URL + _SL_AUTH_ENDPOINT, json=body_obj, headers=headers, timeout=45)
    token, token_source = _sl_extract_token(resp)
    if not token:
        try:
            preview = str(resp.json())[:200]
        except Exception:
            preview = resp.text[:200]
        raise ValueError(f"Token tidak ditemukan. Status: {resp.status_code}. {preview}")
    return token, token_source


def parse_batch_bearer_rows(uploaded_file):
    rows = get_import_rows(uploaded_file)
    if not rows:
        raise ValueError("File import kosong.")

    first_row = [normalize_import_header(cell) for cell in rows[0]]
    has_header = (
        len(first_row) >= 2
        and first_row[0] in {"username", "user", "identifier", "kc username", "login username"}
        and first_row[1] in {"password", "pass", "kc password", "login password"}
    )
    data_rows = rows[1:] if has_header else rows

    parsed_rows = []
    skipped = 0
    for row_number, row in enumerate(data_rows, start=2 if has_header else 1):
        if row is None or all(import_cell_to_text(cell) == "" for cell in row):
            skipped += 1
            continue
        username = import_cell_to_text(row[0] if len(row) > 0 else "")
        password = import_cell_to_text(row[1] if len(row) > 1 else "")
        if not username or not password:
            skipped += 1
            continue
        parsed_rows.append({"row": row_number, "username": username, "password": password})

    if not parsed_rows:
        raise ValueError("Tidak ada row valid. Isi kolom A=username dan B=password.")
    return parsed_rows, skipped


def batch_import_bearer_tokens(uploaded_file):
    secret = os.environ.get("APP_HMAC_SECRET", "").strip()
    if not secret:
        raise ValueError("APP_HMAC_SECRET belum di-set.")

    captcha_key = os.environ.get("TWOCAPTCHA_API_KEY", "").strip()
    if not captcha_key:
        raise ValueError("TWOCAPTCHA_API_KEY belum di-set.")

    rows, skipped = parse_batch_bearer_rows(uploaded_file)
    inserted = 0
    updated = 0
    failed = 0
    sample_errors = []
    result_rows = []

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        for item in rows:
            processed_at = get_now_wib().strftime("%Y-%m-%d %H:%M:%S")
            processed_date = processed_at[:10]
            username = item["username"]
            password = item["password"]
            try:
                bearer_token, token_source = login_kc_credentials_for_bearer(username, password, captcha_key, secret)
                existing = find_kc_token_by_username(username, cur)
                if existing:
                    cur.execute(
                        """
                        UPDATE valid_kc_tokens
                        SET kc_username = %s,
                            kc_password = %s,
                            bearer_token = %s,
                            last_bearer_updated_at = %s
                        WHERE kc_token = %s
                        """,
                        (username, password, bearer_token, get_now_wib_string(), existing["kc_token"]),
                    )
                    updated += 1
                    result_rows.append({
                        "row": item["row"],
                        "tanggal": processed_date,
                        "waktu_proses": processed_at,
                        "username": username,
                        "password": password,
                        "kc_token": existing["kc_token"],
                        "kc_name": existing["kc_name"],
                        "bearer_token": bearer_token,
                        "status": "UPDATED",
                        "error": "",
                    })
                    logger.info(
                        "Batch bearer updated kc_token=%s username=%s source=%s",
                        existing["kc_token"],
                        username,
                        token_source,
                    )
                else:
                    kc_token = generate_unique_kc_token(cur)
                    cur.execute(
                        """
                        INSERT INTO valid_kc_tokens (
                            kc_token, kc_name, team, token_area, kc_username, kc_password,
                            bearer_token, daily_limit, is_active, created_date, last_bearer_updated_at
                        )
                        VALUES (%s, %s, '', '', %s, %s, %s, %s, 1, %s, %s)
                        """,
                        (
                            kc_token,
                            username,
                            username,
                            password,
                            bearer_token,
                            DEFAULT_DAILY_LIMIT,
                            get_today_wib(),
                            get_now_wib_string(),
                        ),
                    )
                    inserted += 1
                    result_rows.append({
                        "row": item["row"],
                        "tanggal": processed_date,
                        "waktu_proses": processed_at,
                        "username": username,
                        "password": password,
                        "kc_token": kc_token,
                        "kc_name": username,
                        "bearer_token": bearer_token,
                        "status": "INSERTED",
                        "error": "",
                    })
                    logger.info("Batch bearer inserted kc_token=%s username=%s source=%s", kc_token, username, token_source)
                conn.commit()
            except Exception as exc:
                conn.rollback()
                failed += 1
                result_rows.append({
                    "row": item["row"],
                    "tanggal": processed_date,
                    "waktu_proses": processed_at,
                    "username": username,
                    "password": password,
                    "kc_token": "",
                    "kc_name": "",
                    "bearer_token": "",
                    "status": "FAILED",
                    "error": str(exc),
                })
                logger.warning("Batch bearer failed row=%s username=%s error=%s", item["row"], username, exc)
                if len(sample_errors) < 5:
                    sample_errors.append(f"Baris {item['row']} ({username}): {exc}")
    finally:
        conn.close()

    if inserted + updated == 0:
        detail = f" Contoh error: {' | '.join(sample_errors)}" if sample_errors else ""
        raise ValueError(f"Tidak ada bearer token yang berhasil diambil.{detail}")

    return {
        "total": len(rows),
        "inserted": inserted,
        "updated": updated,
        "failed": failed,
        "skipped": skipped,
        "sample_errors": sample_errors,
        "rows": result_rows,
    }


def build_batch_bearer_export_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Bearer"
    ws.append(["row", "tanggal", "waktu_proses", "username", "password", "kc_token", "kc_name", "bearer_token", "status", "error"])
    for item in rows:
        ws.append([
            item.get("row", ""),
            item.get("tanggal", ""),
            item.get("waktu_proses", ""),
            item.get("username", ""),
            item.get("password", ""),
            item.get("kc_token", ""),
            item.get("kc_name", ""),
            item.get("bearer_token", ""),
            item.get("status", ""),
            item.get("error", ""),
        ])
    widths = {
        "A": 8,
        "B": 14,
        "C": 22,
        "D": 28,
        "E": 28,
        "F": 24,
        "G": 32,
        "H": 90,
        "I": 14,
        "J": 70,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def save_batch_bearer_export(rows):
    export_id = secrets.token_urlsafe(12)
    export_dir = os.path.join(tempfile.gettempdir(), "kc_batch_bearer_exports")
    os.makedirs(export_dir, exist_ok=True)
    export_path = os.path.join(export_dir, f"{export_id}.xlsx")
    with open(export_path, "wb") as fh:
        fh.write(build_batch_bearer_export_excel(rows))
    rows_path = os.path.join(export_dir, f"{export_id}.json")
    with open(rows_path, "w", encoding="utf-8") as fh:
        json.dump(rows, fh, ensure_ascii=False)
    return export_id


def get_batch_bearer_export_path(export_id):
    safe_id = re.sub(r"[^A-Za-z0-9_-]", "", str(export_id or ""))
    if not safe_id:
        return ""
    export_dir = os.path.join(tempfile.gettempdir(), "kc_batch_bearer_exports")
    return os.path.join(export_dir, f"{safe_id}.xlsx")


def get_batch_bearer_rows(export_id):
    safe_id = re.sub(r"[^A-Za-z0-9_-]", "", str(export_id or ""))
    if not safe_id:
        return []
    export_dir = os.path.join(tempfile.gettempdir(), "kc_batch_bearer_exports")
    rows_path = os.path.join(export_dir, f"{safe_id}.json")
    if not os.path.exists(rows_path):
        return []
    with open(rows_path, "r", encoding="utf-8") as fh:
        rows = json.load(fh)
    return rows if isinstance(rows, list) else []


def get_local_master_data_options():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT name
        FROM bumo_master
        WHERE is_active = 1
        ORDER BY name ASC
        """
    )
    bumo_rows = cur.fetchall()
    cur.execute(
        """
        SELECT area_id, area_name
        FROM kc_area_master
        WHERE is_active = 1
        ORDER BY area_name ASC
        """
    )
    area_rows = cur.fetchall()
    conn.close()
    return {
        "bumo_options": [{"label": row["name"], "value": row["name"]} for row in bumo_rows],
        "kc_area_options": [{"label": row["area_name"], "value": str(row["area_id"])} for row in area_rows],
    }


def get_local_master_data_summary():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT name
        FROM bumo_master
        WHERE is_active = 1
        ORDER BY name ASC
        """
    )
    bumo_rows = cur.fetchall()
    cur.execute("SELECT COUNT(*) AS total FROM kc_area_master WHERE is_active = 1")
    area_count = cur.fetchone()["total"]
    conn.close()
    return {
        "bumo_options": [{"label": row["name"], "value": row["name"]} for row in bumo_rows],
        "kc_area_count": int(area_count or 0),
    }


def search_local_kc_area_options(query="", selected_area_id="", limit=60):
    keyword = str(query or "").strip()
    selected_id = str(selected_area_id or "").strip()
    safe_limit = max(1, min(int(limit or 60), 100))

    conn = get_db_connection()
    cur = conn.cursor()
    rows = []
    if keyword:
        cur.execute(
            """
            SELECT area_id, area_name
            FROM kc_area_master
            WHERE is_active = 1
              AND area_name ILIKE %s
            ORDER BY
              CASE WHEN area_name ILIKE %s THEN 0 ELSE 1 END,
              area_name ASC
            LIMIT %s
            """,
            (f"%{keyword}%", f"{keyword}%", safe_limit),
        )
        rows = cur.fetchall()

    if selected_id and not any(str(row["area_id"]) == selected_id for row in rows):
        cur.execute(
            """
            SELECT area_id, area_name
            FROM kc_area_master
            WHERE is_active = 1
              AND area_id = %s
            LIMIT 1
            """,
            (selected_id,),
        )
        selected_row = cur.fetchone()
        if selected_row:
            rows.insert(0, selected_row)

    conn.close()
    return [{"label": row["area_name"], "value": str(row["area_id"])} for row in rows[:safe_limit]]


def get_master_data_counts():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) AS total FROM bumo_master")
    bumo_total = cur.fetchone()["total"]
    cur.execute("SELECT COUNT(*) AS total FROM bumo_master WHERE is_active = 1")
    bumo_active = cur.fetchone()["total"]
    cur.execute("SELECT COUNT(*) AS total FROM kc_area_master")
    area_total = cur.fetchone()["total"]
    cur.execute("SELECT COUNT(*) AS total FROM kc_area_master WHERE is_active = 1")
    area_active = cur.fetchone()["total"]
    conn.close()
    return {
        "bumo_total": bumo_total,
        "bumo_active": bumo_active,
        "kc_area_total": area_total,
        "kc_area_active": area_active,
    }


def get_master_data_cache_version():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT COALESCE(MAX(updated_at), '') AS version
        FROM (
            SELECT updated_at FROM bumo_master
            UNION ALL
            SELECT updated_at FROM kc_area_master
        ) master_updates
        """
    )
    row = cur.fetchone()
    conn.close()
    return row["version"] if row else ""


def get_master_data_sync_tokens():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT kc_token, kc_name, bearer_token
        FROM valid_kc_tokens
        WHERE is_active = 1
          AND BTRIM(COALESCE(bearer_token, '')) != ''
        ORDER BY kc_name ASC, kc_token ASC
        """
    )
    rows = cur.fetchall()
    conn.close()
    if not rows:
        raise ValueError("Tidak ada KC token aktif dengan bearer token untuk load master data.")
    return rows


def upsert_master_data_options(bumo_options, kc_area_options):
    now_str = get_now_db_string()
    inserted_bumo = 0
    updated_bumo = 0
    inserted_area = 0
    updated_area = 0

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        for item in bumo_options:
            name = str(item.get("value") or item.get("label") or "").strip()
            if not name:
                continue
            cur.execute("SELECT 1 FROM bumo_master WHERE name = %s", (name,))
            existing = cur.fetchone()
            cur.execute(
                """
                INSERT INTO bumo_master (name, is_active, created_at, updated_at)
                VALUES (%s, 1, %s, %s)
                ON CONFLICT (name)
                DO UPDATE SET
                    is_active = 1,
                    updated_at = EXCLUDED.updated_at
                """,
                (name, now_str, now_str),
            )
            if existing:
                updated_bumo += 1
            else:
                inserted_bumo += 1

        for item in kc_area_options:
            area_id = str(item.get("value") or "").strip()
            area_name = str(item.get("label") or "").strip()
            if not area_id or not area_name:
                continue
            cur.execute("SELECT 1 FROM kc_area_master WHERE area_id = %s", (area_id,))
            existing = cur.fetchone()
            cur.execute(
                """
                INSERT INTO kc_area_master (area_id, area_name, is_active, created_at, updated_at)
                VALUES (%s, %s, 1, %s, %s)
                ON CONFLICT (area_id)
                DO UPDATE SET
                    area_name = EXCLUDED.area_name,
                    is_active = 1,
                    updated_at = EXCLUDED.updated_at
                """,
                (area_id, area_name, now_str, now_str),
            )
            if existing:
                updated_area += 1
            else:
                inserted_area += 1

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    if inserted_bumo + updated_bumo == 0 or inserted_area + updated_area == 0:
        raise ValueError("Data dari API kosong atau formatnya tidak sesuai.")

    return {
        "bumo_inserted": inserted_bumo,
        "bumo_updated": updated_bumo,
        "kc_area_inserted": inserted_area,
        "kc_area_updated": updated_area,
    }


def sync_master_data_from_api(manual_bearer_token=""):
    current_manual_token = str(manual_bearer_token or "").strip()
    if current_manual_token:
        bumo_options = fetch_bumo_options(current_manual_token)
        kc_area_options = fetch_kc_area_options(current_manual_token)
        result = upsert_master_data_options(bumo_options, kc_area_options)
        result["kc_token"] = "manual"
        result["kc_name"] = "Bearer token manual"
        result["tried_count"] = 1
        result["failed_tokens"] = []
        return result

    token_rows = get_master_data_sync_tokens()
    failed_tokens = []

    for token_row in token_rows:
        bearer_token = (token_row["bearer_token"] or "").strip()
        try:
            bumo_options = fetch_bumo_options(bearer_token)
            kc_area_options = fetch_kc_area_options(bearer_token)
            result = upsert_master_data_options(bumo_options, kc_area_options)
            result["kc_token"] = token_row["kc_token"]
            result["kc_name"] = token_row["kc_name"]
            result["tried_count"] = len(failed_tokens) + 1
            result["failed_tokens"] = failed_tokens
            return result
        except Exception as exc:
            failed_tokens.append(f"{token_row['kc_name']} ({token_row['kc_token']}): {exc}")
            logger.warning(
                "sync master data failed with kc_token=%s kc_name=%s error=%s",
                token_row["kc_token"],
                token_row["kc_name"],
                exc,
            )

    preview = "; ".join(failed_tokens[:3])
    if len(failed_tokens) > 3:
        preview += f"; dan {len(failed_tokens) - 3} token lain"
    raise ValueError(f"Semua bearer token aktif gagal load master data. {preview}")


def import_master_data(uploaded_file, master_type):
    rows = get_master_import_rows(uploaded_file, master_type)
    first_row = [normalize_import_header(cell) for cell in rows[0]]
    current_type = str(master_type or "").strip().lower()

    header_aliases = {
        "bumo": {
            "name": {"name", "nama", "bumo", "nama bumo", "current bumo"},
            "is_active": {"is active", "aktif", "active", "status"},
        },
        "kc_area": {
            "area_id": {"id", "area id", "kc area id", "id area", "value", "kode"},
            "area_name": {"name", "nama", "area", "area name", "kc area", "nama area"},
            "is_active": {"is active", "aktif", "active", "status"},
        },
    }

    if current_type not in header_aliases:
        raise ValueError("Jenis master data tidak valid.")

    header_index = {}
    for field_name, aliases in header_aliases[current_type].items():
        for idx, header in enumerate(first_row):
            if header in aliases:
                header_index[field_name] = idx
                break

    required_fields = ["name"] if current_type == "bumo" else ["area_id", "area_name"]
    missing_fields = [field for field in required_fields if field not in header_index]
    if missing_fields:
        if current_type == "bumo":
            raise ValueError("Header wajib untuk BUMO: name.")
        raise ValueError("Header wajib untuk KC Area: area_id dan area_name.")

    inserted = 0
    updated = 0
    invalid = 0
    skipped = 0
    sample_errors = []
    now_str = get_now_db_string()

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        for row_number, row in enumerate(rows[1:], start=2):
            if row is None or all(str(cell or "").strip() == "" for cell in row):
                skipped += 1
                continue

            try:
                is_active = parse_optional_active_value(get_import_cell(row, header_index, "is_active"), 1)
                if current_type == "bumo":
                    name = str(get_import_cell(row, header_index, "name") or "").strip()
                    if not name:
                        raise ValueError("Nama BUMO kosong.")

                    cur.execute("SELECT 1 FROM bumo_master WHERE name = %s", (name,))
                    existing = cur.fetchone()
                    cur.execute(
                        """
                        INSERT INTO bumo_master (name, is_active, created_at, updated_at)
                        VALUES (%s, %s, %s, %s)
                        ON CONFLICT (name)
                        DO UPDATE SET
                            is_active = EXCLUDED.is_active,
                            updated_at = EXCLUDED.updated_at
                        """,
                        (name, int(is_active), now_str, now_str),
                    )
                else:
                    area_id = str(get_import_cell(row, header_index, "area_id") or "").strip()
                    area_name = str(get_import_cell(row, header_index, "area_name") or "").strip()
                    if not area_id:
                        raise ValueError("ID KC Area kosong.")
                    if not area_name:
                        raise ValueError("Nama KC Area kosong.")

                    cur.execute("SELECT 1 FROM kc_area_master WHERE area_id = %s", (area_id,))
                    existing = cur.fetchone()
                    cur.execute(
                        """
                        INSERT INTO kc_area_master (area_id, area_name, is_active, created_at, updated_at)
                        VALUES (%s, %s, %s, %s, %s)
                        ON CONFLICT (area_id)
                        DO UPDATE SET
                            area_name = EXCLUDED.area_name,
                            is_active = EXCLUDED.is_active,
                            updated_at = EXCLUDED.updated_at
                        """,
                        (area_id, area_name, int(is_active), now_str, now_str),
                    )

                if existing:
                    updated += 1
                else:
                    inserted += 1
            except Exception as exc:
                invalid += 1
                if len(sample_errors) < 3:
                    sample_errors.append(f"Baris {row_number}: {exc}")

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    if inserted + updated == 0:
        detail = f" Contoh error: {' | '.join(sample_errors)}" if sample_errors else ""
        raise ValueError(f"Tidak ada master data valid yang berhasil diimport.{detail}")

    return {
        "inserted": inserted,
        "updated": updated,
        "invalid": invalid,
        "skipped": skipped,
        "sample_errors": sample_errors,
    }


def reshuffle_ready_customer_numbers():
    conn = get_db_connection()
    cur = conn.cursor()
    rows = cur.execute(
        """
        SELECT phone_number
        FROM customer_directory
        WHERE is_active = 1
          AND is_used = 0
          AND (reserved_by_token IS NULL OR reserved_by_token = '')
        """
    ).fetchall()
    now_str = get_now_db_string()
    for row in rows:
        cur.execute(
            "UPDATE customer_directory SET shuffle_order = %s, updated_at = %s WHERE phone_number = %s",
            (secrets.randbits(63), now_str, row["phone_number"]),
        )
    conn.commit()
    conn.close()
    return len(rows)


def reset_customer_distribution():
    conn = get_db_connection()
    cur = conn.cursor()
    now_str = get_now_db_string()
    cur.execute(
        """
        UPDATE customer_directory
        SET reserved_by_token = NULL,
            reserved_at = NULL,
            updated_at = %s
        WHERE is_used = 0
        """,
        (now_str,),
    )
    rows = cur.execute(
        """
        SELECT phone_number
        FROM customer_directory
        WHERE is_active = 1
          AND is_used = 0
          AND (reserved_by_token IS NULL OR reserved_by_token = '')
        """
    ).fetchall()
    for row in rows:
        cur.execute(
            "UPDATE customer_directory SET shuffle_order = %s, updated_at = %s WHERE phone_number = %s",
            (secrets.randbits(63), now_str, row["phone_number"]),
        )
    conn.commit()
    conn.close()
    session.pop("assigned_phone_number", None)
    return len(rows)


def delete_unused_customer_numbers():
    release_stale_reserved_phones()
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        DELETE FROM customer_directory
        WHERE is_used = 0
          AND (reserved_by_token IS NULL OR reserved_by_token = '')
        """
    )
    deleted_count = cur.rowcount
    conn.commit()
    conn.close()
    return deleted_count


def delete_used_customer_numbers():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        DELETE FROM customer_directory
        WHERE is_used = 1
        """
    )
    deleted_count = cur.rowcount
    conn.commit()
    conn.close()
    return deleted_count


def get_kc_token_detail(kc_token):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT kc_token, kc_name, team, token_area, kc_username, kc_password,
               bearer_token, daily_limit, is_active, created_date, last_bearer_updated_at
        FROM valid_kc_tokens
        WHERE kc_token = %s
    """, (kc_token,))
    row = cur.fetchone()
    conn.close()
    return row


def get_all_kc_tokens():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT v.kc_token, v.kc_name, v.team, v.token_area, v.kc_username, v.kc_password,
               v.bearer_token, v.daily_limit, v.is_active,
               COALESCE(v.last_bearer_updated_at, '') AS last_bearer_updated_at,
               COALESCE(last_submit.last_submit_at, '') AS last_submit_at
        FROM valid_kc_tokens v
        LEFT JOIN (
            SELECT kc_token, MAX(created_at) AS last_submit_at
            FROM submission_attempts
            GROUP BY kc_token
        ) last_submit ON last_submit.kc_token = v.kc_token
        ORDER BY v.token_area ASC, v.kc_name ASC, v.kc_token ASC
    """)
    rows = cur.fetchall()
    conn.close()
    return rows


def ensure_kc_username_unique(kc_username, exclude_kc_token=None, cur=None):
    username = str(kc_username or "").strip()
    if not username:
        return

    own_connection = None
    if cur is None:
        own_connection = get_db_connection()
        cur = own_connection.cursor()

    try:
        query = [
            """
            SELECT kc_token
            FROM valid_kc_tokens
            WHERE LOWER(BTRIM(kc_username)) = LOWER(BTRIM(%s))
            """
        ]
        params = [username]
        if exclude_kc_token:
            query.append("AND kc_token != %s")
            params.append(exclude_kc_token)
        query.append("LIMIT 1")
        cur.execute(" ".join(query), params)
        existing = cur.fetchone()
        if existing:
            raise ValueError(
                f"Username KC \"{username}\" sudah dipakai oleh KC token {existing['kc_token']}."
            )
    finally:
        if own_connection:
            own_connection.close()


def normalize_token_rows_param(raw_value):
    allowed_values = {"3", "10", "100", "all"}
    current = str(raw_value or "10").strip().lower()
    if current not in allowed_values:
        current = "10"
    return current


def token_rows_value_to_limit(rows_value):
    return None if rows_value == "all" else int(rows_value)


def normalize_token_page_param(raw_value):
    try:
        page = int(str(raw_value or "1").strip())
    except (TypeError, ValueError):
        page = 1
    return max(1, page)


def paginate_token_rows(rows, rows_value="10", page=1):
    rows_value = normalize_token_rows_param(rows_value)
    total_count = len(rows)
    limit = token_rows_value_to_limit(rows_value)
    if limit is None:
        return list(rows), 1, 1

    total_pages = max(1, (total_count + limit - 1) // limit)
    current_page = min(normalize_token_page_param(page), total_pages)
    start = (current_page - 1) * limit
    end = start + limit
    return list(rows)[start:end], current_page, total_pages


def normalize_token_sort(sort_by, sort_dir):
    allowed_sort_by = {
        "kc_name",
        "team",
        "token_area",
        "kc_token",
        "kc_username",
        "daily_limit",
        "total_submit",
        "is_active",
        "last_bearer_updated_at",
        "last_submit_at",
    }
    current_sort_by = str(sort_by or "kc_name").strip().lower()
    if current_sort_by not in allowed_sort_by:
        current_sort_by = "kc_name"

    current_sort_dir = "desc" if str(sort_dir or "asc").strip().lower() == "desc" else "asc"
    return current_sort_by, current_sort_dir


def get_token_sort_value(row, sort_by):
    if sort_by in {"daily_limit", "total_submit", "is_active"}:
        return int(row.get(sort_by) or 0)
    return str(row.get(sort_by) or "").lower()


def get_token_last_submit_date(row):
    last_submit_at = str(row.get("last_submit_at") or "").strip()
    if not last_submit_at or last_submit_at == "-":
        return ""
    try:
        return datetime.strptime(last_submit_at[:10], "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        return ""


def filter_sort_limit_token_rows(
    rows,
    filter_text="",
    status_filter="",
    area_filter="",
    team_filter="",
    last_submit_date_from="",
    last_submit_date_to="",
    sort_by="kc_name",
    sort_dir="asc",
    rows_value="10",
):
    current_filter = str(filter_text or "").strip().lower()
    current_status = str(status_filter or "").strip().lower()
    current_area = str(area_filter or "").strip().lower()
    current_team = str(team_filter or "").strip().lower()
    current_last_submit_from = normalize_submission_date_filter(last_submit_date_from)
    current_last_submit_to = normalize_submission_date_filter(last_submit_date_to)
    sort_by, sort_dir = normalize_token_sort(sort_by, sort_dir)
    rows_value = normalize_token_rows_param(rows_value)

    filtered_rows = list(rows)
    if current_status == "aktif":
        filtered_rows = [row for row in filtered_rows if row.get("is_active") == 1]
    elif current_status == "nonaktif":
        filtered_rows = [row for row in filtered_rows if row.get("is_active") != 1]

    if current_area:
        filtered_rows = [row for row in filtered_rows if str(row.get("token_area") or "").strip().lower() == current_area]

    if current_team:
        filtered_rows = [row for row in filtered_rows if str(row.get("team") or "").strip().lower() == current_team]

    if current_last_submit_from:
        filtered_rows = [row for row in filtered_rows if get_token_last_submit_date(row) >= current_last_submit_from]

    if current_last_submit_to:
        filtered_rows = [
            row for row in filtered_rows
            if get_token_last_submit_date(row) and get_token_last_submit_date(row) <= current_last_submit_to
        ]

    if current_filter:
        filter_terms = [term.strip() for term in re.split(r"[\r\n,;|]+", current_filter) if term.strip()]
        if not filter_terms:
            filter_terms = [current_filter]

        def row_matches_term(row, term):
            token_value = str(row.get("kc_token") or "").strip().lower()
            if term.startswith("kc-"):
                return token_value == term

            haystack = " ".join([
                str(row.get("kc_name") or ""),
                str(row.get("team") or ""),
                str(row.get("token_area") or ""),
                str(row.get("kc_username") or ""),
                str(row.get("kc_password") or ""),
                str(row.get("bearer_token_masked") or ""),
                str(row.get("last_bearer_updated_at") or ""),
                str(row.get("last_submit_at") or ""),
                str(row.get("daily_limit") or ""),
                str(row.get("total_submit") or ""),
                "aktif" if row.get("is_active") == 1 else "nonaktif",
            ]).lower()
            return term in haystack

        filtered_rows = [
            row for row in filtered_rows
            if any(row_matches_term(row, term) for term in filter_terms)
        ]

    filtered_rows.sort(
        key=lambda row: (
            get_token_sort_value(row, sort_by),
            str(row.get("kc_name") or "").lower(),
            str(row.get("kc_token") or "").lower(),
        ),
        reverse=sort_dir == "desc",
    )

    total_filtered = len(filtered_rows)
    limit = token_rows_value_to_limit(rows_value)
    if limit is not None:
        filtered_rows = filtered_rows[:limit]
    return filtered_rows, total_filtered


def create_kc_token(kc_token, kc_name, token_area, bearer_token, daily_limit, kc_username="", kc_password="", team=""):
    conn = get_db_connection()
    cur = conn.cursor()
    ensure_kc_username_unique(kc_username, cur=cur)
    cur.execute("""
        INSERT INTO valid_kc_tokens (
            kc_token, kc_name, team, token_area, kc_username, kc_password,
            bearer_token, daily_limit, is_active, created_date, last_bearer_updated_at
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 1, %s, %s)
    """, (
        kc_token,
        kc_name,
        team,
        token_area,
        kc_username,
        kc_password,
        bearer_token,
        daily_limit,
        get_today_wib(),
        get_now_wib_string(),
    ))
    conn.commit()
    conn.close()


def update_kc_token(
    old_kc_token,
    new_kc_token,
    kc_name,
    bearer_token,
    daily_limit,
    is_active=None,
    team=None,
    token_area=None,
    kc_username=None,
    kc_password=None,
):
    conn = get_db_connection()
    cur = conn.cursor()
    if kc_username is not None:
        ensure_kc_username_unique(kc_username, exclude_kc_token=old_kc_token, cur=cur)
    cur.execute("SELECT bearer_token, last_bearer_updated_at FROM valid_kc_tokens WHERE kc_token = %s", (old_kc_token,))
    existing_token_row = cur.fetchone()
    last_bearer_updated_at = (
        get_now_wib_string()
        if not existing_token_row or (existing_token_row["bearer_token"] or "") != (bearer_token or "")
        else existing_token_row["last_bearer_updated_at"]
    )

    if old_kc_token != new_kc_token:
        cur.execute("""
            UPDATE kc_token_usage
            SET kc_token = %s
            WHERE kc_token = %s
        """, (new_kc_token, old_kc_token))

    if is_active is None:
        cur.execute("""
            UPDATE valid_kc_tokens
            SET kc_token = %s,
                kc_name = %s,
                team = COALESCE(%s, team),
                token_area = COALESCE(%s, token_area),
                kc_username = COALESCE(%s, kc_username),
                kc_password = COALESCE(%s, kc_password),
                bearer_token = %s,
                last_bearer_updated_at = %s,
                daily_limit = %s
            WHERE kc_token = %s
        """, (
            new_kc_token,
            kc_name,
            team,
            token_area,
            kc_username,
            kc_password,
            bearer_token,
            last_bearer_updated_at,
            daily_limit,
            old_kc_token,
        ))
    else:
        cur.execute("""
            UPDATE valid_kc_tokens
            SET kc_token = %s,
                kc_name = %s,
                team = COALESCE(%s, team),
                token_area = COALESCE(%s, token_area),
                kc_username = COALESCE(%s, kc_username),
                kc_password = COALESCE(%s, kc_password),
                bearer_token = %s,
                last_bearer_updated_at = %s,
                daily_limit = %s,
                is_active = %s
            WHERE kc_token = %s
        """, (
            new_kc_token,
            kc_name,
            team,
            token_area,
            kc_username,
            kc_password,
            bearer_token,
            last_bearer_updated_at,
            daily_limit,
            int(is_active),
            old_kc_token,
        ))

    conn.commit()
    conn.close()


def toggle_kc_token_status(kc_token):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        UPDATE valid_kc_tokens
        SET is_active = CASE WHEN is_active = 1 THEN 0 ELSE 1 END
        WHERE kc_token = %s
    """, (kc_token,))
    conn.commit()
    conn.close()


def deactivate_all_kc_tokens():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        UPDATE valid_kc_tokens
        SET is_active = 0
        WHERE is_active = 1
    """)
    updated_count = cur.rowcount
    conn.commit()
    conn.close()
    return updated_count


def activate_all_kc_tokens():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        UPDATE valid_kc_tokens
        SET is_active = 1
        WHERE is_active = 0
    """)
    updated_count = cur.rowcount
    conn.commit()
    conn.close()
    return updated_count


def set_kc_tokens_status(kc_tokens, is_active):
    tokens = normalize_kc_token_selection(kc_tokens)
    if not tokens:
        return 0
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE valid_kc_tokens
        SET is_active = %s
        WHERE kc_token = ANY(%s)
        """,
        (1 if is_active else 0, tokens),
    )
    updated_count = cur.rowcount
    conn.commit()
    conn.close()
    return updated_count


def delete_kc_token(kc_token):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE customer_directory
        SET reserved_by_token = NULL, reserved_at = NULL, updated_at = %s
        WHERE is_used = 0 AND reserved_by_token = %s
        """,
        (get_now_db_string(), kc_token),
    )
    cur.execute("DELETE FROM kc_token_usage WHERE kc_token = %s", (kc_token,))
    cur.execute("DELETE FROM valid_kc_tokens WHERE kc_token = %s", (kc_token,))
    conn.commit()
    conn.close()


def delete_kc_tokens(kc_tokens):
    tokens = normalize_kc_token_selection(kc_tokens)
    if not tokens:
        return 0

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE customer_directory
        SET reserved_by_token = NULL, reserved_at = NULL, updated_at = %s
        WHERE is_used = 0 AND reserved_by_token = ANY(%s)
        """,
        (get_now_db_string(), tokens),
    )
    cur.execute("DELETE FROM kc_token_usage WHERE kc_token = ANY(%s)", (tokens,))
    cur.execute("DELETE FROM valid_kc_tokens WHERE kc_token = ANY(%s)", (tokens,))
    deleted_count = cur.rowcount
    conn.commit()
    conn.close()
    return deleted_count


def get_today_kc_usage_summary(date_from=None, date_to=None):
    today = get_today_wib()
    qdate_from = date_from if date_from else today
    qdate_to = date_to if date_to else (date_from if date_from else today)
    conn = get_db_connection()
    cur = conn.cursor()
    if qdate_from == qdate_to:
        # Single-day view: carry over previous day's usage if limit was not reached
        cur.execute("""
            SELECT v.kc_token, v.kc_name, v.team, v.token_area, v.kc_username, v.kc_password, v.bearer_token, v.daily_limit, v.is_active,
                   CASE
                       WHEN today_u.total_submit IS NOT NULL THEN today_u.total_submit
                       WHEN prev_u.total_submit IS NOT NULL AND prev_u.total_submit < v.daily_limit THEN prev_u.total_submit
                       ELSE 0
                   END AS total_submit
            FROM valid_kc_tokens v
            LEFT JOIN kc_token_usage today_u
              ON today_u.kc_token = v.kc_token AND today_u.usage_date = %s
            LEFT JOIN LATERAL (
                SELECT total_submit FROM kc_token_usage
                WHERE kc_token = v.kc_token AND usage_date < %s
                ORDER BY usage_date DESC LIMIT 1
            ) prev_u ON TRUE
            ORDER BY v.kc_name ASC, v.kc_token ASC
        """, (qdate_from, qdate_from))
    else:
        cur.execute("""
            SELECT v.kc_token, v.kc_name, v.team, v.token_area, v.kc_username, v.kc_password, v.bearer_token, v.daily_limit, v.is_active,
                   COALESCE(SUM(u.total_submit), 0) AS total_submit
            FROM valid_kc_tokens v
            LEFT JOIN kc_token_usage u
              ON v.kc_token = u.kc_token
             AND u.usage_date >= %s AND u.usage_date <= %s
            GROUP BY v.kc_token, v.kc_name, v.team, v.token_area, v.kc_username, v.kc_password, v.bearer_token, v.daily_limit, v.is_active
            ORDER BY v.kc_name ASC, v.kc_token ASC
        """, (qdate_from, qdate_to))
    rows = cur.fetchall()
    conn.close()
    display = qdate_from if qdate_from == qdate_to else f"{qdate_from} s/d {qdate_to}"
    return rows, display


def get_all_kc_usage_data(date_from="", date_to="", active_only=False):
    conn = get_db_connection()
    cur = conn.cursor()
    query = ["""
        SELECT u.usage_date, v.kc_name, v.token_area, u.kc_token, v.kc_username, v.daily_limit, u.total_submit
        FROM kc_token_usage u
        LEFT JOIN valid_kc_tokens v ON v.kc_token = u.kc_token
        WHERE 1=1
    """]
    params = []
    if date_from:
        query.append("AND u.usage_date >= %s")
        params.append(date_from)
    if date_to:
        query.append("AND u.usage_date <= %s")
        params.append(date_to)
    if active_only:
        query.append("AND u.total_submit > 0")
    query.append("ORDER BY u.usage_date DESC, v.kc_name ASC")
    cur.execute(" ".join(query), params)
    rows = cur.fetchall()
    conn.close()
    return rows


def build_kc_usage_export_excel(date_from="", date_to="", active_only=False):
    rows = get_all_kc_usage_data(date_from=date_from, date_to=date_to, active_only=active_only)
    wb = Workbook()
    ws = wb.active
    ws.title = "Usage KC Token"
    ws.append(["Tanggal", "Nama KC", "Area", "KC Token", "Username", "Limit Harian", "Submit"])
    for row in rows:
        ws.append([
            row["usage_date"],
            row["kc_name"] or "",
            row["token_area"] or "",
            row["kc_token"],
            row["kc_username"] or "",
            row["daily_limit"] or 0,
            row["total_submit"],
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def build_kc_token_export_csv():
    rows, usage_date = get_today_kc_usage_summary()
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "kc_name",
        "team",
        "token_area",
        "kc_token",
        "kc_username",
        "kc_password",
        "bearer_token",
        "daily_limit",
        "is_active",
        "sudah_terpakai",
    ])
    for row in rows:
        writer.writerow([
            row["kc_name"],
            row["team"] if "team" in row else "",
            row["token_area"],
            row["kc_token"],
            row["kc_username"],
            row["kc_password"],
            row["bearer_token"],
            row["daily_limit"],
            row["is_active"],
            row["total_submit"],
        ])
    return "\ufeff" + output.getvalue(), usage_date


def get_selected_kc_token_export_rows(kc_tokens):
    tokens = normalize_kc_token_selection(kc_tokens)
    if not tokens:
        return []

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT kc_name, token_area, kc_username, kc_password, kc_token
        FROM valid_kc_tokens
        WHERE kc_token = ANY(%s)
        ORDER BY array_position(%s, kc_token)
        """,
        (tokens, tokens),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def build_selected_kc_token_export_text(kc_tokens):
    rows = get_selected_kc_token_export_rows(kc_tokens)
    blocks = []
    for row in rows:
        blocks.append(
            "\n".join([
                f"Nama KC : {row['kc_name'] or ''}",
                f"Area : {row['token_area'] or ''}",
                f"Username : {row['kc_username'] or ''}",
                f"Pasword : {row['kc_password'] or ''}",
                f"Token : {row['kc_token'] or ''}",
            ])
        )
    return "\n\n".join(blocks), len(rows)


def build_selected_kc_token_export_excel(kc_tokens):
    rows = get_selected_kc_token_export_rows(kc_tokens)
    wb = Workbook()
    ws = wb.active
    ws.title = "KC Token Terpilih"
    ws.append(["Nama KC", "Area", "Username", "Pasword", "Token"])
    for row in rows:
        ws.append([
            row["kc_name"] or "",
            row["token_area"] or "",
            row["kc_username"] or "",
            row["kc_password"] or "",
            row["kc_token"] or "",
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read(), len(rows)


def build_full_kc_token_export_excel(args):
    def parse_date_arg(key):
        return normalize_submission_date_filter(args.get(key))

    usage_date_from = parse_date_arg("usage_date_from")
    usage_date_to = parse_date_arg("usage_date_to")
    token_filter = (args.get("token_filter") or "").strip()
    status_filter = (args.get("token_status_filter") or "").strip().lower()
    area_filter = (args.get("token_area_filter") or "").strip()
    team_filter = (args.get("token_team_filter") or "").strip()
    last_submit_date_from = parse_date_arg("token_last_submit_date_from")
    last_submit_date_to = parse_date_arg("token_last_submit_date_to")
    if status_filter not in ("", "aktif", "nonaktif"):
        status_filter = ""
    sort_by, sort_dir = normalize_token_sort(
        args.get("token_sort_by", "kc_name"),
        args.get("token_sort_dir", "asc"),
    )

    token_rows = get_all_kc_tokens()
    usage_rows, usage_date = get_today_kc_usage_summary(
        date_from=usage_date_from or None,
        date_to=usage_date_to or None,
    )
    usage_by_token = {row["kc_token"]: row["total_submit"] for row in usage_rows}
    purchase_counts_by_token = get_kc_purchase_counts(
        date_from=usage_date_from,
        date_to=usage_date_to,
    )

    export_rows = []
    for row in token_rows:
        purchase_counts = purchase_counts_by_token.get(row["kc_token"], {})
        export_rows.append({
            "kc_token": row["kc_token"],
            "kc_name": row["kc_name"],
            "team": row.get("team") or "-",
            "token_area": row["token_area"] or "-",
            "kc_username": row["kc_username"] or "-",
            "kc_password": row["kc_password"] or "-",
            "bearer_token": row["bearer_token"] or "",
            "last_bearer_updated_at": row.get("last_bearer_updated_at") or "-",
            "last_submit_at": row.get("last_submit_at") or "-",
            "daily_limit": row["daily_limit"],
            "total_submit": usage_by_token.get(row["kc_token"], 0),
            "purchase_yes": purchase_counts.get("purchase_yes", 0),
            "purchase_no": purchase_counts.get("purchase_no", 0),
            "lighter_yes": purchase_counts.get("lighter_yes", 0),
            "is_active": row["is_active"],
        })

    export_rows, export_count = filter_sort_limit_token_rows(
        rows=export_rows,
        filter_text=token_filter,
        status_filter=status_filter,
        area_filter=area_filter,
        team_filter=team_filter,
        last_submit_date_from=last_submit_date_from,
        last_submit_date_to=last_submit_date_to,
        sort_by=sort_by,
        sort_dir=sort_dir,
        rows_value="all",
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar KC Token"
    ws.append([
        "Nama KC",
        "Team",
        "Area",
        "KC Token",
        "Username",
        "Password",
        "Bearer Token",
        "Update Bearer",
        "Submit Terakhir",
        "Limit Harian",
        f"Total Submit (NOC) - {usage_date}",
        "Beli (NOT)",
        "Tidak Beli",
        "Lighter",
        "Status",
    ])
    for row in export_rows:
        ws.append([
            row["kc_name"],
            row["team"],
            row["token_area"],
            row["kc_token"],
            row["kc_username"],
            row["kc_password"],
            row["bearer_token"],
            row["last_bearer_updated_at"],
            row["last_submit_at"],
            row["daily_limit"],
            row["total_submit"],
            row["purchase_yes"],
            row["purchase_no"],
            row["lighter_yes"],
            "Aktif" if row["is_active"] == 1 else "Nonaktif",
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 48)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read(), export_count


def get_kc_token_usage(kc_token, usage_date):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT total_submit
        FROM kc_token_usage
        WHERE kc_token = %s AND usage_date = %s
    """, (kc_token, usage_date))
    row = cur.fetchone()
    conn.close()
    return row["total_submit"] if row else 0


def _get_effective_usage(kc_token, today, daily_limit):
    """Returns today's usage, carrying over from the most recent previous day if limit was not reached."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT total_submit FROM kc_token_usage WHERE kc_token = %s AND usage_date = %s",
        (kc_token, today),
    )
    row = cur.fetchone()
    if row:
        conn.close()
        return row["total_submit"]
    cur.execute("""
        SELECT total_submit FROM kc_token_usage
        WHERE kc_token = %s AND usage_date < %s
        ORDER BY usage_date DESC LIMIT 1
    """, (kc_token, today))
    prev = cur.fetchone()
    conn.close()
    if prev and prev["total_submit"] < daily_limit:
        return prev["total_submit"]
    return 0


def increment_kc_token_usage(kc_token, usage_date, daily_limit=None):
    conn = get_db_connection()
    cur = conn.cursor()

    if daily_limit is not None:
        cur.execute(
            "SELECT total_submit FROM kc_token_usage WHERE kc_token = %s AND usage_date = %s",
            (kc_token, usage_date),
        )
        if cur.fetchone() is None:
            cur.execute("""
                SELECT total_submit FROM kc_token_usage
                WHERE kc_token = %s AND usage_date < %s
                ORDER BY usage_date DESC LIMIT 1
            """, (kc_token, usage_date))
            prev = cur.fetchone()
            if prev and prev["total_submit"] < daily_limit:
                cur.execute("""
                    INSERT INTO kc_token_usage (kc_token, usage_date, total_submit)
                    VALUES (%s, %s, %s)
                """, (kc_token, usage_date, prev["total_submit"] + 1))
                conn.commit()
                conn.close()
                return

    cur.execute("""
        INSERT INTO kc_token_usage (kc_token, usage_date, total_submit)
        VALUES (%s, %s, 1)
        ON CONFLICT (kc_token, usage_date)
        DO UPDATE SET total_submit = kc_token_usage.total_submit + 1
    """, (kc_token, usage_date))

    conn.commit()
    conn.close()


def reset_kc_token_usage_today(kc_token, usage_date=None):
    target_date = usage_date or get_today_wib()
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO kc_token_usage (kc_token, usage_date, total_submit)
        VALUES (%s, %s, 0)
        ON CONFLICT (kc_token, usage_date)
        DO UPDATE SET total_submit = 0
    """, (kc_token, target_date))
    conn.commit()
    conn.close()
    return target_date


def reset_all_kc_token_usage_today(usage_date=None):
    target_date = usage_date or get_today_wib()
    token_rows = get_all_kc_tokens()
    if not token_rows:
        return 0, target_date

    conn = get_db_connection()
    cur = conn.cursor()
    for row in token_rows:
        cur.execute("""
            INSERT INTO kc_token_usage (kc_token, usage_date, total_submit)
            VALUES (%s, %s, 0)
            ON CONFLICT (kc_token, usage_date)
            DO UPDATE SET total_submit = 0
        """, (row["kc_token"], target_date))
    conn.commit()
    conn.close()
    return len(token_rows), target_date


def normalize_kc_token_selection(raw_tokens):
    tokens = []
    seen = set()
    if not isinstance(raw_tokens, list):
        return tokens
    for raw_token in raw_tokens:
        token = str(raw_token or "").strip()
        if token and token not in seen:
            tokens.append(token)
            seen.add(token)
    return tokens


def reset_selected_kc_token_usage_today(kc_tokens, usage_date=None):
    target_date = usage_date or get_today_wib()
    tokens = normalize_kc_token_selection(kc_tokens)
    if not tokens:
        return 0, target_date

    conn = get_db_connection()
    cur = conn.cursor()
    reset_count = 0
    for kc_token in tokens:
        cur.execute("""
            INSERT INTO kc_token_usage (kc_token, usage_date, total_submit)
            SELECT kc_token, %s, 0
            FROM valid_kc_tokens
            WHERE kc_token = %s
            ON CONFLICT (kc_token, usage_date)
            DO UPDATE SET total_submit = 0
        """, (target_date, kc_token))
        if cur.rowcount:
            reset_count += 1
    conn.commit()
    conn.close()
    return reset_count, target_date


def get_remaining_quota(kc_token, daily_limit):
    today = get_today_wib()
    used = _get_effective_usage(kc_token, today, daily_limit)
    remaining = max(0, daily_limit - used)
    return used, remaining, today


def is_daily_quota_exhausted(used_today, daily_limit):
    try:
        limit = int(daily_limit or DEFAULT_DAILY_LIMIT)
    except (TypeError, ValueError):
        limit = DEFAULT_DAILY_LIMIT
    if limit < 1:
        limit = DEFAULT_DAILY_LIMIT
    try:
        used = int(used_today or 0)
    except (TypeError, ValueError):
        used = 0
    return used >= limit


def _has_any_usage(kc_token):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM kc_token_usage WHERE kc_token = %s LIMIT 1", (kc_token,))
    row = cur.fetchone()
    conn.close()
    return row is not None


def auto_disable_kc_token_if_limit_reached(kc_token):
    kc_detail = get_kc_token_detail(kc_token)
    if not kc_detail:
        return False, 0, 0

    daily_limit = kc_detail["daily_limit"] or DEFAULT_DAILY_LIMIT
    if daily_limit < 1:
        daily_limit = DEFAULT_DAILY_LIMIT

    today = get_today_wib()
    used_today = _get_effective_usage(kc_token, today, daily_limit)

    if is_daily_quota_exhausted(used_today, daily_limit) and kc_detail["is_active"] == 1:
        update_kc_token(
            old_kc_token=kc_token,
            new_kc_token=kc_detail["kc_token"],
            kc_name=kc_detail["kc_name"],
            bearer_token=kc_detail["bearer_token"],
            daily_limit=daily_limit,
            is_active=0,
        )
        return True, used_today, daily_limit

    # Auto-disable token yang tidak pernah dipakai sama sekali dan sudah melewati hari pembuatannya
    created_date = (kc_detail.get("created_date") or "").strip()
    if used_today == 0 and created_date and created_date < today and kc_detail["is_active"] == 1:
        has_any_usage = _has_any_usage(kc_token)
        if not has_any_usage:
            update_kc_token(
                old_kc_token=kc_token,
                new_kc_token=kc_detail["kc_token"],
                kc_name=kc_detail["kc_name"],
                bearer_token=kc_detail["bearer_token"],
                daily_limit=daily_limit,
                is_active=0,
            )
            return True, 0, daily_limit

    return False, used_today, daily_limit



def pack_label_to_quantity(label):
    current = (label or "").strip().lower()
    if current.startswith("0"):
        return 0
    if current.startswith("1"):
        return 1
    if current.startswith("2"):
        return 2
    raise ValueError(f"Pack label tidak valid: {label}")


def quantity_to_pack_label(quantity):
    qty = int(quantity or 0)
    if qty <= 0:
        return "0 pack"
    if qty == 1:
        return "1 pack"
    if qty == 2:
        return "2 pack"
    raise ValueError(f"Quantity pack tidak valid: {quantity}")


def build_product_transactions_json(sp12_label, include_lighter=False):
    sp12_qty = pack_label_to_quantity(sp12_label)
    if sp12_qty <= 0:
        raise ValueError("Minimal CMKT SP12 harus 1 pack.")

    items = [{"product_name": "CMKT SP12", "quantity": sp12_qty}]
    if include_lighter:
        items.append({"product_name": "Lighter", "quantity": 1})
    return json.dumps(items, ensure_ascii=False, separators=(",", ":"))


def parse_product_transactions_to_pack_labels(value):
    sp12_qty = 0
    current = (value or "").strip()
    if current:
        items = json.loads(current)
        if not isinstance(items, list):
            raise ValueError("product_transactions harus berupa list JSON.")
        for item in items:
            if not isinstance(item, dict):
                continue
            name = str(item.get("product_name", "")).strip()
            qty = int(item.get("quantity", 0) or 0)
            if name == "CMKT SP12":
                sp12_qty = qty
            elif name == "Lighter":
                pass
            elif name:
                raise ValueError(f"Produk tidak valid: {name}")
    return quantity_to_pack_label(sp12_qty)


def normalize_product_transactions_from_form(form):
    sp12_label = (form.get("sp12_pack") or form.get("cmkt12_pack") or "").strip()
    include_lighter = (form.get("lighter") or "").strip() == "Ya"

    if sp12_label:
        return build_product_transactions_json(sp12_label, include_lighter), sp12_label

    current_value = (form.get("product_transactions") or "").strip()
    if current_value:
        sp12_label = parse_product_transactions_to_pack_labels(current_value)
        return build_product_transactions_json(sp12_label, include_lighter), sp12_label

    raise ValueError("Paket CMKT SP12 wajib dipilih.")


def _make_json_safe(value, _seen=None):
    if _seen is None:
        _seen = set()

    if value is None or isinstance(value, (str, int, float, bool)):
        return value

    if hasattr(value, "read") or hasattr(value, "stream"):
        return ""

    obj_id = id(value)
    if obj_id in _seen:
        return "[circular]"

    if isinstance(value, dict):
        _seen.add(obj_id)
        safe_dict = {}
        for key, item in value.items():
            safe_dict[str(key)] = _make_json_safe(item, _seen)
        _seen.discard(obj_id)
        return safe_dict

    if isinstance(value, (list, tuple, set)):
        _seen.add(obj_id)
        safe_list = [_make_json_safe(item, _seen) for item in value]
        _seen.discard(obj_id)
        return safe_list

    return str(value)


def safe_json_dumps(value, **kwargs):
    return json.dumps(_make_json_safe(value), **kwargs)


def normalize_submission_identity_text(value):
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def canonicalize_product_transactions(value):
    current = str(value or "").strip()
    if not current:
        return []
    try:
        items = json.loads(current)
    except Exception:
        return current
    if not isinstance(items, list):
        return current

    normalized = []
    for item in items:
        if not isinstance(item, dict):
            continue
        normalized.append({
            "product_name": normalize_submission_identity_text(item.get("product_name")),
            "quantity": int(item.get("quantity") or 0),
        })
    normalized.sort(key=lambda item: (item["product_name"], item["quantity"]))
    return normalized


def build_submission_identity_payload(request_summary):
    summary = request_summary or {}
    return {
        "kc_name": normalize_submission_identity_text(summary.get("kc_name")),
        "customer_name": normalize_submission_identity_text(summary.get("customer_name")),
        "age_range": normalize_submission_identity_text(summary.get("age_range")),
        "current_bumo": normalize_submission_identity_text(summary.get("current_bumo")),
        "kc_area": normalize_submission_identity_text(summary.get("kc_area")),
        "has_purchased": normalize_submission_identity_text(summary.get("has_purchased")),
        "lighter": normalize_submission_identity_text(summary.get("lighter")),
        "non_purchase_reasons": normalize_submission_identity_text(summary.get("non_purchase_reasons")),
        "product_transactions": canonicalize_product_transactions(summary.get("product_transactions")),
    }


def build_submission_identity_key(request_summary):
    canonical_payload = build_submission_identity_payload(request_summary)
    payload_text = safe_json_dumps(canonical_payload, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    return hashlib.sha256(payload_text.encode("utf-8")).hexdigest()


class DuplicateSubmissionBlocked(Exception):
    pass


def is_pending_duplicate_still_blocking(created_at_value):
    raw_value = str(created_at_value or "").strip()
    if not raw_value:
        return False
    try:
        created_at = datetime.strptime(raw_value[:19], "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return False
    age_seconds = (get_now_wib() - created_at).total_seconds()
    return age_seconds <= PENDING_DUPLICATE_BLOCK_SECONDS


def find_recent_duplicate_submission(kc_token, request_summary, exclude_submission_id=None):
    identity_key = build_submission_identity_key(request_summary)
    cutoff = (get_now_wib() - timedelta(minutes=DUPLICATE_SUBMISSION_WINDOW_MINUTES)).strftime("%Y-%m-%d %H:%M:%S")

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT submission_id, phone_number, status_local, created_at, request_summary_json
        FROM submission_attempts
        WHERE kc_token = %s
          AND created_at >= %s
          AND status_local IN ('PENDING', 'SUCCESS', 'LIKELY_SUCCESS')
        ORDER BY created_at DESC
        LIMIT 50
        """,
        (kc_token, cutoff),
    )
    rows = cur.fetchall()
    conn.close()

    for row in rows:
        if exclude_submission_id and row["submission_id"] == exclude_submission_id:
            continue
        if row["status_local"] == "PENDING" and not is_pending_duplicate_still_blocking(row["created_at"]):
            continue
        try:
            existing_summary = json.loads(row["request_summary_json"] or "{}")
        except Exception:
            existing_summary = {}
        if build_submission_identity_key(existing_summary) == identity_key:
            return {
                "submission_id": row["submission_id"],
                "phone_number": row["phone_number"],
                "status_local": row["status_local"],
                "created_at": row["created_at"],
            }
    return None


def build_hash(secret, method, endpoint, payload_obj, timestamp):
    parsed = urlparse(endpoint)
    pathname = parsed.path if parsed.path else endpoint
    query_string = f"?{parsed.query}" if parsed.query else ""
    body_string = safe_json_dumps(payload_obj, separators=(",", ":"), ensure_ascii=False)
    string_to_hash = timestamp + method.upper() + pathname + query_string + body_string
    hash_val = hmac.new(
        secret.encode("utf-8"),
        string_to_hash.encode("utf-8"),
        hashlib.sha256
    ).hexdigest()
    return hash_val, body_string, string_to_hash


def build_get_hash(secret, endpoint, timestamp):
    parsed = urlparse(endpoint)
    pathname = parsed.path if parsed.path else endpoint
    query_string = f"?{parsed.query}" if parsed.query else ""
    string_to_hash = timestamp + "GET" + pathname + query_string
    hash_val = hmac.new(
        secret.encode("utf-8"),
        string_to_hash.encode("utf-8"),
        hashlib.sha256
    ).hexdigest()
    return hash_val


def build_headers(timestamp, hash_val, bearer_token=""):
    headers = {
        "Accept": "application/json, text/plain, */*",
        "Timestamp": timestamp,
        "Hash": hash_val,
    }
    if bearer_token:
        headers["Authorization"] = f"Bearer {bearer_token}"
    return headers


def build_browser_style_headers(timestamp, hash_value, bearer_token=""):
    headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.9",
        "Timestamp": timestamp,
        "Hash": hash_value,
        "Origin": "https://letscml.id",
        "Referer": "https://letscml.id/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36",
        "Sec-Ch-Ua": '"Not-A.Brand";v="24", "Chromium";v="146"',
        "Sec-Ch-Ua-Mobile": "?0",
        "Sec-Ch-Ua-Platform": '"Windows"',
        "Sec-Fetch-Site": "cross-site",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Storage-Access": "active",
    }
    if bearer_token:
        headers["Authorization"] = f"Bearer {bearer_token}"
    return headers


def build_browser_style_headers_for_master(timestamp, hash_value, bearer_token=""):
    return build_browser_style_headers(timestamp, hash_value, bearer_token)


def guess_mime_type(filepath):
    mime_type, _ = mimetypes.guess_type(filepath)
    return mime_type or "application/octet-stream"


def generate_webkit_boundary():
    alphabet = string.ascii_letters + string.digits
    suffix = "".join(secrets.choice(alphabet) for _ in range(16))
    return "----WebKitFormBoundary" + suffix


def get_multipart_file_order(file_obj):
    ordered = []
    for key in ["transaction_photo", "chat_photo"]:
        if key in file_obj:
            ordered.append(key)
    for key in file_obj.keys():
        if key not in ordered:
            ordered.append(key)
    return ordered


def get_curl_style_field_order(body_obj):
    ordered = []
    preferred = [
        "phone_number",
        "customer_name",
        "age_range",
        "current_bumo",
        "campaign_type",
        "has_purchased",
        "submission_location",
        "non_purchase_reasons",
        "kc_area",
        "product_transactions",
    ]
    for key in preferred:
        if key in body_obj:
            ordered.append(key)
    for key in body_obj.keys():
        if key not in ordered:
            ordered.append(key)
    return ordered


def save_upload_to_temp(upload_file, prefix):
    if not upload_file or not upload_file.filename:
        return None

    original_name = secure_filename(upload_file.filename) or "upload.bin"
    suffix = os.path.splitext(original_name)[1] or ".bin"

    tmp = tempfile.NamedTemporaryFile(
        prefix=f"{prefix}_",
        suffix=suffix,
        delete=False,
    )
    tmp_path = tmp.name
    tmp.close()

    upload_file.save(tmp_path)
    return {
        "path": tmp_path,
        "filename": original_name,
    }


def build_multipart_body(body_obj, file_obj, boundary):
    crlf = b"\r\n"
    body = bytearray()

    for key in get_multipart_file_order(file_obj):
        file_info = file_obj.get(key)
        if not file_info:
            continue

        if isinstance(file_info, dict):
            path = file_info.get("path")
            filename = file_info.get("filename") or os.path.basename(path or "")
        else:
            path = file_info
            filename = os.path.basename(path)

        if not path:
            continue
        if not os.path.isfile(path):
            raise FileNotFoundError(f"File tidak ditemukan untuk field {key}: {path}")

        content_type = guess_mime_type(filename)

        body.extend(f"--{boundary}".encode("utf-8"))
        body.extend(crlf)
        body.extend(f'Content-Disposition: form-data; name="{key}"; filename="{filename}"'.encode("utf-8"))
        body.extend(crlf)
        body.extend(f"Content-Type: {content_type}".encode("utf-8"))
        body.extend(crlf)
        body.extend(crlf)

        with open(path, "rb") as f:
            body.extend(f.read())

        body.extend(crlf)

    for key in get_curl_style_field_order(body_obj):
        if key not in body_obj:
            continue

        value = body_obj[key]
        body.extend(f"--{boundary}".encode("utf-8"))
        body.extend(crlf)
        body.extend(f'Content-Disposition: form-data; name="{key}"'.encode("utf-8"))
        body.extend(crlf)
        body.extend(crlf)
        body.extend(str(value).encode("utf-8"))
        body.extend(crlf)

    body.extend(f"--{boundary}--".encode("utf-8"))
    body.extend(crlf)
    return bytes(body)

def fetch_bumo_options(bearer_token):
    if not bearer_token:
        return []

    timestamp = datetime.now(timezone.utc).isoformat(timespec="milliseconds").replace("+00:00", "Z")
    hash_val = build_get_hash(MASTERDATA_HMAC_SECRET, DEFAULT_BUMO_ENDPOINT, timestamp)
    headers = build_browser_style_headers_for_master(timestamp, hash_val, bearer_token)

    url = DEFAULT_BASE_URL.rstrip("/") + DEFAULT_BUMO_ENDPOINT
    response = requests.get(url, headers=headers, timeout=MASTERDATA_API_TIMEOUT_SECONDS, verify=False)
    response.raise_for_status()

    data = response.json()
    items = data["data"]["data"]
    return [{"label": item["name"], "value": item["name"]} for item in items]


def fetch_kc_area_options(bearer_token):
    if not bearer_token:
        return []

    timestamp = datetime.now(timezone.utc).isoformat(timespec="milliseconds").replace("+00:00", "Z")
    hash_val = build_get_hash(MASTERDATA_HMAC_SECRET, DEFAULT_KC_AREA_ENDPOINT, timestamp)
    headers = build_browser_style_headers_for_master(timestamp, hash_val, bearer_token)

    url = DEFAULT_BASE_URL.rstrip("/") + DEFAULT_KC_AREA_ENDPOINT
    response = requests.get(url, headers=headers, timeout=MASTERDATA_API_TIMEOUT_SECONDS, verify=False)
    response.raise_for_status()

    data = response.json()
    items = data["data"]["areas"]
    return [{"label": item["name"], "value": str(item["id"])} for item in items]


def send_survey_request(

    secret,
    base_url,
    endpoint,
    bearer_token,
    phone_number,
    customer_name,
    age_range,
    current_bumo,
    campaign_type,
    has_purchased,
    submission_location,
    kc_area,
    product_transactions,
    non_purchase_reasons,
    transaction_photo,
    chat_photo,
):
    body_obj = {
        "phone_number": phone_number,
        "customer_name": customer_name,
        "age_range": age_range,
        "current_bumo": current_bumo,
        "campaign_type": campaign_type,
        "has_purchased": has_purchased,
        "submission_location": submission_location,
        "kc_area": kc_area,
    }

    if product_transactions:
        body_obj["product_transactions"] = product_transactions

    if has_purchased == "false" and non_purchase_reasons:
        body_obj["non_purchase_reasons"] = non_purchase_reasons

    file_obj = {}
    temp_paths = []

    if transaction_photo and transaction_photo.filename:
        saved_file = save_upload_to_temp(transaction_photo, "transaction_photo")
        file_obj["transaction_photo"] = saved_file
        temp_paths.append(saved_file["path"])

    if chat_photo and chat_photo.filename:
        saved_file = save_upload_to_temp(chat_photo, "chat_photo")
        file_obj["chat_photo"] = saved_file
        temp_paths.append(saved_file["path"])

    try:
        def build_hash_payload():
            hash_payload = {}
            for key in get_multipart_file_order(file_obj):
                if key in file_obj:
                    hash_payload[key] = ""
            for key in get_curl_style_field_order(body_obj):
                if key in body_obj:
                    hash_payload[key] = body_obj[key]
            return hash_payload

        def build_request_once(attempt_no):
            started_at = time.monotonic()
            timestamp = datetime.now(timezone.utc).isoformat(timespec="milliseconds").replace("+00:00", "Z")
            hash_payload = build_hash_payload()

            hash_val, body_string, string_to_hash = build_hash(
                secret=secret,
                method="POST",
                endpoint=endpoint,
                payload_obj=hash_payload,
                timestamp=timestamp
            )

            boundary = generate_webkit_boundary()
            multipart_body = build_multipart_body(body_obj, file_obj, boundary)

            headers = build_browser_style_headers(timestamp, hash_val, bearer_token)
            headers["Content-Type"] = f"multipart/form-data; boundary={boundary}"

            url = base_url.rstrip("/") + endpoint

            try:
                response = requests.request(
                    method="POST",
                    url=url,
                    data=multipart_body,
                    headers=headers,
                    timeout=60,
                    verify=False
                )
            except requests.exceptions.Timeout as exc:
                elapsed_ms = int((time.monotonic() - started_at) * 1000)
                logger.warning(
                    "submit attempt=%s status=None error=timeout elapsed_ms=%s phone=%s timestamp=%s",
                    attempt_no,
                    elapsed_ms,
                    phone_number,
                    timestamp,
                )
                return {
                    "attempt_no": attempt_no,
                    "timestamp": timestamp,
                    "hash": hash_val,
                    "string_to_hash": string_to_hash,
                    "body_string_for_hash": body_string,
                    "status_code": None,
                    "response_body": {
                        "error": "timeout",
                        "message": str(exc),
                    },
                    "error_type": "timeout",
                    "url": url,
                    "boundary": boundary,
                    "elapsed_ms": elapsed_ms,
                }
            except requests.exceptions.RequestException as exc:
                elapsed_ms = int((time.monotonic() - started_at) * 1000)
                logger.warning(
                    "submit attempt=%s status=None error=network elapsed_ms=%s phone=%s timestamp=%s",
                    attempt_no,
                    elapsed_ms,
                    phone_number,
                    timestamp,
                )
                return {
                    "attempt_no": attempt_no,
                    "timestamp": timestamp,
                    "hash": hash_val,
                    "string_to_hash": string_to_hash,
                    "body_string_for_hash": body_string,
                    "status_code": None,
                    "response_body": {
                        "error": "network",
                        "message": str(exc),
                    },
                    "error_type": "network",
                    "url": url,
                    "boundary": boundary,
                    "elapsed_ms": elapsed_ms,
                }

            try:
                response_body = response.json()
            except Exception:
                response_body = response.text

            elapsed_ms = int((time.monotonic() - started_at) * 1000)
            logger.info(
                "submit attempt=%s status=%s elapsed_ms=%s phone=%s timestamp=%s",
                attempt_no,
                response.status_code,
                elapsed_ms,
                phone_number,
                timestamp,
            )

            return {
                "attempt_no": attempt_no,
                "timestamp": timestamp,
                "hash": hash_val,
                "string_to_hash": string_to_hash,
                "body_string_for_hash": body_string,
                "status_code": response.status_code,
                "response_body": response_body,
                "url": url,
                "boundary": boundary,
                "elapsed_ms": elapsed_ms,
            }

        attempts = [build_request_once(1)]
        final_result = attempts[-1]
        final_result = dict(final_result)
        final_result["attempts"] = attempts
        final_result["final_state"] = normalize_final_submit_state(final_result)
        return final_result

    finally:
        for path in temp_paths:
            try:
                if os.path.isfile(path):
                    os.remove(path)
            except Exception:
                pass


def mask_bearer_token(token):
    if not token:
        return "-"
    if len(token) <= 10:
        return token[:2] + "***"
    return token[:6] + "..." + token[-4:]


def build_submit_success_message(customer_name, phone_number, venue, final_state, quota_exhausted=False):
    headline = "Data berhasil di submit, silahkan cek di website camlet!"

    lines = [
        headline,
        "",
        f"Nama : {customer_name or '-'}",
        f"No. tlpn : {phone_number or '-'}",
        f"Venue : {venue or '-'}",
    ]

    if quota_exhausted:
        lines.extend(["", "Kuota harian token ini sudah habis, token otomatis dinonaktifkan."])

    return "\n".join(lines)


@app.route("/health")
def healthcheck():
    return jsonify({"status": "ok"})


@app.route("/", methods=["GET", "POST"])
def home():
    error = None

    if request.method == "POST":
        kc_token = request.form.get("kc_token", "").strip()

        if not kc_token:
            error = "KC token wajib diisi."
        else:
            kc_detail = get_kc_token_detail(kc_token)

            if not kc_detail or kc_detail["is_active"] != 1:
                error = "KC token tidak valid atau sudah nonaktif."
            else:
                auto_disabled, used_today, daily_limit = auto_disable_kc_token_if_limit_reached(kc_token)
                if auto_disabled or is_daily_quota_exhausted(used_today, daily_limit):
                    error = (
                        f"Kuota KC token hari ini sudah habis ({used_today}/{daily_limit}). "
                        "Token otomatis dinonaktifkan. Hubungi admin atau gunakan token lain."
                    )
                else:
                    session["kc_token"] = kc_token
                    session["token_login_date"] = get_today_wib()
                    session["kc_name"] = kc_detail["kc_name"] or "-"
                    session["daily_limit"] = daily_limit
                    return redirect(url_for("user_app"))

    return render_template("token_page.html", error=error)


@app.route("/logout")
def logout():
    clear_user_session()
    return redirect(url_for("home"))


@app.route("/api/master-data", methods=["POST"])
def api_master_data():
    try:
        session.pop("bearer_token", None)
        if "kc_token" not in session:
            return jsonify({"error": "Session token tidak ditemukan. Silakan login ulang."}), 401

        if clear_expired_user_session():
            return jsonify({"error": "Token sudah expired karena hari sudah berganti. Silakan login ulang."}), 401

        kc_token = session.get("kc_token", "").strip()
        kc_detail = get_kc_token_detail(kc_token)

        if not kc_detail or kc_detail["is_active"] != 1:
            clear_user_session()
            return jsonify({"error": "KC token tidak valid atau sudah nonaktif."}), 401

        auto_disabled, used_today, daily_limit = auto_disable_kc_token_if_limit_reached(kc_token)
        if auto_disabled or is_daily_quota_exhausted(used_today, daily_limit):
            clear_user_session()
            return jsonify({"error": "Kuota KC token hari ini sudah habis dan token otomatis dinonaktifkan."}), 401

        local_master_data = get_local_master_data_summary()
        if local_master_data["bumo_options"] and local_master_data["kc_area_count"] > 0:
            return jsonify({
                "bumo_options": local_master_data["bumo_options"],
                "kc_area_count": local_master_data["kc_area_count"],
                "source": "database",
            })

        return jsonify({
            "error": "Master data BUMO/KC Area lokal belum tersedia. Admin perlu load master data terlebih dahulu.",
        }), 400
    except requests.exceptions.HTTPError as e:
        status_code = e.response.status_code if e.response is not None else None
        if status_code == 401:
            logger.warning("api_master_data bearer token expired while loading BUMO/KC Area")
            return jsonify({"error": "Token bearer expired"}), 401

        logger.exception("api_master_data http error")
        return jsonify({"error": str(e)}), status_code or 500
    except Exception as e:
        logger.exception("api_master_data error")
        return jsonify({"error": str(e)}), 500


@app.route("/api/kc-areas/search", methods=["GET"])
def api_search_kc_areas():
    try:
        if "kc_token" not in session:
            return jsonify({"error": "Session token tidak ditemukan. Silakan login ulang."}), 401

        if clear_expired_user_session():
            return jsonify({"error": "Token sudah expired karena hari sudah berganti. Silakan login ulang."}), 401

        kc_token = session.get("kc_token", "").strip()
        kc_detail = get_kc_token_detail(kc_token)
        if not kc_detail or kc_detail["is_active"] != 1:
            clear_user_session()
            return jsonify({"error": "KC token tidak valid atau sudah nonaktif."}), 401

        query = (request.args.get("q") or "").strip()
        selected_id = (request.args.get("selected") or "").strip()
        if len(query) < 2 and not selected_id:
            return jsonify({"kc_area_options": [], "source": "database"})

        return jsonify({
            "kc_area_options": search_local_kc_area_options(query=query, selected_area_id=selected_id, limit=60),
            "source": "database",
        })
    except Exception as e:
        logger.exception("api_search_kc_areas error")
        return jsonify({"error": str(e)}), 500


@app.route("/user", methods=["GET", "POST"])
def user_app():
    session.pop("bearer_token", None)
    if "kc_token" not in session:
        return redirect(url_for("home"))

    if clear_expired_user_session():
        return redirect(url_for("home"))

    error = None
    error_detail = None
    can_retry_failed_submit = False
    success_message = None
    reset_form = False
    delayed_reset_seconds = 0
    result = None
    selected_age_range = request.form.get("age_range", "age-31-35")
    selected_has_purchased = (request.form.get("has_purchased") or "").strip().lower()
    if selected_has_purchased not in {"true", "false"}:
        selected_has_purchased = ""
    selected_non_purchase_reason = (request.form.get("non_purchase_reasons") or "").strip()
    selected_sp12_pack = request.form.get("sp12_pack", request.form.get("cmkt12_pack", DEFAULT_SP12_PACK))

    kc_token = (session.get("kc_token") or "").strip()

    kc_detail = get_kc_token_detail(kc_token)
    if not kc_detail or kc_detail["is_active"] != 1:
        clear_user_session()
        return redirect(url_for("home"))

    auto_disabled, used_today, daily_limit = auto_disable_kc_token_if_limit_reached(kc_token)
    if auto_disabled or is_daily_quota_exhausted(used_today, daily_limit):
        clear_user_session()
        return redirect(url_for("home"))

    bearer_token = (kc_detail["bearer_token"] or "").strip()
    kc_name = kc_detail["kc_name"] or "-"
    token_card_stats = get_kc_token_card_stats(kc_token, daily_limit)
    used_today = token_card_stats["used_today"]
    remaining_today = token_card_stats["remaining_today"]
    quota_date = token_card_stats["quota_date"]
    total_submit_noc = token_card_stats["used_today"]
    purchase_yes = token_card_stats["purchase_yes"]
    purchase_no = token_card_stats["purchase_no"]
    lighter_yes = token_card_stats["lighter_yes"]

    release_stale_reserved_phones(kc_token=kc_token)
    assigned_phone_number = session.get("assigned_phone_number")
    if assigned_phone_number and not refresh_reserved_phone(assigned_phone_number, kc_token):
        session.pop("assigned_phone_number", None)
        assigned_phone_number = ""

    if not assigned_phone_number:
        assigned_phone_number = reserve_phone_for_kc(kc_token)
        if assigned_phone_number:
            session["assigned_phone_number"] = assigned_phone_number

    if request.method == "POST":
        submission_id = secrets.token_hex(16)
        submission_attempt_created = False
        tracking_phone_number = (session.get("assigned_phone_number") or "").strip() or "UNKNOWN"
        lighter_val = ""
        try:
            create_submission_attempt(
                submission_id,
                tracking_phone_number,
                kc_token,
                {
                    "stage": "REQUEST_RECEIVED",
                    "kc_name": kc_name,
                    "message": "Submit sudah diterima backend. Menunggu validasi form dan upload.",
                },
            )
            submission_attempt_created = True

            if remaining_today <= 0:
                raise ValueError(
                    f"Kuota KC token hari ini sudah habis. "
                    f"Maksimal {daily_limit} form per hari. "
                    f"Silakan login ulang dan coba lagi besok setelah jam 00:00 WIB."
                )

            if not bearer_token:
                raise ValueError("Bearer token untuk KC ini belum diset.")

            secret = APP_HMAC_SECRET
            base_url = DEFAULT_BASE_URL
            endpoint = DEFAULT_ENDPOINT

            phone_number = (session.get("assigned_phone_number") or request.form.get("phone_number", "")).strip()
            customer_name = request.form.get("customer_name", "").strip()
            age_range = request.form.get("age_range", "").strip()
            current_bumo = request.form.get("current_bumo", "").strip()
            campaign_type = "kc"
            has_purchased = normalize_has_purchased_value(request.form.get("has_purchased"))
            submission_location = ""
            kc_area = request.form.get("kc_area", "").strip()
            kc_area_label = request.form.get("kc_area_label", "").strip() or kc_area
            product_transactions = ""
            non_purchase_reasons = ""

            transaction_photo = request.files.get("transaction_photo")
            chat_photo = request.files.get("chat_photo")

            if not phone_number.startswith("08") or not phone_number.isdigit() or len(phone_number) < 10 or len(phone_number) > 14:
                raise ValueError("Nomor HP harus diawali 08 dan panjang 10-14 digit.")
            if len(customer_name) < 2:
                raise ValueError("Nama customer minimal 2 karakter.")
            if not age_range:
                raise ValueError("Age range wajib dipilih.")
            if age_range not in VALID_AGE_RANGES:
                raise ValueError("Age range tidak valid.")
            if not current_bumo:
                raise ValueError("BUMO wajib dipilih.")
            if not kc_area:
                raise ValueError("KC Area wajib dipilih.")
            if has_purchased == "true":
                lighter_val = request.form.get("lighter", "").strip()
                if lighter_val not in ("Ya", "Tidak"):
                    raise ValueError("Pilihan lighter wajib diisi.")
                product_transactions, selected_sp12_pack = normalize_product_transactions_from_form(request.form)

                if not transaction_photo or not transaction_photo.filename:
                    raise ValueError("Foto transaksi wajib diupload.")
            else:
                selected_sp12_pack = DEFAULT_SP12_PACK
                transaction_photo = None
                non_purchase_reasons = (request.form.get("non_purchase_reasons") or "").strip()
                selected_non_purchase_reason = non_purchase_reasons
                if non_purchase_reasons not in VALID_NON_PURCHASE_REASONS:
                    raise ValueError("Alasan tidak membeli wajib dipilih.")

            if not chat_photo or not chat_photo.filename:
                raise ValueError("Screenshot chat wajib diupload.")

            request_summary = {
                "stage": "SUBMITTING_TO_API",
                "phone_number": phone_number,
                "kc_name": kc_name,
                "customer_name": customer_name,
                "age_range": age_range,
                "current_bumo": current_bumo,
                "kc_area": kc_area,
                "kc_area_label": kc_area_label,
                "has_purchased": has_purchased,
                "lighter": lighter_val,
                "non_purchase_reasons": non_purchase_reasons,
                "has_transaction_photo": bool(transaction_photo and transaction_photo.filename),
                "has_chat_photo": bool(chat_photo and chat_photo.filename),
                "product_transactions": product_transactions,
            }
            update_submission_request_summary(submission_id, phone_number, kc_token, request_summary)

            duplicate_submission = find_recent_duplicate_submission(
                kc_token=kc_token,
                request_summary=request_summary,
                exclude_submission_id=submission_id,
            )
            if duplicate_submission:
                duplicate_created_at = duplicate_submission.get("created_at") or "-"
                duplicate_phone = duplicate_submission.get("phone_number") or "-"
                duplicate_status = duplicate_submission.get("status_local") or "PENDING"
                duplicate_message = (
                    f"Submit dengan data yang sama sudah terdeteksi pada {duplicate_created_at} "
                    f"dengan nomor {duplicate_phone}. Sistem membatalkan kirim ulang otomatis. "
                    "Jika ini customer berbeda, ubah Nama Customer lalu kirim ulang."
                )
                update_submission_attempt(
                    submission_id,
                    "FAILED",
                    None,
                    duplicate_message,
                    [],
                )
                logger.warning(
                    "duplicate submission blocked submission_id=%s duplicate_submission_id=%s kc_token=%s duplicate_phone=%s duplicate_status=%s",
                    submission_id,
                    duplicate_submission.get("submission_id"),
                    kc_token,
                    duplicate_phone,
                    duplicate_status,
                )
                if duplicate_status == "PENDING":
                    error = (
                        "Data yang sama masih sedang diproses dari submit sebelumnya. "
                        f"Form ini akan direset otomatis dalam {PENDING_DUPLICATE_BLOCK_SECONDS} detik. "
                        "Sistem membatalkan kirim ulang agar tidak dobel submit. "
                        "Jika ini customer berbeda, isi ulang dan ganti Nama Customer sebelum kirim lagi."
                    )
                    delayed_reset_seconds = PENDING_DUPLICATE_BLOCK_SECONDS
                else:
                    success_message = (
                        "Data yang sama sudah pernah terkirim beberapa saat lalu, jadi sistem tidak mengirim ulang.\n\n"
                        f"Nomor sebelumnya : {duplicate_phone}\n"
                        f"Waktu submit : {duplicate_created_at}\n\n"
                        "Jika ini customer berbeda, ganti Nama Customer lalu kirim ulang."
                    )
                    reset_form = True
                    selected_age_range = "age-31-35"
                    selected_has_purchased = ""
                    selected_non_purchase_reason = ""
                    selected_sp12_pack = DEFAULT_SP12_PACK
                result = None
                raise DuplicateSubmissionBlocked(duplicate_message)

            result = send_survey_request(
                secret=secret,
                base_url=base_url,
                endpoint=endpoint,
                bearer_token=bearer_token,
                phone_number=phone_number,
                customer_name=customer_name,
                age_range=age_range,
                current_bumo=current_bumo,
                campaign_type=campaign_type,
                has_purchased=has_purchased,
                submission_location=submission_location,
                kc_area=kc_area,
                product_transactions=product_transactions,
                non_purchase_reasons=non_purchase_reasons,
                transaction_photo=transaction_photo,
                chat_photo=chat_photo,
            )

            final_state = result.get("final_state") or normalize_final_submit_state(result)
            mark_phone_invalid = should_mark_phone_invalid(result)

            if mark_phone_invalid:
                final_state = "INVALID"

            update_submission_attempt(
                submission_id,
                final_state,
                result.get("status_code"),
                result.get("response_body"),
                result.get("attempts") or [],
            )

            logger.info(
                "submit final submission_id=%s phone=%s state=%s summary=%s",
                submission_id,
                phone_number,
                final_state,
                summarize_submit_result(result),
            )

            if mark_phone_invalid:
                mark_phone_as_invalid(phone_number, kc_token)
                new_phone_number = reserve_next_phone_for_session(kc_token, previous_phone_number=phone_number)
                invalid_reason = "Submit ditolak oleh API dengan status 400."
                if new_phone_number:
                    assigned_phone_number = new_phone_number
                    logger.info(
                        "submit phone action submission_id=%s final_state=%s phone_changed=%s old_phone=%s new_phone=%s reason=%s",
                        submission_id,
                        final_state,
                        True,
                        phone_number,
                        new_phone_number,
                        invalid_reason,
                    )
                    error = (
                        f"{invalid_reason} "
                        "Nomor lama ditandai invalid dan diganti otomatis ke nomor lain. "
                        f"Nomor baru: {new_phone_number}"
                    )
                else:
                    assigned_phone_number = ""
                    logger.info(
                        "submit phone action submission_id=%s final_state=%s phone_changed=%s old_phone=%s new_phone=%s reason=%s",
                        submission_id,
                        final_state,
                        False,
                        phone_number,
                        "",
                        invalid_reason,
                    )
                    error = (
                        f"{invalid_reason} "
                        "Nomor lama ditandai invalid, tetapi saat ini tidak ada nomor pengganti yang tersedia."
                    )

            elif final_state in {"SUCCESS", "LIKELY_SUCCESS"}:
                mark_phone_as_used(phone_number, kc_token)
                increment_kc_token_usage(kc_token, quota_date, daily_limit)
                auto_disabled, _used_after_submit, _daily_limit_after_submit = auto_disable_kc_token_if_limit_reached(kc_token)
                used_today, remaining_today, quota_date = get_remaining_quota(kc_token, daily_limit)
                success_venue = kc_area_label or current_bumo
                if auto_disabled:
                    success_message = build_submit_success_message(
                        customer_name=customer_name,
                        phone_number=phone_number,
                        venue=success_venue,
                        final_state=final_state,
                        quota_exhausted=True,
                    )
                    reset_form = True
                    selected_age_range = "age-31-35"
                    selected_has_purchased = ""
                    selected_non_purchase_reason = ""
                    selected_sp12_pack = DEFAULT_SP12_PACK
                    assigned_phone_number = ""
                    logger.info(
                        "submit phone action submission_id=%s final_state=%s phone_changed=%s old_phone=%s new_phone=%s reason=%s",
                        submission_id,
                        final_state,
                        False,
                        phone_number,
                        "",
                        "quota auto-disabled",
                    )
                    clear_user_session()
                else:
                    assigned_phone_number = reserve_next_phone_for_session(kc_token, previous_phone_number=phone_number)
                    reset_form = True
                    selected_age_range = "age-31-35"
                    selected_has_purchased = ""
                    selected_non_purchase_reason = ""
                    selected_sp12_pack = DEFAULT_SP12_PACK
                    logger.info(
                        "submit phone action submission_id=%s final_state=%s phone_changed=%s old_phone=%s new_phone=%s reason=%s",
                        submission_id,
                        final_state,
                        bool(assigned_phone_number),
                        phone_number,
                        assigned_phone_number or "",
                        "success next phone reserved",
                    )
                    success_message = build_submit_success_message(
                        customer_name=customer_name,
                        phone_number=phone_number,
                        venue=success_venue,
                        final_state=final_state,
                    )
            else:
                body_msg = result.get("response_body")
                is_unauthorized = is_all_attempts_unauthorized(result)
                logger.info(
                    "submit phone action submission_id=%s final_state=%s phone_changed=%s old_phone=%s new_phone=%s reason=%s",
                    submission_id,
                    final_state,
                    False,
                    phone_number,
                    "",
                    "bearer token unauthorized" if is_unauthorized else "failed final state",
                )
                can_retry_failed_submit = True
                if is_unauthorized:
                    error = (
                        "Bearer token KC expired atau tidak valid. "
                        "Nomor tetap sama. Update bearer token di admin, lalu coba kirim lagi."
                    )
                else:
                    error = (
                        "Submit belum berhasil. "
                        "Nomor tetap sama. Upload ulang bukti, lalu coba kirim lagi."
                    )
                error_detail = (
                    f"Status code: {result.get('status_code')} | "
                    f"Attempts: {summarize_submit_result(result)} | "
                    f"Response: {body_msg}"
                )

        except DuplicateSubmissionBlocked:
            pass
        except Exception as e:
            logger.exception("submit route error")
            if submission_attempt_created:
                try:
                    update_submission_attempt(
                        submission_id,
                        "FAILED",
                        None,
                        {
                            "error": "submit_route_error",
                            "message": str(e),
                        },
                        [],
                    )
                except Exception:
                    logger.exception("failed to update submission attempt after route error")
            error = str(e)

    if session.get("kc_token") == kc_token:
        token_card_stats = get_kc_token_card_stats(kc_token, daily_limit)
        used_today = token_card_stats["used_today"]
        remaining_today = token_card_stats["remaining_today"]
        total_submit_noc = token_card_stats["used_today"]
        purchase_yes = token_card_stats["purchase_yes"]
        purchase_no = token_card_stats["purchase_no"]
        lighter_yes = token_card_stats["lighter_yes"]

    return render_template(
        "user_app.html",
        error=error,
        error_detail=error_detail,
        can_retry_failed_submit=can_retry_failed_submit,
        success_message=success_message,
        result=result,
        bumo_options=[],
        kc_area_options=[],
        used_today=used_today,
        remaining_today=remaining_today,
        daily_limit=daily_limit,
        total_submit_noc=total_submit_noc,
        purchase_yes=purchase_yes,
        purchase_no=purchase_no,
        lighter_yes=lighter_yes,
        kc_name=kc_name,
        age_range_options=AGE_RANGE_OPTIONS,
        product_pack_options=PRODUCT_PACK_OPTIONS,
        non_purchase_reason_options=NON_PURCHASE_REASON_OPTIONS,
        selected_age_range=selected_age_range,
        selected_has_purchased=selected_has_purchased,
        selected_non_purchase_reason=selected_non_purchase_reason,
        selected_sp12_pack=selected_sp12_pack,
        assigned_phone_number=assigned_phone_number or "",
        reset_form=reset_form,
        delayed_reset_seconds=delayed_reset_seconds,
        kc_token=kc_token,
        master_data_cache_version=get_master_data_cache_version(),
    )


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    error = None

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        if username == ADMIN_PAGE_USERNAME and password == ADMIN_PAGE_PASSWORD:
            session["is_admin_logged_in"] = True
            session["admin_page_username"] = username
            if wants_json_response():
                return jsonify({"success": True, "redirect_url": url_for("admin_dashboard")})
            return redirect(url_for("admin_dashboard"))
        else:
            error = "Username atau password admin salah."
            if wants_json_response():
                return jsonify({"success": False, "error": error}), 401

    return render_template("admin_login.html", error=error)


@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin_logged_in", None)
    session.pop("admin_page_username", None)
    return redirect(url_for("admin_login"))


@app.route("/leader/login", methods=["GET", "POST"])
def team_leader_login():
    error = None

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        account = get_team_leader(username)

        if not account or account["is_active"] != 1 or not check_password_hash(account["password_hash"], password):
            error = "Username atau password Team Leader salah."
            if wants_json_response():
                return jsonify({"success": False, "error": error}), 401
        else:
            clear_team_leader_session()
            session["is_team_leader_logged_in"] = True
            session["team_leader_username"] = account["username"]
            session["team_leader_name"] = account["leader_name"] or account["username"]
            if wants_json_response():
                return jsonify({"success": True, "redirect_url": url_for("team_leader_dashboard")})
            return redirect(url_for("team_leader_dashboard"))

    return render_template("team_leader_login.html", error=error)


@app.route("/leader/logout")
def team_leader_logout():
    clear_team_leader_session()
    return redirect(url_for("team_leader_login"))


def build_admin_dashboard_context(args):
    token_rows = get_all_kc_tokens()

    def parse_date_param(key):
        val = (args.get(key) or "").strip()
        if val:
            try:
                datetime.strptime(val, "%Y-%m-%d")
                return val
            except ValueError:
                pass
        return ""

    selected_usage_date_from = parse_date_param("usage_date_from")
    selected_usage_date_to = parse_date_param("usage_date_to")
    selected_usage_activity_filter = (args.get("usage_activity_filter") or "all").strip().lower()
    if selected_usage_activity_filter not in {"all", "active"}:
        selected_usage_activity_filter = "all"
    usage_rows, usage_date = get_today_kc_usage_summary(
        date_from=selected_usage_date_from or None,
        date_to=selected_usage_date_to or None,
    )
    recent_submissions = get_recent_submission_attempts(limit=10, include_response_text=False)

    selected_token_filter = (args.get("token_filter") or "").strip()
    selected_token_status_filter = (args.get("token_status_filter") or "").strip().lower()
    selected_token_area_filter = (args.get("token_area_filter") or "").strip()
    selected_token_team_filter = (args.get("token_team_filter") or "").strip()
    selected_token_last_submit_date_from = parse_date_param("token_last_submit_date_from")
    selected_token_last_submit_date_to = parse_date_param("token_last_submit_date_to")
    if selected_token_status_filter not in ("", "aktif", "nonaktif"):
        selected_token_status_filter = ""
    selected_token_rows = normalize_token_rows_param(args.get("token_rows", "10"))
    selected_token_page = normalize_token_page_param(args.get("token_page", "1"))
    selected_token_sort_by, selected_token_sort_dir = normalize_token_sort(
        args.get("token_sort_by", "kc_name"),
        args.get("token_sort_dir", "asc"),
    )

    total_tokens = len(token_rows)
    active_tokens = len([r for r in token_rows if r["is_active"] == 1])
    total_submit_today = sum(r["total_submit"] for r in usage_rows)
    token_area_options = sorted({str((r.get("token_area") or "")).strip() for r in token_rows if str((r.get("token_area") or "")).strip()}, key=str.lower)
    token_team_options = sorted({str((r.get("team") or "")).strip() for r in token_rows if str((r.get("team") or "")).strip()}, key=str.lower)

    usage_by_token = {row["kc_token"]: row["total_submit"] for row in usage_rows}
    purchase_counts_by_token = get_kc_purchase_counts(
        date_from=selected_usage_date_from,
        date_to=selected_usage_date_to,
    )
    masked_token_rows = []
    for row in token_rows:
        purchase_counts = purchase_counts_by_token.get(row["kc_token"], {})
        masked_token_rows.append({
            "kc_token": row["kc_token"],
            "kc_name": row["kc_name"],
            "team": row.get("team") or "-",
            "token_area": row["token_area"] or "-",
            "kc_username": row["kc_username"] or "-",
            "kc_password": row["kc_password"] or "-",
            "bearer_token_masked": mask_bearer_token(row["bearer_token"]),
            "last_bearer_updated_at": row.get("last_bearer_updated_at") or "-",
            "last_submit_at": row.get("last_submit_at") or "-",
            "daily_limit": row["daily_limit"],
            "total_submit": usage_by_token.get(row["kc_token"], 0),
            "purchase_yes": purchase_counts.get("purchase_yes", 0),
            "purchase_no": purchase_counts.get("purchase_no", 0),
            "lighter_yes": purchase_counts.get("lighter_yes", 0),
            "lighter_no": purchase_counts.get("lighter_no", 0),
            "is_active": row["is_active"],
        })

    filtered_token_rows, token_filtered_count = filter_sort_limit_token_rows(
        rows=masked_token_rows,
        filter_text=selected_token_filter,
        status_filter=selected_token_status_filter,
        area_filter=selected_token_area_filter,
        team_filter=selected_token_team_filter,
        last_submit_date_from=selected_token_last_submit_date_from,
        last_submit_date_to=selected_token_last_submit_date_to,
        sort_by=selected_token_sort_by,
        sort_dir=selected_token_sort_dir,
        rows_value="all",
    )
    masked_token_rows, selected_token_page, token_total_pages = paginate_token_rows(
        filtered_token_rows,
        rows_value=selected_token_rows,
        page=selected_token_page,
    )

    masked_usage_rows = []
    for row in usage_rows:
        if selected_usage_activity_filter == "active" and int(row["total_submit"] or 0) <= 0:
            continue
        masked_usage_rows.append({
            "kc_token": row["kc_token"],
            "kc_name": row["kc_name"],
            "team": row.get("team") or "-",
            "token_area": row["token_area"] or "-",
            "kc_username": row["kc_username"] or "-",
            "kc_password": row["kc_password"] or "-",
            "bearer_token_masked": mask_bearer_token(row["bearer_token"]),
            "daily_limit": row["daily_limit"],
            "is_active": row["is_active"],
            "total_submit": row["total_submit"],
        })

    submission_counts = get_submission_status_counts()
    return {
        "token_rows": masked_token_rows,
        "usage_rows": masked_usage_rows,
        "usage_date": usage_date,
        "total_tokens": total_tokens,
        "active_tokens": active_tokens,
        "total_submit_today": total_submit_today,
        "submission_counts": submission_counts,
        "recent_submissions": recent_submissions,
        "token_import_message": args.get("token_import_message", ""),
        "token_import_error": args.get("token_import_error", ""),
        "selected_token_filter": selected_token_filter,
        "selected_token_status_filter": selected_token_status_filter,
        "selected_token_area_filter": selected_token_area_filter,
        "selected_token_team_filter": selected_token_team_filter,
        "selected_token_last_submit_date_from": selected_token_last_submit_date_from,
        "selected_token_last_submit_date_to": selected_token_last_submit_date_to,
        "selected_token_rows": selected_token_rows,
        "selected_token_page": selected_token_page,
        "selected_token_sort_by": selected_token_sort_by,
        "selected_token_sort_dir": selected_token_sort_dir,
        "token_filtered_count": token_filtered_count,
        "token_visible_count": len(masked_token_rows),
        "token_total_pages": token_total_pages,
        "token_area_options": token_area_options,
        "token_team_options": token_team_options,
        "selected_usage_date_from": selected_usage_date_from,
        "selected_usage_date_to": selected_usage_date_to,
        "selected_usage_activity_filter": selected_usage_activity_filter,
        "token_filter_help_text": get_token_filter_help_text(),
    }


def build_team_leader_dashboard_context(args, allowed_kc_tokens=None, viewer_name="", viewer_role="admin"):
    token_rows = get_all_kc_tokens()
    allowed_token_set = None if allowed_kc_tokens is None else {str(token).strip() for token in allowed_kc_tokens if str(token).strip()}
    if allowed_token_set is not None:
        token_rows = [row for row in token_rows if row["kc_token"] in allowed_token_set]

    def parse_date_param(key):
        val = (args.get(key) or "").strip()
        if val:
            try:
                datetime.strptime(val, "%Y-%m-%d")
                return val
            except ValueError:
                pass
        return ""

    selected_date_from = parse_date_param("date_from")
    selected_date_to = parse_date_param("date_to")
    selected_filter = (args.get("token_filter") or "").strip()
    selected_status_filter = (args.get("token_status_filter") or "").strip().lower()
    selected_area_filter = (args.get("token_area_filter") or "").strip()
    selected_team_filter = (args.get("token_team_filter") or "").strip()
    if selected_status_filter not in ("", "aktif", "nonaktif"):
        selected_status_filter = ""
    selected_rows = normalize_token_rows_param(args.get("token_rows", "10"))
    selected_sort_by, selected_sort_dir = normalize_token_sort(
        args.get("token_sort_by", "total_submit"),
        args.get("token_sort_dir", "desc"),
    )

    usage_rows, usage_date = get_today_kc_usage_summary(
        date_from=selected_date_from or None,
        date_to=selected_date_to or None,
    )
    usage_by_token = {row["kc_token"]: row["total_submit"] for row in usage_rows}
    purchase_counts_by_token = get_kc_purchase_counts(
        date_from=selected_date_from,
        date_to=selected_date_to,
    )

    leader_rows = []
    for row in token_rows:
        purchase_counts = purchase_counts_by_token.get(row["kc_token"], {})
        total_submit = int(usage_by_token.get(row["kc_token"], 0) or 0)
        daily_limit = int(row["daily_limit"] or 0)
        leader_rows.append({
            "kc_token": row["kc_token"],
            "kc_name": row["kc_name"],
            "team": row.get("team") or "-",
            "token_area": row["token_area"] or "-",
            "kc_username": row["kc_username"] or "-",
            "daily_limit": daily_limit,
            "total_submit": total_submit,
            "purchase_yes": int(purchase_counts.get("purchase_yes", 0) or 0),
            "purchase_no": int(purchase_counts.get("purchase_no", 0) or 0),
            "lighter_yes": int(purchase_counts.get("lighter_yes", 0) or 0),
            "is_active": row["is_active"],
            "remaining_quota": max(0, daily_limit - total_submit),
        })

    token_area_options = sorted(
        {str((r.get("token_area") or "")).strip() for r in token_rows if str((r.get("token_area") or "")).strip()},
        key=str.lower,
    )
    token_team_options = sorted(
        {str((r.get("team") or "")).strip() for r in token_rows if str((r.get("team") or "")).strip()},
        key=str.lower,
    )

    filtered_rows_all, token_filtered_count = filter_sort_limit_token_rows(
        rows=leader_rows,
        filter_text=selected_filter,
        status_filter=selected_status_filter,
        area_filter=selected_area_filter,
        team_filter=selected_team_filter,
        sort_by=selected_sort_by,
        sort_dir=selected_sort_dir,
        rows_value="all",
    )
    visible_limit = token_rows_value_to_limit(selected_rows)
    visible_rows = filtered_rows_all if visible_limit is None else filtered_rows_all[:visible_limit]

    total_tokens = len(filtered_rows_all)
    active_tokens = len([r for r in filtered_rows_all if r["is_active"] == 1])
    total_submit = sum(r["total_submit"] for r in filtered_rows_all)
    total_purchase_yes = sum(r["purchase_yes"] for r in filtered_rows_all)
    total_purchase_no = sum(r["purchase_no"] for r in filtered_rows_all)
    total_lighter_yes = sum(r["lighter_yes"] for r in filtered_rows_all)
    total_remaining_quota = sum(r["remaining_quota"] for r in filtered_rows_all)
    low_quota_count = len([r for r in filtered_rows_all if r["remaining_quota"] <= 10])

    active_rows = [r for r in filtered_rows_all if r["is_active"] == 1]
    top_submitter = max(active_rows, key=lambda r: (r["total_submit"], r["purchase_yes"], r["kc_name"])) if active_rows else None
    top_buyer = max(active_rows, key=lambda r: (r["purchase_yes"], r["total_submit"], r["kc_name"])) if active_rows else None
    zero_submit_rows = [r for r in filtered_rows_all if r["total_submit"] == 0]
    low_quota_rows = sorted(
        [r for r in filtered_rows_all if r["remaining_quota"] <= 10],
        key=lambda r: (r["remaining_quota"], -r["total_submit"], str(r["kc_name"]).lower()),
    )[:5]
    inactive_rows = [r for r in filtered_rows_all if r["is_active"] != 1][:5]

    token_meta_by_token = {
        row["kc_token"]: {
            "team": row.get("team") or "-",
            "token_area": row.get("token_area") or "-",
        }
        for row in leader_rows
    }
    recent_submissions = get_recent_submission_attempts(
        limit=60,
        date_from=selected_date_from,
        date_to=selected_date_to,
        include_response_text=False,
    )
    visible_token_set = {row["kc_token"] for row in filtered_rows_all}
    filtered_recent_submissions = []
    for row in recent_submissions:
        meta = token_meta_by_token.get(row["kc_token"], {})
        row_team = meta.get("team", "-")
        row_area = meta.get("token_area", "-")
        if row["kc_token"] not in visible_token_set:
            continue
        if selected_team_filter and str(row_team).strip().lower() != selected_team_filter.strip().lower():
            continue
        if selected_area_filter and str(row_area).strip().lower() != selected_area_filter.strip().lower():
            continue
        filtered_recent_submissions.append({
            **row,
            "team": row_team,
            "token_area": row_area,
        })
    filtered_recent_submissions = filtered_recent_submissions[:10]

    return {
        "token_rows": visible_rows,
        "usage_date": usage_date,
        "total_tokens": total_tokens,
        "active_tokens": active_tokens,
        "total_submit": total_submit,
        "total_purchase_yes": total_purchase_yes,
        "total_purchase_no": total_purchase_no,
        "total_lighter_yes": total_lighter_yes,
        "total_remaining_quota": total_remaining_quota,
        "low_quota_count": low_quota_count,
        "recent_submissions": filtered_recent_submissions,
        "selected_token_filter": selected_filter,
        "selected_token_status_filter": selected_status_filter,
        "selected_token_area_filter": selected_area_filter,
        "selected_token_team_filter": selected_team_filter,
        "selected_token_rows": selected_rows,
        "selected_token_sort_by": selected_sort_by,
        "selected_token_sort_dir": selected_sort_dir,
        "token_filtered_count": token_filtered_count,
        "token_visible_count": len(visible_rows),
        "token_area_options": token_area_options,
        "token_team_options": token_team_options,
        "selected_date_from": selected_date_from,
        "selected_date_to": selected_date_to,
        "token_filter_help_text": get_token_filter_help_text(),
        "top_submitter": top_submitter,
        "top_buyer": top_buyer,
        "zero_submit_rows": zero_submit_rows[:5],
        "zero_submit_count": len(zero_submit_rows),
        "low_quota_rows": low_quota_rows,
        "low_quota_preview_count": len(low_quota_rows),
        "inactive_rows": inactive_rows,
        "inactive_count": len([r for r in filtered_rows_all if r["is_active"] != 1]),
        "viewer_name": viewer_name or ("Admin" if viewer_role == "admin" else "Team Leader"),
        "viewer_role": viewer_role,
        "allowed_kc_count": len(allowed_token_set) if allowed_token_set is not None else len(token_rows),
    }


def build_admin_submissions_context(args):
    selected_status = (args.get("status") or "").strip()
    selected_kc_token = (args.get("kc_token") or "").strip()
    selected_phone = (args.get("phone_number") or "").strip()
    selected_date_from = normalize_submission_date_filter(args.get("date_from")) or get_today_wib()
    selected_date_to = normalize_submission_date_filter(args.get("date_to")) or selected_date_from
    selected_limit_raw = (args.get("limit") or "100").strip()
    requested_limit = normalize_submission_log_limit(selected_limit_raw, default=100)
    display_limit = 50 if requested_limit is None else min(requested_limit, 50)

    rows = get_recent_submission_attempts(
        limit=display_limit,
        status_filter=selected_status,
        kc_token_filter=selected_kc_token,
        phone_filter=selected_phone,
        date_from=selected_date_from,
        date_to=selected_date_to,
        include_response_text=False,
    )
    counts = get_submission_status_counts(
        kc_token_filter=selected_kc_token,
        phone_filter=selected_phone,
        date_from=selected_date_from,
        date_to=selected_date_to,
    )

    direct_success_count = 0
    retried_success_count = 0
    retried_failed_count = 0
    retried_likely_success_count = 0
    invalid_count = 0

    for row in rows:
        attempt_count = int(row.get("attempt_count") or 0)
        row["is_retried"] = attempt_count > 1
        if row["status_local"] == "SUCCESS":
            if row["is_retried"]:
                retried_success_count += 1
                row["retry_label"] = "Retried Success"
            else:
                direct_success_count += 1
                row["retry_label"] = "Direct Success"
        elif row["status_local"] == "LIKELY_SUCCESS":
            if row["is_retried"]:
                retried_likely_success_count += 1
                row["retry_label"] = "Retried Likely"
            else:
                row["retry_label"] = "Likely Success"
        elif row["status_local"] == "FAILED":
            if row["is_retried"]:
                retried_failed_count += 1
                row["retry_label"] = "Retried Failed"
            else:
                row["retry_label"] = "Failed"
        elif row["status_local"] == "INVALID":
            invalid_count += 1
            row["retry_label"] = "Invalid"
        else:
            row["retry_label"] = "Pending"

    total_rows = len(rows)
    retried_total = sum(1 for row in rows if row.get("is_retried"))
    retry_rate = round((retried_total / total_rows) * 100, 1) if total_rows else 0

    retry_stats = {
        "direct_success_count": direct_success_count,
        "retried_success_count": retried_success_count,
        "retried_likely_success_count": retried_likely_success_count,
        "retried_failed_count": retried_failed_count,
        "invalid_count": invalid_count,
        "retried_total": retried_total,
        "retry_rate": retry_rate,
    }

    return {
        "rows": rows,
        "counts": counts,
        "retry_stats": retry_stats,
        "selected_status": selected_status,
        "selected_kc_token": selected_kc_token,
        "selected_phone": selected_phone,
        "selected_date_from": selected_date_from,
        "selected_date_to": selected_date_to,
        "selected_limit": "all" if requested_limit is None else str(requested_limit),
        "limit_options": SUBMISSION_LOG_LIMIT_OPTIONS,
    }


def get_token_filter_help_text():
    return "Bisa cari beberapa nama sekaligus, pisahkan dengan koma atau baris baru. Contoh: Ariska, Cici, Fahmi"


def clear_team_leader_session():
    session.pop("is_team_leader_logged_in", None)
    session.pop("team_leader_username", None)
    session.pop("team_leader_name", None)


def get_team_leader(username):
    current = str(username or "").strip()
    if not current:
        return None
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT username, password_hash, leader_name, is_active, created_at, updated_at
        FROM team_leaders
        WHERE username = %s
        """,
        (current,),
    )
    row = cur.fetchone()
    conn.close()
    return row


def get_team_leader_access_tokens(username):
    current = str(username or "").strip()
    if not current:
        return []
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT a.kc_token
        FROM team_leader_kc_access a
        WHERE a.leader_username = %s
        ORDER BY a.kc_token ASC
        """,
        (current,),
    )
    rows = cur.fetchall()
    conn.close()
    return [row["kc_token"] for row in rows]


def get_all_team_leaders():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT tl.username, tl.leader_name, tl.is_active, tl.created_at, tl.updated_at,
               COALESCE(COUNT(a.kc_token), 0) AS kc_count
        FROM team_leaders tl
        LEFT JOIN team_leader_kc_access a ON a.leader_username = tl.username
        GROUP BY tl.username, tl.leader_name, tl.is_active, tl.created_at, tl.updated_at
        ORDER BY tl.leader_name ASC, tl.username ASC
        """
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def get_team_leader_access_rows():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT tl.username, tl.leader_name, tl.is_active, a.kc_token,
               COALESCE(v.kc_name, '') AS kc_name,
               COALESCE(v.team, '') AS team,
               COALESCE(v.token_area, '') AS token_area
        FROM team_leaders tl
        LEFT JOIN team_leader_kc_access a ON a.leader_username = tl.username
        LEFT JOIN valid_kc_tokens v ON v.kc_token = a.kc_token
        ORDER BY tl.leader_name ASC, tl.username ASC, a.kc_token ASC
        """
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def normalize_kc_token_list(raw_value):
    if isinstance(raw_value, list):
        values = raw_value
    else:
        values = re.split(r"[\r\n,;|]+", str(raw_value or ""))
    unique_tokens = []
    seen = set()
    for value in values:
        token = str(value or "").strip()
        if not token:
            continue
        token_key = token.lower()
        if token_key in seen:
            continue
        seen.add(token_key)
        unique_tokens.append(token)
    return unique_tokens


def save_team_leader(username, password, leader_name, is_active, allowed_kc_tokens):
    current_username = str(username or "").strip()
    current_leader_name = str(leader_name or "").strip()
    if not current_username:
        raise ValueError("Username TL wajib diisi.")
    if not current_leader_name:
        raise ValueError("Nama TL wajib diisi.")

    current_password = str(password or "")
    existing = get_team_leader(current_username)
    if not existing and not current_password:
        raise ValueError("Password TL wajib diisi saat membuat akun baru.")

    normalized_tokens = normalize_kc_token_list(allowed_kc_tokens)
    valid_kc_tokens = {row["kc_token"] for row in get_all_kc_tokens()}
    invalid_tokens = [token for token in normalized_tokens if token not in valid_kc_tokens]
    if invalid_tokens:
        raise ValueError("KC token tidak ditemukan: " + ", ".join(invalid_tokens))

    now_str = get_now_db_string()
    conn = get_db_connection()
    cur = conn.cursor()
    if existing:
        if current_password:
            cur.execute(
                """
                UPDATE team_leaders
                SET password_hash = %s,
                    leader_name = %s,
                    is_active = %s,
                    updated_at = %s
                WHERE username = %s
                """,
                (generate_password_hash(current_password), current_leader_name, int(bool(is_active)), now_str, current_username),
            )
        else:
            cur.execute(
                """
                UPDATE team_leaders
                SET leader_name = %s,
                    is_active = %s,
                    updated_at = %s
                WHERE username = %s
                """,
                (current_leader_name, int(bool(is_active)), now_str, current_username),
            )
    else:
        cur.execute(
            """
            INSERT INTO team_leaders (username, password_hash, leader_name, is_active, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (current_username, generate_password_hash(current_password), current_leader_name, int(bool(is_active)), now_str, now_str),
        )

    cur.execute("DELETE FROM team_leader_kc_access WHERE leader_username = %s", (current_username,))
    for token in normalized_tokens:
        cur.execute(
            """
            INSERT INTO team_leader_kc_access (leader_username, kc_token, created_at)
            VALUES (%s, %s, %s)
            ON CONFLICT (leader_username, kc_token) DO NOTHING
            """,
            (current_username, token, now_str),
        )
    conn.commit()
    conn.close()
    return current_username


def delete_team_leader(username):
    current = str(username or "").strip()
    if not current:
        raise ValueError("Username TL tidak valid.")
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM team_leader_kc_access WHERE leader_username = %s", (current,))
    cur.execute("DELETE FROM team_leaders WHERE username = %s", (current,))
    deleted = cur.rowcount
    conn.commit()
    conn.close()
    return deleted > 0


@app.route("/admin")
@admin_required
def admin_dashboard():
    return render_template("admin_dashboard.html", **build_admin_dashboard_context(request.args))


@app.route("/admin/data")
@admin_required
def admin_dashboard_data():
    return jsonify(build_admin_dashboard_context(request.args))


@app.route("/leader")
@leader_required
def team_leader_dashboard():
    if is_admin_logged_in():
        context = build_team_leader_dashboard_context(
            request.args,
            allowed_kc_tokens=None,
            viewer_name=session.get("admin_page_username") or "Admin",
            viewer_role="admin",
        )
    else:
        leader_username = session.get("team_leader_username") or ""
        context = build_team_leader_dashboard_context(
            request.args,
            allowed_kc_tokens=get_team_leader_access_tokens(leader_username),
            viewer_name=session.get("team_leader_name") or leader_username,
            viewer_role="leader",
        )
    return render_template("team_leader_dashboard.html", **context)


@app.route("/admin/team-leaders", methods=["GET", "POST"])
@admin_required
def admin_team_leaders():
    error = ""
    success = ""
    if request.method == "POST":
        try:
            username = (request.form.get("username") or "").strip()
            password = request.form.get("password") or ""
            leader_name = (request.form.get("leader_name") or "").strip()
            is_active = (request.form.get("is_active") or "1").strip() == "1"
            allowed_kc_tokens = request.form.get("allowed_kc_tokens") or ""
            saved_username = save_team_leader(username, password, leader_name, is_active, allowed_kc_tokens)
            success = f"Akun Team Leader {saved_username} berhasil disimpan."
        except Exception as e:
            error = str(e)

    access_rows = get_team_leader_access_rows()
    grouped_access = {}
    for row in access_rows:
        key = row["username"]
        item = grouped_access.setdefault(key, {
            "username": row["username"],
            "leader_name": row["leader_name"] or row["username"],
            "is_active": row["is_active"],
            "kc_items": [],
        })
        if row.get("kc_token"):
            item["kc_items"].append({
                "kc_token": row["kc_token"],
                "kc_name": row.get("kc_name") or "-",
                "team": row.get("team") or "-",
                "token_area": row.get("token_area") or "-",
            })

    return render_template(
        "admin_team_leaders.html",
        error=error,
        success=success,
        leaders=get_all_team_leaders(),
        leader_access=list(grouped_access.values()),
        kc_tokens=get_all_kc_tokens(),
    )


@app.route("/admin/team-leaders/<username>/delete", methods=["POST"])
@admin_required
def admin_delete_team_leader(username):
    try:
        deleted = delete_team_leader(username)
        if not deleted:
            raise ValueError("Akun Team Leader tidak ditemukan.")
        return redirect(url_for("admin_team_leaders"))
    except Exception:
        return redirect(url_for("admin_team_leaders"))


@app.route("/admin/token/add", methods=["GET", "POST"])
@admin_required
def admin_add_token():
    error = None

    if request.method == "POST":
        kc_token = request.form.get("kc_token", "").strip()
        kc_name = request.form.get("kc_name", "").strip()
        team = request.form.get("team", "").strip()
        token_area = request.form.get("token_area", "").strip()
        kc_username = request.form.get("kc_username", "").strip()
        kc_password = request.form.get("kc_password", "").strip()
        bearer_token = request.form.get("bearer_token", "").strip()
        daily_limit = request.form.get("daily_limit", "").strip()

        try:
            if not kc_token:
                raise ValueError("KC token wajib diisi.")
            if not kc_name:
                raise ValueError("Nama KC wajib diisi.")
            if not bearer_token:
                raise ValueError("Bearer token wajib diisi.")
            if not daily_limit:
                raise ValueError("Daily limit wajib diisi.")

            daily_limit = int(daily_limit)
            if daily_limit < 1:
                raise ValueError("Daily limit minimal 1.")

            existing = get_kc_token_detail(kc_token)
            if existing:
                raise ValueError("KC token sudah ada.")

            create_kc_token(
                kc_token,
                kc_name,
                token_area,
                bearer_token,
                daily_limit,
                kc_username=kc_username,
                kc_password=kc_password,
                team=team,
            )
            if wants_json_response():
                return jsonify({
                    "success": True,
                    "message": "KC token berhasil disimpan.",
                    "redirect_url": url_for("admin_dashboard"),
                    "kc_token": kc_token,
                })
            return redirect(url_for("admin_dashboard"))

        except Exception as e:
            logger.exception("submit route error")
            error = str(e)
            token_data = {
                "kc_token": kc_token,
                "kc_name": kc_name,
                "team": team,
                "token_area": token_area,
                "kc_username": kc_username,
                "kc_password": kc_password,
                "bearer_token": bearer_token,
                "daily_limit": daily_limit if daily_limit else 40,
            }
            if wants_json_response():
                return jsonify({"success": False, "error": error, "token_data": token_data}), 400
            return render_template("admin_token_form.html", error=error, mode="add", token_data=token_data)

    return render_template("admin_token_form.html", error=error, mode="add", token_data=None)


@app.route("/admin/token/<path:kc_token>/edit", methods=["GET", "POST"])
@admin_required
def admin_edit_token(kc_token):
    token_data = get_kc_token_detail(kc_token)
    if not token_data:
        return redirect(url_for("admin_dashboard"))

    error = None

    if request.method == "POST":
        new_kc_token = request.form.get("kc_token", "").strip()
        kc_name = request.form.get("kc_name", "").strip()
        team = request.form.get("team", "").strip()
        token_area = request.form.get("token_area", "").strip()
        kc_username = request.form.get("kc_username", "").strip()
        kc_password = request.form.get("kc_password", "").strip()
        bearer_token = request.form.get("bearer_token", "").strip()
        daily_limit = request.form.get("daily_limit", "").strip()

        try:
            if not new_kc_token:
                raise ValueError("KC token wajib diisi.")
            if not kc_name:
                raise ValueError("Nama KC wajib diisi.")
            if not bearer_token:
                raise ValueError("Bearer token wajib diisi.")
            if not daily_limit:
                raise ValueError("Daily limit wajib diisi.")

            daily_limit = int(daily_limit)
            if daily_limit < 1:
                raise ValueError("Daily limit minimal 1.")

            if new_kc_token != kc_token:
                existing = get_kc_token_detail(new_kc_token)
                if existing:
                    raise ValueError("KC token baru sudah digunakan.")

            update_kc_token(
                kc_token,
                new_kc_token,
                kc_name,
                bearer_token,
                daily_limit,
                team=team,
                token_area=token_area,
                kc_username=kc_username,
                kc_password=kc_password,
            )
            if wants_json_response():
                return jsonify({
                    "success": True,
                    "message": "KC token berhasil diperbarui.",
                    "redirect_url": url_for("admin_dashboard"),
                    "kc_token": new_kc_token,
                })
            return redirect(url_for("admin_dashboard"))

        except Exception as e:
            logger.exception("submit route error")
            error = str(e)
            token_data = {
                "kc_token": new_kc_token,
                "kc_name": kc_name,
                "team": team,
                "token_area": token_area,
                "kc_username": kc_username,
                "kc_password": kc_password,
                "bearer_token": bearer_token,
                "daily_limit": daily_limit,
                "is_active": token_data["is_active"],
            }
            if wants_json_response():
                return jsonify({"success": False, "error": error, "token_data": token_data}), 400

    return render_template("admin_token_form.html", error=error, mode="edit", token_data=token_data)


def refresh_kc_bearer_token_with_retry(kc_token, max_attempts=3):
    kc_detail = get_kc_token_detail(kc_token)
    if not kc_detail:
        raise ValueError("KC token tidak ditemukan.")

    username = (kc_detail.get("kc_username") or "").strip()
    password = (kc_detail.get("kc_password") or "").strip()
    if not username or not password:
        raise ValueError("Username atau password belum diisi di data KC token ini.")

    secret = os.environ.get("APP_HMAC_SECRET", "").strip()
    if not secret:
        raise ValueError("APP_HMAC_SECRET belum di-set.")

    captcha_key = os.environ.get("TWOCAPTCHA_API_KEY", "").strip()
    if not captcha_key:
        raise ValueError("TWOCAPTCHA_API_KEY belum di-set.")

    attempts = max(1, int(max_attempts or 1))
    last_error = ""
    for attempt in range(1, attempts + 1):
        try:
            recaptcha_token = _sl_solve_recaptcha(captcha_key)
            ts = _sl_utc_timestamp_ms()
            body_obj = {
                "identifier": username,
                "password": base64.b64encode(password.encode()).decode(),
                "type_of_web": _SL_DEFAULT_TYPE_OF_WEB,
                "recaptcha_token": recaptcha_token,
            }
            hash_val = _sl_build_hash(secret, ts, _SL_AUTH_ENDPOINT, body_obj)
            headers = {
                "Accept": "application/json, text/plain, */*",
                "Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
                "Content-Type": "application/json",
                "Dnt": "1",
                "Hash": hash_val,
                "Origin": "https://letscml.id",
                "Referer": "https://letscml.id/",
                "Timestamp": ts,
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36",
            }
            resp = requests.post(_SL_BASE_URL + _SL_AUTH_ENDPOINT, json=body_obj, headers=headers, timeout=45)
            token, token_source = _sl_extract_token(resp)
            if not token:
                try:
                    preview = str(resp.json())[:200]
                except Exception:
                    preview = resp.text[:200]
                raise ValueError(f"Token tidak ditemukan. Status: {resp.status_code}. {preview}")

            update_kc_token(
                old_kc_token=kc_token,
                new_kc_token=kc_detail["kc_token"],
                kc_name=kc_detail["kc_name"],
                bearer_token=token,
                daily_limit=kc_detail["daily_limit"],
                team=kc_detail.get("team", ""),
                token_area=kc_detail.get("token_area", ""),
                kc_username=username,
                kc_password=password,
            )
            return {"success": True, "token_source": token_source, "attempts": attempt}
        except requests.exceptions.Timeout:
            last_error = "Timeout (45s)."
        except Exception as e:
            last_error = str(e)
        logger.warning("Refresh bearer token failed kc_token=%s attempt=%s/%s error=%s", kc_token, attempt, attempts, last_error)

    raise ValueError(last_error or "Gagal refresh bearer token.")


@app.route("/admin/token/<path:kc_token>/refresh-bearer", methods=["POST"])
@admin_required
def admin_refresh_bearer_token(kc_token):
    try:
        result = refresh_kc_bearer_token_with_retry(kc_token, max_attempts=1)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    logger.info(f"Bearer token refreshed for KC {kc_token} via refresh-bearer route (source: {result['token_source']})")
    return jsonify({"success": True, "token_source": result["token_source"], "attempts": result["attempts"]})


@app.route("/admin/token/import", methods=["POST"])
@admin_required
def admin_import_tokens():
    try:
        token_file = request.files.get("token_file")
        if not token_file or not token_file.filename:
            raise ValueError("File import wajib dipilih.")

        result = import_kc_tokens(token_file)
        message = (
            f"Import token selesai. Baru: {result['inserted']} | "
            f"Update: {result['updated']} | Invalid: {result['invalid']} | "
            f"Kosong dilewati: {result['skipped']}"
        )
        if result["sample_errors"]:
            message += f" | Contoh error: {'; '.join(result['sample_errors'])}"
        return redirect(url_for("admin_dashboard", token_import_message=message))
    except Exception as e:
        logger.exception("admin_import_tokens error")
        return redirect(url_for("admin_dashboard", token_import_error=str(e)))


@app.route("/admin/batch-bearer")
@admin_required
def admin_batch_bearer_page():
    export_id = request.args.get("export_id", "")
    return render_template(
        "admin_batch_bearer.html",
        message=request.args.get("message", ""),
        error=request.args.get("error", ""),
        export_id=export_id,
        result_rows=get_batch_bearer_rows(export_id),
    )


@app.route("/admin/token/batch-bearer", methods=["POST"])
@admin_required
def admin_batch_bearer_tokens():
    try:
        batch_file = request.files.get("batch_bearer_file")
        if not batch_file or not batch_file.filename:
            raise ValueError("File batch bearer wajib dipilih.")

        result = batch_import_bearer_tokens(batch_file)
        message = (
            f"Batch ambil bearer selesai. Total: {result['total']} | "
            f"Baru: {result['inserted']} | Update: {result['updated']} | "
            f"Gagal: {result['failed']} | Kosong dilewati: {result['skipped']}"
        )
        if result["sample_errors"]:
            message += f" | Contoh error: {'; '.join(result['sample_errors'])}"
        export_id = save_batch_bearer_export(result["rows"])
        return redirect(url_for("admin_batch_bearer_page", message=message, export_id=export_id))
    except Exception as e:
        logger.exception("admin_batch_bearer_tokens error")
        return redirect(url_for("admin_batch_bearer_page", error=str(e)))


@app.route("/admin/token/batch-bearer/export/<export_id>")
@admin_required
def admin_export_batch_bearer_tokens(export_id):
    export_path = get_batch_bearer_export_path(export_id)
    if not export_path or not os.path.exists(export_path):
        return redirect(url_for("admin_batch_bearer_page", error="File export hasil batch tidak ditemukan atau sudah dibersihkan."))

    with open(export_path, "rb") as fh:
        excel_content = fh.read()
    today = get_today_wib()
    return Response(
        excel_content,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="hasil-batch-bearer-{today}.xlsx"'},
    )


@app.route("/admin/master-data", methods=["GET", "POST"])
@admin_required
def admin_master_data():
    error = request.args.get("error", "")
    success = request.args.get("success", "")

    if request.method == "POST":
        try:
            master_type = request.form.get("master_type", "").strip()
            master_file = request.files.get("master_file")
            if not master_file or not master_file.filename:
                raise ValueError("File import wajib dipilih.")

            result = import_master_data(master_file, master_type)
            label = "BUMO" if master_type == "bumo" else "KC Area"
            success = (
                f"Import {label} selesai. Baru: {result['inserted']} | "
                f"Update: {result['updated']} | Invalid: {result['invalid']} | "
                f"Kosong dilewati: {result['skipped']}"
            )
            if result["sample_errors"]:
                success += f" | Contoh error: {'; '.join(result['sample_errors'])}"
        except Exception as e:
            logger.exception("admin_master_data import error")
            error = str(e)

    return render_template(
        "admin_master_data.html",
        error=error,
        success=success,
        counts=get_master_data_counts(),
    )


@app.route("/admin/master-data/sync", methods=["POST"])
@admin_required
def admin_sync_master_data():
    try:
        bearer_token = request.form.get("bearer_token", "").strip()
        if not bearer_token:
            raise ValueError("Bearer token manual wajib diisi untuk load BUMO/KC Area.")

        result = sync_master_data_from_api(bearer_token)
        message = (
            "Load BUMO/KC Area dari API selesai. "
            f"Token sumber: {result['kc_name']} ({result['kc_token']}) | "
            f"Percobaan token: {result['tried_count']} | "
            f"BUMO baru: {result['bumo_inserted']} | BUMO update: {result['bumo_updated']} | "
            f"KC Area baru: {result['kc_area_inserted']} | KC Area update: {result['kc_area_updated']}"
        )
        return redirect(url_for("admin_master_data", success=message))
    except Exception as e:
        logger.exception("admin_sync_master_data error")
        return redirect(url_for("admin_master_data", error=str(e)))


@app.route("/admin/token/export")
@admin_required
def admin_export_tokens():
    csv_content, usage_date = build_kc_token_export_csv()
    filename = f"kc-token-detail-{usage_date}.csv"
    return Response(
        csv_content,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.route("/admin/tokens/export-excel")
@admin_required
def admin_export_tokens_excel():
    excel_content, _export_count = build_full_kc_token_export_excel(request.args)
    today = get_today_wib()
    filename = f"daftar-kc-token-{today}.xlsx"
    return Response(
        excel_content,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.route("/admin/tokens/export-selected", methods=["POST"])
@admin_required
def admin_export_selected_tokens():
    payload = request.get_json(silent=True) or {}
    kc_tokens = normalize_kc_token_selection(payload.get("kc_tokens") or [])
    export_format = str(payload.get("format") or "txt").strip().lower()
    if not kc_tokens:
        return jsonify({"success": False, "error": "Pilih minimal satu akun."}), 400

    today = get_today_wib()
    if export_format == "xlsx":
        excel_content, export_count = build_selected_kc_token_export_excel(kc_tokens)
        if export_count == 0:
            return jsonify({"success": False, "error": "Akun terpilih tidak ditemukan."}), 404
        filename = f"kc-token-terpilih-{today}.xlsx"
        return Response(
            excel_content,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    text_content, export_count = build_selected_kc_token_export_text(kc_tokens)
    if export_count == 0:
        return jsonify({"success": False, "error": "Akun terpilih tidak ditemukan."}), 404
    filename = f"kc-token-terpilih-{today}.txt"
    return Response(
        "\ufeff" + text_content,
        mimetype="text/plain; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.route("/admin/usage/export")
@admin_required
def admin_export_usage():
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    active_only = (request.args.get("active_only") or "").strip().lower() in {"1", "true", "yes", "active"}
    excel_content = build_kc_usage_export_excel(date_from=date_from, date_to=date_to, active_only=active_only)
    today = get_today_wib()
    filename = f"usage-kc-token-{today}.xlsx"
    return Response(
        excel_content,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.route("/admin/token/<path:kc_token>/toggle", methods=["POST"])
@admin_required
def admin_toggle_token(kc_token):
    token_data = get_kc_token_detail(kc_token)
    if token_data:
        toggle_kc_token_status(kc_token)

        current_user_token = session.get("kc_token")
        if current_user_token == kc_token:
            clear_user_session()
        if wants_json_response():
            updated = get_kc_token_detail(kc_token)
            return jsonify({
                "success": True,
                "kc_token": kc_token,
                "is_active": updated["is_active"] if updated else 0,
            })

    if wants_json_response():
        return jsonify({"success": False, "error": "KC token tidak ditemukan."}), 404
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/tokens/deactivate-all", methods=["POST"])
@admin_required
def admin_deactivate_all_tokens():
    deactivated_count = deactivate_all_kc_tokens()
    if session.get("kc_token"):
        clear_user_session()
    if wants_json_response():
        return jsonify({"success": True, "deactivated_count": deactivated_count})
    return redirect(request.referrer or url_for("admin_dashboard"))


@app.route("/admin/tokens/activate-all", methods=["POST"])
@admin_required
def admin_activate_all_tokens():
    activated_count = activate_all_kc_tokens()
    if wants_json_response():
        return jsonify({"success": True, "activated_count": activated_count})
    return redirect(request.referrer or url_for("admin_dashboard"))


@app.route("/admin/tokens/bulk-action", methods=["POST"])
@admin_required
def admin_bulk_token_action():
    payload = request.get_json(silent=True) or {}
    action = str(payload.get("action") or "").strip()
    kc_tokens = normalize_kc_token_selection(payload.get("kc_tokens") or [])
    if not kc_tokens:
        return jsonify({"success": False, "error": "Pilih minimal satu akun."}), 400

    current_user_token = (session.get("kc_token") or "").strip()
    try:
        if action == "activate":
            updated_count = set_kc_tokens_status(kc_tokens, 1)
            return jsonify({"success": True, "action": action, "updated_count": updated_count})

        if action == "deactivate":
            updated_count = set_kc_tokens_status(kc_tokens, 0)
            if current_user_token in kc_tokens:
                clear_user_session()
            return jsonify({"success": True, "action": action, "updated_count": updated_count})

        if action == "reset_usage_today":
            reset_count, target_date = reset_selected_kc_token_usage_today(kc_tokens)
            if current_user_token in kc_tokens:
                session["used_today"] = 0
            return jsonify({
                "success": True,
                "action": action,
                "reset_count": reset_count,
                "usage_date": target_date,
            })

        if action == "delete":
            deleted_count = delete_kc_tokens(kc_tokens)
            if current_user_token in kc_tokens:
                clear_user_session()
            return jsonify({"success": True, "action": action, "deleted_count": deleted_count})

        if action == "refresh_bearer":
            successes = []
            failures = []
            for token in kc_tokens:
                try:
                    result = refresh_kc_bearer_token_with_retry(token, max_attempts=3)
                    successes.append({
                        "kc_token": token,
                        "token_source": result["token_source"],
                        "attempts": result["attempts"],
                    })
                except Exception as e:
                    failures.append({"kc_token": token, "error": str(e)})
            return jsonify({
                "success": True,
                "action": action,
                "success_count": len(successes),
                "failure_count": len(failures),
                "successes": successes[:10],
                "failures": failures[:10],
            })

        return jsonify({"success": False, "error": "Aksi massal tidak valid."}), 400
    except Exception as e:
        logger.exception("admin_bulk_token_action error")
        return jsonify({"success": False, "error": str(e)}), 400


@app.route("/admin/token/<path:kc_token>/delete", methods=["POST"])
@admin_required
def admin_delete_token_route(kc_token):
    token_data = get_kc_token_detail(kc_token)
    if token_data:
        current_user_token = session.get("kc_token")
        delete_kc_token(kc_token)

        if current_user_token == kc_token:
            clear_user_session()
        if wants_json_response():
            return jsonify({"success": True, "kc_token": kc_token})

    if wants_json_response():
        return jsonify({"success": False, "error": "KC token tidak ditemukan."}), 404
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/token/<path:kc_token>/reset-usage-today", methods=["POST"])
@admin_required
def admin_reset_token_usage_today(kc_token):
    token_data = get_kc_token_detail(kc_token)
    if token_data:
        target_date = reset_kc_token_usage_today(kc_token)
        current_user_token = (session.get("kc_token") or "").strip()
        if current_user_token == kc_token:
            session["used_today"] = 0
        if wants_json_response():
            return jsonify({"success": True, "kc_token": kc_token, "usage_date": target_date})
    if wants_json_response():
        return jsonify({"success": False, "error": "KC token tidak ditemukan."}), 404
    return redirect(request.referrer or url_for("admin_dashboard"))


@app.route("/admin/token/reset-usage-today", methods=["POST"])
@admin_required
def admin_reset_all_token_usage_today():
    reset_count, target_date = reset_all_kc_token_usage_today()
    current_user_token = (session.get("kc_token") or "").strip()
    if current_user_token:
        session["used_today"] = 0
    if wants_json_response():
        return jsonify({"success": True, "reset_count": reset_count, "usage_date": target_date})
    return redirect(request.referrer or url_for("admin_dashboard"))


@app.route("/admin/customers")
@admin_required
def admin_customers():
    selected_row_limit = normalize_customer_rows_param(request.args.get("rows", "10"))
    selected_sort_by, selected_sort_dir = normalize_customer_sort(
        request.args.get("sort_by", "reserved_at"),
        request.args.get("sort_dir", "desc"),
    )

    stats = get_customer_stats()
    return render_template(
        "admin_customer_db.html",
        selected_row_limit=selected_row_limit,
        selected_sort_by=selected_sort_by,
        selected_sort_dir=selected_sort_dir,
        **stats,
    )


@app.route("/admin/customers/data")
@admin_required
def admin_customers_data():
    rows_value = normalize_customer_rows_param(request.args.get("rows", "10"))
    sort_by, sort_dir = normalize_customer_sort(
        request.args.get("sort_by", "reserved_at"),
        request.args.get("sort_dir", "desc"),
    )
    limit = customer_rows_value_to_limit(rows_value)
    customer_rows = [
        serialize_customer_row(row)
        for row in get_all_customer_numbers(limit=limit, sort_by=sort_by, sort_dir=sort_dir)
    ]
    stats = get_customer_stats()
    return jsonify({
        "rows": customer_rows,
        "stats": stats,
        "rows_value": rows_value,
        "sort_by": sort_by,
        "sort_dir": sort_dir,
    })


@app.route("/admin/customers/item")
@admin_required
def admin_customers_item():
    phone_number = request.args.get("phone_number", "")
    row = get_customer_number(phone_number)
    if not row:
        return jsonify({"error": "Nomor tidak ditemukan."}), 404
    return jsonify({"item": serialize_customer_row(row)})


@app.route("/admin/customers/save", methods=["POST"])
@admin_required
def admin_customers_save():
    try:
        normalized = upsert_customer_number(
            phone_number=request.form.get("phone_number", ""),
            is_active=1 if request.form.get("is_active") == "1" else 0,
            old_phone_number=request.form.get("old_phone_number", ""),
        )
        stats = get_customer_stats()
        return jsonify({"ok": True, "message": f"Nomor {normalized} berhasil disimpan.", "stats": stats})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/admin/customers/delete", methods=["POST"])
@admin_required
def admin_customers_delete():
    try:
        delete_customer_number(request.form.get("phone_number", ""))
        stats = get_customer_stats()
        return jsonify({"ok": True, "message": "Nomor berhasil dihapus.", "stats": stats})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/admin/customers/reset", methods=["POST"])
@admin_required
def admin_customers_reset():
    try:
        reset_customer_status(request.form.get("phone_number", ""))
        stats = get_customer_stats()
        return jsonify({"ok": True, "message": "Status nomor berhasil direset.", "stats": stats})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/admin/customers/import", methods=["POST"])
@admin_required
def admin_customers_import():
    try:
        excel_file = request.files.get("excel_file")
        if not excel_file or not excel_file.filename:
            raise ValueError("File import wajib dipilih.")
        result = import_customer_numbers(excel_file, is_active=1 if request.form.get("is_active") == "1" else 0)
        stats = get_customer_stats()
        message = (
            f"Import selesai. Masuk: {result['inserted']} nomor | "
            f"Duplikat: {result['duplicates']} | Invalid: {result['invalid']}"
        )
        return jsonify({"ok": True, "message": message, "stats": stats})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/admin/customers/export")
@admin_required
def admin_customers_export():
    sort_by, sort_dir = normalize_customer_sort(
        request.args.get("sort_by", "reserved_at"),
        request.args.get("sort_dir", "desc"),
    )

    excel_content = build_customer_numbers_export_excel(sort_by=sort_by, sort_dir=sort_dir)
    filename = f"customer_database_{get_now_wib().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        excel_content,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.route("/admin/customers/reshuffle", methods=["POST"])
@admin_required
def admin_customers_reshuffle():
    try:
        total = reshuffle_ready_customer_numbers()
        stats = get_customer_stats()
        return jsonify({"ok": True, "message": f"Reshuffle selesai untuk {total} nomor siap pakai.", "stats": stats})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/admin/customers/reset-distribution", methods=["POST"])
@admin_required
def admin_customers_reset_distribution():
    try:
        total = reset_customer_distribution()
        stats = get_customer_stats()
        return jsonify({"ok": True, "message": f"Reset distribusi selesai. {total} nomor siap pakai diacak ulang.", "stats": stats})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/admin/customers/delete-unused", methods=["POST"])
@admin_required
def admin_customers_delete_unused():
    try:
        total = delete_unused_customer_numbers()
        stats = get_customer_stats()
        return jsonify({"ok": True, "message": f"Hapus database selesai. {total} nomor belum terpakai berhasil dihapus.", "stats": stats})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/admin/customers/delete-used", methods=["POST"])
@admin_required
def admin_customers_delete_used():
    try:
        total = delete_used_customer_numbers()
        stats = get_customer_stats()
        return jsonify({"ok": True, "message": f"Hapus database selesai. {total} nomor sudah terpakai berhasil dihapus.", "stats": stats})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/admin/submissions")
@admin_required
def admin_submissions():
    return render_template("admin_submit_logs.html", **build_admin_submissions_context(request.args))


@app.route("/admin/submissions/data")
@admin_required
def admin_submissions_data():
    return jsonify(build_admin_submissions_context(request.args))


@app.route("/admin/submissions/<submission_id>/response")
@admin_required
def admin_submission_response_detail(submission_id):
    row = get_submission_response_detail(submission_id)
    if not row:
        return jsonify({"error": "Data submit tidak ditemukan."}), 404
    return jsonify({
        "submission_id": row["submission_id"],
        "final_response_text": row["final_response_text"] or "",
    })


@app.route("/admin/submissions/delete", methods=["POST"])
@admin_required
def admin_submissions_delete():
    payload = request.get_json(silent=True) or {}
    confirm_text = str(payload.get("confirm_text") or "").strip()
    if confirm_text != "HAPUS":
        return jsonify({"success": False, "error": "Ketik HAPUS untuk konfirmasi hapus riwayat submit."}), 400

    deleted_count = delete_submission_attempts_by_filter(
        status_filter=payload.get("status") or "",
        kc_token_filter=payload.get("kc_token") or "",
        phone_filter=payload.get("phone_number") or "",
        date_from=payload.get("date_from") or "",
        date_to=payload.get("date_to") or "",
    )
    return jsonify({"success": True, "deleted_count": deleted_count})


@app.route("/admin/submissions/export")
@admin_required
def admin_submissions_export():
    selected_status = (request.args.get("status") or "").strip()
    selected_kc_token = (request.args.get("kc_token") or "").strip()
    selected_phone = (request.args.get("phone_number") or "").strip()
    selected_date_from = normalize_submission_date_filter(request.args.get("date_from"))
    selected_date_to = normalize_submission_date_filter(request.args.get("date_to"))
    limit = normalize_submission_log_limit(request.args.get("limit", "10000"), default=10000)

    csv_content = build_submission_attempts_export_csv(
        limit=limit,
        status_filter=selected_status,
        kc_token_filter=selected_kc_token,
        phone_filter=selected_phone,
        date_from=selected_date_from,
        date_to=selected_date_to,
    )
    filename = f"submission_attempts_{get_now_wib().strftime('%Y%m%d_%H%M%S')}.csv"
    return Response(
        csv_content,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

# ── Single Login ──────────────────────────────────────────────────────────────

_SL_BASE_URL = "https://ca-msfsax05-22-be-letscml-prd.mangosmoke-bb4ae1b7.southeastasia.azurecontainerapps.io"
_SL_AUTH_ENDPOINT = "/api/auth/local"
_SL_DEFAULT_TYPE_OF_WEB = "spgCMKT"


def _sl_utc_timestamp_ms() -> str:
    now = datetime.now(timezone.utc)
    ms = int(now.timestamp() * 1000)
    base = now.strftime("%Y-%m-%dT%H:%M:%S")
    return f"{base}.{ms % 1000:03d}Z"


def _sl_build_hash(secret: str, timestamp: str, endpoint: str, body_obj: dict) -> str:
    body_json = json.dumps(body_obj, separators=(",", ":"), ensure_ascii=False)
    raw = timestamp + "POST" + endpoint + "" + body_json
    return hmac.new(secret.encode(), raw.encode(), hashlib.sha256).hexdigest()


def _sl_extract_token(resp: requests.Response):
    token = resp.cookies.get("token")
    if token:
        return token, "cookie"
    set_cookie = resp.headers.get("Set-Cookie", "")
    if set_cookie:
        m = re.search(r"token=([^;]+)", set_cookie)
        if m:
            return m.group(1), "set-cookie"
    try:
        data = resp.json()
        jwt = data.get("jwt") or data.get("token") or data.get("access_token")
        if jwt and isinstance(jwt, str) and jwt.startswith("eyJ"):
            return jwt, "json"
        text = json.dumps(data, ensure_ascii=False)
    except Exception:
        text = resp.text
    m = re.search(r"eyJ[A-Za-z0-9_-]+\.[A-Za-z0-9_-]+\.[A-Za-z0-9_-]+", text)
    if m:
        src = "json" if "application/json" in resp.headers.get("Content-Type", "") else "text"
        return m.group(0), src
    return "", "not_found"


@app.route("/admin/single-login")
@admin_required
def admin_single_login_page():
    return render_template("admin_single_login.html")


@app.route("/admin/single-login/token", methods=["POST"])
@admin_required
def admin_single_login_token():
    secret = os.environ.get("APP_HMAC_SECRET", "").strip()
    if not secret:
        return jsonify({"success": False, "error": "APP_HMAC_SECRET belum di-set.", "status_code": None,
                        "token": "", "token_source": "", "timestamp": "", "hash": "", "response_preview": ""})

    data = request.get_json(silent=True) or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    type_of_web = (data.get("type_of_web") or _SL_DEFAULT_TYPE_OF_WEB).strip()
    recaptcha_token = (data.get("recaptcha_token") or "").strip()

    if not username:
        return jsonify({"success": False, "error": "Username kosong.", "status_code": None,
                        "token": "", "token_source": "", "timestamp": "", "hash": "", "response_preview": ""})
    if not password:
        return jsonify({"success": False, "error": "Password kosong.", "status_code": None,
                        "token": "", "token_source": "", "timestamp": "", "hash": "", "response_preview": ""})
    if not recaptcha_token:
        captcha_key = os.environ.get("TWOCAPTCHA_API_KEY", "").strip()
        if not captcha_key:
            return jsonify({"success": False, "error": "reCAPTCHA token kosong dan TWOCAPTCHA_API_KEY belum di-set.", "status_code": None,
                            "token": "", "token_source": "", "timestamp": "", "hash": "", "response_preview": ""})
        try:
            recaptcha_token = _sl_solve_recaptcha(captcha_key)
        except Exception as e:
            return jsonify({"success": False, "error": f"Auto-solve captcha gagal: {e}", "status_code": None,
                            "token": "", "token_source": "", "timestamp": "", "hash": "", "response_preview": ""})

    ts = _sl_utc_timestamp_ms()
    body_obj = {
        "identifier": username,
        "password": base64.b64encode(password.encode()).decode(),
        "type_of_web": type_of_web,
        "recaptcha_token": recaptcha_token,
    }
    hash_val = _sl_build_hash(secret, ts, _SL_AUTH_ENDPOINT, body_obj)
    headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
        "Content-Type": "application/json",
        "Dnt": "1",
        "Hash": hash_val,
        "Origin": "https://letscml.id",
        "Referer": "https://letscml.id/",
        "Timestamp": ts,
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36",
    }

    try:
        resp = requests.post(_SL_BASE_URL + _SL_AUTH_ENDPOINT, json=body_obj, headers=headers, timeout=45)
    except requests.exceptions.Timeout:
        return jsonify({"success": False, "error": "Timeout (45s).", "status_code": None,
                        "token": "", "token_source": "", "timestamp": ts, "hash": hash_val, "response_preview": ""})
    except Exception as e:
        return jsonify({"success": False, "error": str(e), "status_code": None,
                        "token": "", "token_source": "", "timestamp": ts, "hash": hash_val, "response_preview": ""})

    try:
        preview = json.dumps(resp.json(), ensure_ascii=False, indent=2)[:2000]
    except Exception:
        preview = resp.text[:2000]

    token, token_source = _sl_extract_token(resp)
    success = resp.status_code == 200 and bool(token)
    error_msg = "" if success else (f"Status {resp.status_code}." if not resp.ok else "Token tidak ditemukan.")

    return jsonify({"success": success, "status_code": resp.status_code, "token": token,
                    "token_source": token_source, "timestamp": ts, "hash": hash_val,
                    "response_preview": preview, "error": error_msg})


_SL_WEB_URL = "https://letscml.id/cml-team/cmkt/signin"
_SL_RECAPTCHA_SITE_KEY = "6Lc6hNMqAAAAAJQdVFyxrTTM0yiD54R_iiIOooZF"
_SL_BROWSER_TIMEOUT = 60


def _sl_solve_recaptcha(api_key: str) -> str:
    submit = requests.post("https://2captcha.com/in.php", data={
        "key": api_key,
        "method": "userrecaptcha",
        "googlekey": _SL_RECAPTCHA_SITE_KEY,
        "pageurl": _SL_WEB_URL,
        "json": 1,
    }, timeout=30).json()

    if submit.get("status") != 1:
        raise ValueError(f"2captcha submit error: {submit.get('request')}")

    task_id = submit["request"]

    for _ in range(24):
        time.sleep(5)
        res = requests.get("https://2captcha.com/res.php", params={
            "key": api_key,
            "action": "get",
            "id": task_id,
            "json": 1,
        }, timeout=15).json()
        if res.get("status") == 1:
            return res["request"]
        if res.get("request") != "CAPCHA_NOT_READY":
            raise ValueError(f"2captcha error: {res.get('request')}")

    raise ValueError("2captcha timeout: token tidak diterima dalam 2 menit.")


def _sl_get_url_origin(url: str) -> str:
    parsed = urlparse(url)
    return f"{parsed.scheme}://{parsed.netloc}" if parsed.scheme and parsed.netloc else ""


def _sl_capture_token_with_browser(username: str, password: str, headless: bool = False):
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        raise ValueError(
            "Playwright belum terinstall.\n"
            "Jalankan: pip install playwright && python -m playwright install chromium"
        )

    captured = {"token": "", "source": ""}

    def _remember(raw, source):
        if not raw or captured["token"]:
            return
        m = re.search(r"eyJ[A-Za-z0-9_-]+\.[A-Za-z0-9_-]+\.[A-Za-z0-9_-]+", str(raw))
        if m:
            captured["token"] = m.group(0)
            captured["source"] = source

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        try:
            context = browser.new_context(ignore_https_errors=True)
            origin = _sl_get_url_origin(_SL_WEB_URL)
            try:
                context.grant_permissions(["local-network-access"], origin=origin or None)
            except Exception:
                pass

            page = context.new_page()

            def on_response(resp):
                try:
                    url = resp.url
                    if "/api/set-token" in url:
                        from urllib.parse import parse_qs
                        params = parse_qs(urlparse(url).query)
                        _remember(params.get("token", [""])[0], "set-token url")
                    if any(p in url for p in ("/api/auth/local", "/api/get-token", "/api/set-token")):
                        _remember(resp.headers.get("set-cookie", ""), f"set-cookie {url}")
                        try:
                            _remember(resp.text(), f"body {url}")
                        except Exception:
                            pass
                except Exception:
                    pass

            page.on("response", on_response)
            page.goto(_SL_WEB_URL, wait_until="domcontentloaded", timeout=60000)

            visible_inputs = []
            for inp in page.locator("input").all():
                try:
                    if inp.is_visible(timeout=300) and inp.is_enabled(timeout=300):
                        t = (inp.get_attribute("type") or "").strip().lower()
                        if t not in ("hidden", "submit", "button", "checkbox", "radio"):
                            visible_inputs.append(inp)
                except Exception:
                    continue

            if len(visible_inputs) >= 2:
                page.wait_for_timeout(800)
                visible_inputs[0].fill(username)
                page.wait_for_timeout(800)
                visible_inputs[1].fill(password)
                page.wait_for_timeout(800)
            else:
                raise ValueError("Field username/password tidak ditemukan di halaman.")

            for sel in ['xpath=//*[@id="button"]', 'button:has-text("LOG IN")',
                        'button:has-text("Login")', 'button[type="submit"]']:
                try:
                    loc = page.locator(sel).first
                    loc.wait_for(state="visible", timeout=800)
                    loc.click(timeout=5000)
                    break
                except Exception:
                    continue

            deadline = time.time() + _SL_BROWSER_TIMEOUT
            while time.time() < deadline and not captured["token"]:
                for cookie in context.cookies():
                    if cookie.get("name") in ("tokenCML", "token"):
                        _remember(cookie.get("value", ""), f"cookie {cookie.get('name')}")
                if captured["token"]:
                    break
                page.wait_for_timeout(1000)

            if not captured["token"]:
                raise ValueError(
                    f"Token tidak tertangkap dalam {_SL_BROWSER_TIMEOUT} detik. "
                    "Pastikan login sukses dan reCAPTCHA sudah diselesaikan."
                )

            return captured["token"], captured["source"]
        finally:
            browser.close()


@app.route("/admin/single-login/browser-token", methods=["POST"])
@admin_required
def admin_browser_login_token():
    data = request.get_json(silent=True) or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    headless = bool(data.get("headless", False))

    if not username:
        return jsonify({"success": False, "error": "Username kosong.", "token": "", "token_source": ""})
    if not password:
        return jsonify({"success": False, "error": "Password kosong.", "token": "", "token_source": ""})

    try:
        token, source = _sl_capture_token_with_browser(username, password, headless=headless)
        return jsonify({"success": True, "token": token, "token_source": source, "error": ""})
    except Exception as e:
        return jsonify({"success": False, "error": str(e), "token": "", "token_source": ""})


# ── End Single Login ───────────────────────────────────────────────────────────

init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5000"))
    app.run(host="0.0.0.0", port=port, debug=not IS_PROD)
